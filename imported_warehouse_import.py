from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from pathlib import Path
import html
import json
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook

DATA_PATH = Path("data/imported_warehouse_model.json")
RESULTS_PATH = Path("results/imported_warehouse_cells.xlsx")

REQUIRED_CELL_FIELDS = {
    "raw_address": ["ссылка", "address", "адрес", "ячейка склада", "складская ячейка"],
    "pallet_capacity": ["количествопаллет", "количество паллет", "паллет", "pallet_capacity"],
    "warehouse_zone": ["складзоныячеекссылка", "зона", "склад", "warehouse_zone"],
    "disabled_raw": ["неиспользовать", "не использовать", "disabled"],
}

AREA_TYPES = [
    "складская зона",
    "ряд",
    "проход",
    "стена / недоступная зона",
    "ворота / старт маршрута",
    "служебная область",
    "игнорировать",
]


@dataclass
class ImportDiagnostics:
    errors: list[str]
    warnings: list[str]
    metrics: dict[str, Any]


def normalize_column_name(value: Any) -> str:
    return re.sub(r"[\s_\-]+", "", str(value).strip().lower().replace("ё", "е"))


def find_column(columns, aliases):
    normalized = {normalize_column_name(col): col for col in columns}
    for alias in aliases:
        found = normalized.get(normalize_column_name(alias))
        if found is not None:
            return found
    return None


def truthy_disabled(value: Any) -> bool:
    text = str(value).strip().lower().replace("ё", "е")
    if text in {"да", "yes", "true", "1", "y", "использовать нельзя", "не использовать"}:
        return True
    if text in {"нет", "no", "false", "0", "n", "", "nan", "none"}:
        return False
    return False


def read_cell_export(file_obj, *, cell_segment=1, row_segment=2, tier_segment=3, tier_mode="Только выбранный ярус", selected_tier="04", source_name="1c_export"):
    source = pd.read_excel(file_obj).dropna(how="all").copy()
    mapping = {field: find_column(source.columns, aliases) for field, aliases in REQUIRED_CELL_FIELDS.items()}
    missing = [field for field, col in mapping.items() if col is None]
    errors, warnings = [], []
    if missing:
        labels = ", ".join(missing)
        return source, pd.DataFrame(), ImportDiagnostics([f"Не найдены обязательные колонки: {labels}"], warnings, {})

    rows = []
    bad_capacity = 0
    empty_address = 0
    unparsed_row = 0
    for idx, src in source.iterrows():
        raw_address = str(src.get(mapping["raw_address"], "")).strip()
        if raw_address in {"", "nan", "None"}:
            empty_address += 1
            continue
        parts = [part.strip() for part in re.split(r"[-–—]", raw_address)]
        def seg(number):
            return parts[number - 1] if 1 <= int(number) <= len(parts) else ""
        cell_number = seg(cell_segment)
        row_number = seg(row_segment)
        tier = seg(tier_segment)
        if not row_number:
            unparsed_row += 1
        capacity = pd.to_numeric(pd.Series([src.get(mapping["pallet_capacity"])]), errors="coerce").iloc[0]
        if pd.isna(capacity):
            bad_capacity += 1
            capacity = 0
        rows.append({
            "warehouse_zone": str(src.get(mapping["warehouse_zone"], "")).strip(),
            "raw_address": raw_address,
            "row_number": row_number,
            "cell_number": cell_number,
            "tier": tier,
            "pallet_capacity": float(capacity),
            "disabled": truthy_disabled(src.get(mapping["disabled_raw"])),
            "source_file": source_name,
        })
    cells = pd.DataFrame(rows)
    duplicate_count = int(cells["raw_address"].duplicated().sum()) if not cells.empty else 0
    if duplicate_count:
        warnings.append(f"Найдены дубли адресов: {duplicate_count}. В модели оставлен первый экземпляр адреса.")
        cells = cells.drop_duplicates("raw_address", keep="first")

    if not cells.empty:
        if tier_mode == "Только выбранный ярус":
            cells = cells[cells["tier"].astype(str) == str(selected_tier)].copy()
        elif tier_mode == "Минимальный ярус по каждому ряду":
            min_tiers = cells.groupby("row_number")["tier"].transform(lambda s: sorted(s.astype(str))[0])
            cells = cells[cells["tier"].astype(str) == min_tiers.astype(str)].copy()

    metrics = {
        "всего строк в выгрузке": len(source),
        "всего уникальных адресов": int(cells["raw_address"].nunique()) if not cells.empty else 0,
        "всего зон": int(cells["warehouse_zone"].nunique()) if not cells.empty else 0,
        "всего рядов": int(cells["row_number"].nunique()) if not cells.empty else 0,
        "всего активных ячеек": int((~cells["disabled"]).sum()) if not cells.empty else 0,
        "всего заблокированных ячеек": int(cells["disabled"].sum()) if not cells.empty else 0,
        "ячеек без распознанного ряда": unparsed_row,
        "дубли адресов": duplicate_count,
        "пустые адреса": empty_address,
        "некорректные КоличествоПаллет": bad_capacity,
    }
    return source, cells.reset_index(drop=True), ImportDiagnostics(errors, warnings, metrics)


def _fill_hex(cell) -> str:
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return ""
    color = fill.fgColor
    if color.type == "rgb" and color.rgb and color.rgb not in {"00000000", "00FFFFFF"}:
        return "#" + color.rgb[-6:].upper()
    if color.type in {"indexed", "theme"}:
        return "#D9EAD3"
    return ""


def read_scheme_colors(file_obj):
    wb = load_workbook(file_obj, data_only=True)
    records = []
    merged_lookup = {}
    for ws in wb.worksheets:
        for rng in ws.merged_cells.ranges:
            top = ws.cell(rng.min_row, rng.min_col)
            for row in range(rng.min_row, rng.max_row + 1):
                for col in range(rng.min_col, rng.max_col + 1):
                    merged_lookup[(ws.title, row, col)] = top
        for row in ws.iter_rows():
            for cell in row:
                owner = merged_lookup.get((ws.title, cell.row, cell.column), cell)
                color = _fill_hex(owner)
                if not color:
                    continue
                value = "" if owner.value is None else str(owner.value).strip()
                records.append({"sheet": ws.title, "excel_row": cell.row, "excel_col": cell.column, "value": value, "fill_hex": color})
    points = pd.DataFrame(records)
    if points.empty:
        return points, pd.DataFrame(columns=["fill_hex", "excel_cells", "example_value", "area_type"]), []
    summary = (points.groupby("fill_hex", as_index=False)
        .agg(excel_cells=("fill_hex", "count"), example_value=("value", lambda s: next((x for x in s if x), ""))))
    summary["area_type"] = "ряд"
    return points, summary, []


def detect_row_labels(points: pd.DataFrame) -> dict[str, dict[str, Any]]:
    labels = {}
    if points.empty:
        return labels
    for color, group in points.groupby("fill_hex"):
        values = [str(v).strip() for v in group["value"].tolist() if str(v).strip()]
        nums = []
        for value in values:
            m = re.search(r"(?:ряд\s*)?(\d{1,4})", value.lower())
            if m:
                nums.append(m.group(1))
        if nums:
            labels[color] = {"row_number": Counter(nums).most_common(1)[0][0], "source": "подпись на схеме"}
    return labels


def build_imported_model(cells: pd.DataFrame, scheme_points: pd.DataFrame, color_types: dict[str, str], manual_rows: dict[str, str], *, cell_length_m=1.2, row_width_m=0.8, aisle_width_m=2.0, numbering_direction="по возрастанию", route_start="минимальная ячейка"):
    warnings = []
    row_labels = detect_row_labels(scheme_points)
    rows_by_color = {color: data["row_number"] for color, data in row_labels.items() if color_types.get(color) == "ряд"}
    rows_by_color.update({color: str(row).strip() for color, row in manual_rows.items() if str(row).strip()})

    scheme_rows = set(rows_by_color.values())
    export_rows = set(cells["row_number"].astype(str)) if not cells.empty else set()
    for row in sorted(scheme_rows - export_rows):
        warnings.append(f"На схеме есть ряд {row}, но в выгрузке по нему нет ячеек.")
    unmapped_rows = sorted(export_rows - scheme_rows)
    if unmapped_rows and not scheme_rows:
        warnings.append("Схема не дала привязок рядов; включён fallback: ряды разложены сеткой по выгрузке 1С.")
    elif unmapped_rows:
        warnings.append(f"В выгрузке есть ряды без сопоставления со схемой: {', '.join(unmapped_rows[:20])}.")

    scheme_geometry = {}
    if not scheme_points.empty:
        for color, row in rows_by_color.items():
            pts = scheme_points[scheme_points["fill_hex"] == color]
            if not pts.empty:
                scheme_geometry[str(row)] = {
                    "x0": float(pts["excel_col"].min()) * cell_length_m,
                    "y0": float(pts["excel_row"].min()) * (row_width_m + aisle_width_m),
                }

    output = []
    for row_idx, (row_number, group) in enumerate(cells.groupby("row_number", sort=True)):
        row_key = str(row_number)
        ordered = group.copy()
        ordered["_sort_cell"] = pd.to_numeric(ordered["cell_number"], errors="coerce")
        ordered = ordered.sort_values(["_sort_cell", "cell_number"], ascending=(numbering_direction == "по возрастанию")).reset_index(drop=True)
        base = scheme_geometry.get(row_key, {"x0": 0.0, "y0": row_idx * (row_width_m + aisle_width_m)})
        for pos, (_, rec) in enumerate(ordered.iterrows()):
            output.append({
                **{k: rec[k] for k in ["warehouse_zone", "raw_address", "row_number", "cell_number", "tier", "pallet_capacity", "disabled", "source_file"]},
                "x": round(float(base["x0"]) + pos * cell_length_m, 3),
                "y": round(float(base["y0"]), 3),
                "width_m": float(cell_length_m),
                "height_m": float(row_width_m),
                "matched_to_scheme": row_key in scheme_geometry,
            })
    model_cells = pd.DataFrame(output)
    warnings.extend([f"Не сопоставлено со схемой ячеек: {int((~model_cells['matched_to_scheme']).sum()) if not model_cells.empty else 0}."])
    return {
        "schema_version": 1,
        "source": "excel_import",
        "settings": {"cell_length_m": cell_length_m, "row_width_m": row_width_m, "aisle_width_m": aisle_width_m, "numbering_direction": numbering_direction, "route_start": route_start},
        "cells": model_cells.to_dict(orient="records"),
        "warnings": warnings,
    }, model_cells, warnings


def save_imported_model(model: dict[str, Any], cells: pd.DataFrame):
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    RESULTS_PATH.parent.mkdir(parents=True, exist_ok=True)
    DATA_PATH.write_text(json.dumps(model, ensure_ascii=False, indent=2), encoding="utf-8")
    with pd.ExcelWriter(RESULTS_PATH, engine="openpyxl") as writer:
        cells.to_excel(writer, sheet_name="imported_cells", index=False)
    return DATA_PATH, RESULTS_PATH


def build_imported_warehouse_html(cells: pd.DataFrame, scale=55):
    if cells.empty:
        return "<p>Нет ячеек для отображения.</p>"
    cells = cells.copy()
    min_x, min_y = cells["x"].min(), cells["y"].min()
    cells["vx"] = (cells["x"] - min_x) * scale + 80
    cells["vy"] = (cells["y"] - min_y) * scale + 70
    width = int(cells["vx"].max() + 220)
    height = int(cells["vy"].max() + 160)
    elements = []
    for _, c in cells.iterrows():
        color = "#bdbdbd" if bool(c["disabled"]) else "#88c999"
        title = html.escape("; ".join([
            f"адрес={c['raw_address']}", f"ряд={c['row_number']}", f"ячейка={c['cell_number']}", f"ярус={c['tier']}",
            f"вместимость={c['pallet_capacity']}", f"зона={c['warehouse_zone']}", f"доступна={not bool(c['disabled'])}",
        ]))
        label = html.escape(str(c["cell_number"]))
        elements.append(f"<rect x='{c['vx']:.1f}' y='{c['vy']:.1f}' width='{max(16, c['width_m']*scale):.1f}' height='{max(16, c['height_m']*scale):.1f}' fill='{color}' stroke='#333'><title>{title}</title></rect><text x='{c['vx']+4:.1f}' y='{c['vy']+14:.1f}' font-size='10'>{label}</text>")
    return f"<div style='height:720px; overflow:auto; border:1px solid #aaa'><svg width='{width}' height='{height}'>{''.join(elements)}</svg></div>"
