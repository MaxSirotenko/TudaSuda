from io import BytesIO
import re
from openpyxl import load_workbook
from warehouse_addressing import FIRST_TIER
from warehouse_model import WarehouseCell, WarehouseModel, WarehouseRow, WarehouseSheet

ROW_LABEL_RE = re.compile(r"(?:^|\b)(?:\u0440\u044f\u0434\s*)?(\d{1,4}|[A-Za-z\u0410-\u042f\u0430-\u044f\u0401\u0451]{1,3}\d{0,3})(?:\b|$)", re.IGNORECASE)


MOJIBAKE_MARKERS = ("Р", "С", "РЃ", "С‘")


def _repair_mojibake(text: str) -> str:
    if not any(marker in text for marker in MOJIBAKE_MARKERS):
        return text
    try:
        repaired = text.encode("cp1251").decode("utf-8")
    except UnicodeError:
        return text
    return repaired if repaired else text


def _text(value) -> str:
    if value is None:
        return ""
    return _repair_mojibake(str(value)).strip()


def _fill_color(cell) -> str:
    fill = cell.fill
    if fill is None or fill.fill_type is None:
        return ""
    color = fill.fgColor
    if color is None:
        return ""
    if color.type == "rgb" and color.rgb and color.rgb not in {"00000000", "00FFFFFF"}:
        rgb = color.rgb[-6:]
        return f"#{rgb}"
    if color.type == "indexed" and color.indexed not in {None, 64}:
        return "#d9ead3"
    if color.type == "theme":
        return "#d9ead3"
    return ""


def _row_number(label: str) -> str:
    match = ROW_LABEL_RE.search(label.replace("\u2116", ""))
    return match.group(1) if match else label.strip()


def _looks_like_row_label(text: str) -> bool:
    clean = text.strip()
    if not clean or len(clean) > 30:
        return False
    low = clean.lower()
    return "\u0440\u044f\u0434" in low or bool(re.fullmatch(r"\d{1,4}|[A-Za-z\u0410-\u042f\u0430-\u044f\u0401\u0451]{1,3}\d{0,3}", clean))


def _find_extent(ws, r: int, c: int) -> tuple[int, int, int, int, str, float, list[str]]:
    warnings = []
    right = c + 1
    while right <= ws.max_column and _text(ws.cell(r, right).value) == "":
        right += 1
    down = r + 1
    while down <= ws.max_row and _text(ws.cell(down, c).value) == "":
        down += 1
    horizontal_span = max(1, min(20, (right - c - 1) if right <= ws.max_column else 8))
    vertical_span = max(1, min(20, (down - r - 1) if down <= ws.max_row else 8))
    if horizontal_span >= vertical_span:
        direction = "left_to_right"
        min_row, max_row = r, r
        min_col, max_col = c + 1, c + horizontal_span
    else:
        direction = "top_to_bottom"
        min_row, max_row = r + 1, r + vertical_span
        min_col, max_col = c, c
    confidence = 0.75 if max(horizontal_span, vertical_span) >= 3 else 0.45
    if confidence < 0.6:
        warnings.append("Ряд найден по подписи, но область ячеек рядом с подписью определена сомнительно.")
    return min_row, min_col, max_row, max_col, direction, confidence, warnings


def parse_warehouse_excel(file_obj, sheet_names: list[str] | None = None) -> WarehouseModel:
    if isinstance(file_obj, bytes):
        file_obj = BytesIO(file_obj)
    wb = load_workbook(file_obj, data_only=True)
    model = WarehouseModel(sheets=[])
    selected_names = set(sheet_names or wb.sheetnames)
    for ws in wb.worksheets:
        if ws.title not in selected_names:
            continue
        values = []
        labels = []
        painted_by_excel_row: dict[int, list[tuple]] = {}
        min_row, min_col, max_row, max_col = 1, 1, ws.max_row, ws.max_column
        if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None and not _fill_color(ws.cell(1, 1)):
            sheet = WarehouseSheet(name=ws.title, max_row=1, max_column=1, values=[], merged_ranges=[])
            sheet.warnings.append("Лист пустой и пропущен без детальной обработки.")
            model.sheets.append(sheet)
            continue
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            if not any(cell.value is not None or cell.fill.fill_type is not None for cell in row):
                continue
            for cell in row:
                text = _text(cell.value)
                fill_color = _fill_color(cell)
                if text:
                    values.append({"row": cell.row, "column": cell.column, "value": text})
                    if _looks_like_row_label(text):
                        labels.append((cell.row, cell.column, text))
                if fill_color:
                    painted_by_excel_row.setdefault(cell.row, []).append((cell, text, fill_color))
        sheet = WarehouseSheet(
            name=ws.title,
            max_row=ws.max_row,
            max_column=ws.max_column,
            values=values,
            merged_ranges=[str(rng) for rng in ws.merged_cells.ranges],
        )

        if painted_by_excel_row:
            for excel_row, painted_cells in sorted(painted_by_excel_row.items()):
                painted_cells = sorted(painted_cells, key=lambda item: item[0].column)
                row_label = ""
                for col in range(max(1, painted_cells[0][0].column - 3), painted_cells[0][0].column):
                    candidate = _text(ws.cell(excel_row, col).value)
                    if candidate:
                        row_label = _row_number(candidate) if _looks_like_row_label(candidate) else candidate
                row_number = row_label or str(excel_row)
                wh_row = WarehouseRow(
                    ws.title,
                    row_number,
                    excel_row,
                    painted_cells[0][0].column,
                    excel_row,
                    painted_cells[-1][0].column,
                    "left_to_right",
                    0.9,
                )
                for idx, (cell, text, fill_color) in enumerate(painted_cells, start=1):
                    cell_number = str(idx)
                    wh_row.potential_cells.append(
                        WarehouseCell(
                            ws.title,
                            row_number,
                            cell_number,
                            FIRST_TIER,
                            f"{cell_number}-{row_number}-{FIRST_TIER}",
                            cell.column,
                            excel_row,
                            fill_color=fill_color,
                            value=text,
                            source="excel_fill",
                        )
                    )
                sheet.rows.append(wh_row)
            sheet.warnings.append("Ячейки построены по заливкам Excel; табличные колонки row_number/pallet_count не требуются.")
        else:
            seen = set()
            for r, c, label in labels:
                row_number = _row_number(label)
                if (row_number, r, c) in seen:
                    continue
                seen.add((row_number, r, c))
                min_row, min_col, max_row, max_col, direction, confidence, warnings = _find_extent(ws, r, c)
                wh_row = WarehouseRow(ws.title, row_number, min_row, min_col, max_row, max_col, direction, confidence, warnings=warnings)
                if confidence >= 0.6:
                    count = (max_col - min_col + 1) if direction == "left_to_right" else (max_row - min_row + 1)
                    for idx in range(1, count + 1):
                        x = min_col + idx - 1 if direction == "left_to_right" else min_col
                        y = min_row if direction == "left_to_right" else min_row + idx - 1
                        wh_row.potential_cells.append(WarehouseCell(ws.title, row_number, str(idx), FIRST_TIER, f"{idx}-{row_number}-{FIRST_TIER}", x, y))
                else:
                    sheet.warnings.append(f"Ряд '{label}' на {r}:{c} не получил автоматические ячейки из-за низкой уверенности.")
                sheet.rows.append(wh_row)
            if not sheet.rows:
                sheet.warnings.append("На листе не найдены цветные ячейки или уверенные текстовые подписи рядов.")
        model.sheets.append(sheet)
    wb.close()
    return model