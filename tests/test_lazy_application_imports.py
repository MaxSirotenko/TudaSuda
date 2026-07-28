from __future__ import annotations

import ast
import hashlib
import json
from pathlib import Path
import subprocess
import sys

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _python(code: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run([sys.executable, "-c", code], cwd=ROOT, capture_output=True, text=True)


def test_plain_application_import_defers_legacy_dependencies_and_openpyxl():
    names = ("warehouse_legacy_excel_ui", "warehouse_excel_parser", "warehouse_visualization",
             "warehouse_placement", "warehouse_diagnostics", "openpyxl")
    process = _python("import json,sys,virtual_warehouse_app; print(json.dumps({n:n in sys.modules for n in %r}))" % (names,))
    assert process.returncode == 0, process.stderr
    assert json.loads(process.stdout.splitlines()[-1]) == {name: False for name in names}


def test_legacy_module_is_one_way_and_import_has_no_persisted_side_effects():
    paths = tuple(ROOT / "data/last_import" / name for name in (
        "warehouse_model.json", "placements.json", "receipts.json", "outbound_orders.json",
        "outbound_execution_state.json", "outbound_execution_log.json", "data_revisions.json",
        "placement_diagnostics.json",
    ))
    before = {p: (p.exists(), hashlib.sha256(p.read_bytes()).hexdigest() if p.exists() else "") for p in paths}
    process = _python("import sys,warehouse_legacy_excel_ui; print('virtual_warehouse_app' in sys.modules)")
    after = {p: (p.exists(), hashlib.sha256(p.read_bytes()).hexdigest() if p.exists() else "") for p in paths}
    assert process.returncode == 0, process.stderr
    assert process.stdout.splitlines()[-1] == "False"
    assert before == after


def test_legacy_module_has_no_top_level_ui_or_application_import():
    source = (ROOT / "warehouse_legacy_excel_ui.py").read_text(encoding="utf-8")
    tree = ast.parse(source)
    assert "virtual_warehouse_app" not in source
    assert "set_page_config" not in source
    assert not any(isinstance(node, ast.If) and isinstance(node.test, ast.Compare) for node in tree.body)
    assert not any(isinstance(node, ast.Expr) and isinstance(node.value, ast.Call) for node in tree.body)


def test_excel_sheet_loader_is_lazy_and_closes_workbook(tmp_path):
    import warehouse_legacy_excel_ui as legacy
    from openpyxl import Workbook

    path = tmp_path / "sheets.xlsx"
    workbook = Workbook()
    workbook.active.title = "Первый"
    workbook.create_sheet("Второй")
    workbook.save(path)
    assert legacy.get_excel_sheet_names(path.read_bytes(), "unique-test-hash") == ["Первый", "Второй"]
    path.unlink()  # Windows also proves the read-only workbook was closed.


def test_dispatcher_preserves_mode_labels_default_and_lazy_import():
    source = (ROOT / "virtual_warehouse_app.py").read_text(encoding="utf-8")
    assert '["Склад из Excel: ряды + ячейки + проезды", "Виртуальный склад по Excel-схеме"]' in source
    assert "index=0" in source
    assert "from warehouse_legacy_excel_ui import render_legacy_excel_warehouse" in source
    assert "from openpyxl import load_workbook" not in source


def test_apptest_default_and_legacy_modes_open_without_exception():
    from streamlit.testing.v1 import AppTest

    app = AppTest.from_file(str(ROOT / "virtual_warehouse_app.py"), default_timeout=20).run()
    assert not app.exception
    mode = next(radio for radio in app.radio if radio.label == "Режим")
    assert mode.value == "Склад из Excel: ряды + ячейки + проезды"
    mode.set_value("Виртуальный склад по Excel-схеме").run()
    assert not app.exception
    assert any("Загрузите Excel-схему" in item.value for item in app.info)
    assert "warehouse_legacy_excel_ui" in sys.modules
