from __future__ import annotations

import copy
import hashlib
import json
import shutil
import subprocess
from dataclasses import asdict
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from time import perf_counter

import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
from streamlit.runtime.scriptrunner import get_script_run_ctx
from openpyxl import load_workbook

from warehouse_diagnostics import build_diagnostics
from warehouse_excel_parser import parse_warehouse_excel
from warehouse_model import WarehouseCell, WarehouseModel, WarehouseRow, WarehouseSheet
from warehouse_placement import apply_cell_addresses, apply_placements, import_cell_addresses, import_placements
from warehouse_visualization import build_virtual_warehouse_html, prepare_render_cache
from warehouse_performance import (
    clear_performance_history,
    finish_performance_run,
    is_performance_enabled,
    load_latest_performance_run,
    measure_step,
    start_performance_run,
)
from warehouse_revisions import (
    REVISION_DOMAINS,
    bump_revisions,
    load_revision_state,
    resolve_model_id,
)
from warehouse_geometry_model import (
    GeometrySettings,
    append_manual_change,
    apply_manual_overrides,
    build_geometry_html,
    build_geometry_model,
    cell_key,
    clear_manual_overrides,
    clear_row_settings,
    default_row_config,
    detect_column_mapping,
    empty_aisle_config,
    export_current_model_excel_bytes,
    get_excel_sheet_names as get_geometry_sheet_names,
    load_geometry_model,
    load_manual_overrides,
    manual_change_counts,
    normalize_cell_table,
    read_cell_table,
    rebuild_geometry_from_cells,
    save_geometry_model,
)

from warehouse_inventory_placement import (
    attach_placements_to_model,
    auto_place_unplaced,
    calculate_basic_weight_placement,
    clear_calculated_placements,
    clear_placement_state,
    delete_placement,
    detect_inventory_columns,
    export_placements_excel_bytes,
    get_inventory_sheet_names,
    import_inventory,
    load_placement_state,
    manual_place,
    normalize_inventory_table,
    placement_diagnostics,
    read_inventory_table,
    reconcile_placements_with_inventory,
    save_placement_state,
    update_placement_qty,
    move_placement,
)
from warehouse_placement_diagnostics import (
    PLACEMENT_CATEGORY_COLORS,
    ZONE_LABELS_RU,
    build_placement_diagnostics,
    enrich_model_with_placement_diagnostics,
    load_pre_placement_snapshot,
    save_pre_placement_snapshot,
)

from warehouse_receipts import (
    build_receipt_diagnostics,
    calculate_receipt_zones,
    clear_receipts_state,
    default_zone_classification_settings,
    detect_receipt_columns,
    detect_zone_classification_columns,
    export_receipts_excel_bytes,
    get_receipt_sheet_names,
    load_receipts_state,
    make_receipts_state,
    normalize_receipt_table,
    read_receipt_table,
    save_receipts_state,
    zone_classification_settings_hash,
)
from warehouse_row_settings import (
    apply_row_settings_transaction,
    changed_row_numbers,
    create_row_settings_state,
    reset_row_settings_state,
    update_row_settings_state,
)
from warehouse_cross_aisles import (
    apply_cross_aisles_transaction,
    changed_cross_aisle_count,
    create_cross_aisle_settings_state,
    reset_cross_aisle_settings_state,
    update_cross_aisle_settings_state,
)
from warehouse_zone_boundaries import (
    ZONE_LABELS,
    ZONE_ORDER,
    apply_active_boundaries_to_model,
    calculate_dynamic_zone_boundaries,
    ensure_zone_boundary_settings,
    set_base_boundaries_from_current_rows,
)
from warehouse_outbound_orders import (
    detect_outbound_columns,
    ensure_pre_outbound_snapshot,
    enrich_model_with_outbound_diagnostics,
    execute_outbound_orders,
    get_outbound_sheet_names,
    load_outbound_execution_log,
    load_outbound_execution_state,
    load_outbound_orders,
    normalize_outbound_table,
    outbound_execution_summary,
    read_outbound_table,
    reset_outbound_execution,
    save_outbound_execution_log,
    save_outbound_execution_state,
    save_outbound_orders,
    summarize_outbound_orders,
)
st.set_page_config(page_title="Симулятор сборки", layout="wide")

APP_BUILD_LABEL = "virtual-excel-only-2026-07-04"
MODEL_VERSION = 1
LAST_IMPORT_DIR = Path("data/last_import")
MODEL_PATH = LAST_IMPORT_DIR / "warehouse_model.json"
RENDER_CACHE_PATH = LAST_IMPORT_DIR / "render_cache.json"
META_PATH = LAST_IMPORT_DIR / "import_meta.json"
RENDER_SETTINGS_PATH = LAST_IMPORT_DIR / "render_settings.json"

DEFAULT_RENDER_LABEL_SETTINGS = {
    "show_row_labels": True,
    "show_cell_labels": True,
    "show_occupancy_labels": True,
    "show_aisle_labels": True,
    "label_mode": "Авто",
    "row_label_position": "авто",
}

DEFAULT_RENDER_COLOR_SETTINGS = {
    "cell_color": "#DCEBFF",
    "deep_lane_cell_color": "#CFE8D5",
    "aisle_color": "#F2F2F2",
    "cross_aisle_color": "#DCE6F2",
    "top_road_color": "#FFE8A3",
    "bottom_road_color": "#FFE8A3",
    "exit_color": "#FFCC80",
    "selected_cell_color": "#FF7043",
    "hover_cell_color": "#FFF59D",
    "occupied_cell_color": "#90CAF9",
    "deep_lane_partial_color": "#A5D6A7",
    "deep_lane_full_color": "#66BB6A",
}


def update_data_revisions(model: dict | object | None, domains: list[str], reason: str) -> bool:
    """Record a completed business operation, reporting persistence failures to UI."""
    try:
        with measure_step("revision_update", metadata={"domains": domains, "reason": reason}):
            bump_revisions(resolve_model_id(model), domains, reason=reason)
        return True
    except OSError as exc:
        st.error(
            "Данные сохранены, но ревизия кеша не обновилась. Исправьте доступ к "
            f"data_revisions.json и повторите операцию или перезапустите приложение: {exc}"
        )
        return False


def render_data_revisions(model: dict | object | None) -> None:
    """Render side-effect-free revision diagnostics in the service section."""
    state = load_revision_state(resolve_model_id(model))
    st.subheader("Ревизии данных")
    if state.get("warning"):
        st.warning(state["warning"])
    st.dataframe(
        pd.DataFrame(
            [{"Домен": domain, "Ревизия": state["revisions"][domain]} for domain in REVISION_DOMAINS]
        ),
        use_container_width=True,
        hide_index=True,
    )
    change = state.get("last_change", {})
    st.caption(f"model_id: {state['model_id']}")
    st.caption(f"Последнее изменение: {state.get('updated_at') or '—'}")
    st.caption(f"Причина: {change.get('reason') or '—'}")
    st.caption(f"Домены последней операции: {', '.join(change.get('domains', [])) or '—'}")

WEIGHT_ZONE_LABELS = {"heavy": "Тяжёлое", "medium": "Среднее", "light": "Лёгкое", "fragile": "Хрупкое", "unassigned": "Не назначено"}
WEIGHT_ZONE_VALUES = list(WEIGHT_ZONE_LABELS)
WEIGHT_ZONE_LABEL_TO_VALUE = {label: value for value, label in WEIGHT_ZONE_LABELS.items()}
STORAGE_TYPE_LABELS = {"normal": "Обычная", "deep_lane": "Набивная"}
STORAGE_TYPE_VALUES = list(STORAGE_TYPE_LABELS)
STORAGE_TYPE_LABEL_TO_VALUE = {label: value for value, label in STORAGE_TYPE_LABELS.items()}
ROW_STORAGE_TYPE_LABELS = {"normal": "Обычный ряд", "deep_lane": "Набивной ряд"}
DIRECTION_LABELS = {"bottom_to_top": "Снизу вверх", "top_to_bottom": "Сверху вниз"}
DIRECTION_VALUES = list(DIRECTION_LABELS)
DIRECTION_LABEL_TO_VALUE = {label: value for value, label in DIRECTION_LABELS.items()}
CELL_STATE_LABELS = {True: "Активна", False: "Заблокирована"}

def display_label(mapping: dict, value, default: str = "—") -> str:
    return mapping.get(str(value), default if value in (None, "") else str(value))

def select_internal(label: str, mapping: dict[str, str], current: str, *, key: str, container=st):
    values = list(mapping)
    index = values.index(current) if current in values else 0
    return container.selectbox(label, values, index=index, format_func=lambda value: mapping.get(value, str(value)), key=key)


@st.cache_data(show_spinner=False)
def get_git_release_info() -> dict[str, str]:
    repo_dir = Path(__file__).resolve().parent

    def git_text(*args: str) -> str:
        try:
            result = subprocess.run(
                ["git", *args],
                cwd=repo_dir,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=2,
                check=False,
            )
        except (OSError, subprocess.SubprocessError):
            return ""
        if result.returncode != 0:
            return ""
        return result.stdout.strip()

    merge_title = git_text("log", "--merges", "-1", "--pretty=%s")
    commit_title = git_text("log", "-1", "--pretty=%s")
    commit_hash = git_text("rev-parse", "--short", "HEAD")
    commit_date = git_text("log", "-1", "--date=short", "--pretty=%cd")
    return {
        "merge_title": merge_title,
        "commit_title": commit_title,
        "display_label": "Последний merge" if merge_title else "Последний commit",
        "display_title": merge_title or commit_title or "нет данных Git",
        "commit_hash": commit_hash or "—",
        "commit_date": commit_date or "—",
    }


def render_git_release_badge() -> None:
    info = get_git_release_info()
    st.sidebar.caption(f"{info['display_label']}: {info['display_title']}")
    st.sidebar.caption(f"Git commit: {info['commit_hash']} · {info['commit_date']}")


def file_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


@st.cache_data(show_spinner=False)
def get_excel_sheet_names(file_bytes: bytes, _content_hash: str) -> list[str]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


@st.cache_data(show_spinner=False)
def parse_warehouse_excel_cached(file_bytes: bytes, _content_hash: str, sheet_names: tuple[str, ...] | None):
    return parse_warehouse_excel(file_bytes, sheet_names=list(sheet_names) if sheet_names else None)


@st.cache_data(show_spinner=False)
def prepare_render_cache_cached(model_payload: dict) -> dict:
    return prepare_render_cache(model_from_dict(model_payload))


@st.cache_data(show_spinner=False)
def build_virtual_warehouse_html_cached(sheet_payload: dict, scale: int, summary_mode: bool):
    return build_virtual_warehouse_html(sheet_from_dict(sheet_payload), scale=scale, summary_mode=summary_mode)


@st.cache_data(show_spinner=False)
def read_cell_table_cached(file_bytes: bytes, content_hash: str, sheet_name: str, header_rows: int) -> pd.DataFrame:
    return read_cell_table(file_bytes, sheet_name, header_rows=header_rows)


@st.cache_data(show_spinner=False)
def normalize_cell_table_cached(table_payload: str, mapping_payload: str) -> tuple[pd.DataFrame, list[dict]]:
    df = pd.read_json(table_payload, orient="split")
    mapping = json.loads(mapping_payload)
    return normalize_cell_table(df, mapping)


@st.cache_data(show_spinner=False)
def build_geometry_html_cached(model_payload: str, scale: float, detailed: bool, label_settings_payload: str) -> str:
    return build_geometry_html(json.loads(model_payload), scale=scale, detailed=detailed, label_settings=json.loads(label_settings_payload))


def invalidate_geometry_render_cache() -> None:
    """Invalidate only artifacts whose contents are derived from geometry."""
    RENDER_CACHE_PATH.unlink(missing_ok=True)
    build_geometry_html_cached.clear()
    prepare_render_cache_cached.clear()


@st.cache_data(show_spinner=False)
def read_inventory_table_cached(file_bytes: bytes, content_hash: str, sheet_name: str, header_rows: int) -> pd.DataFrame:
    return read_inventory_table(file_bytes, sheet_name, header_rows=header_rows)


@st.cache_data(show_spinner=False)
def normalize_inventory_table_cached(table_payload: str, mapping_payload: str) -> tuple[pd.DataFrame, list[dict]]:
    df = pd.read_json(table_payload, orient="split")
    mapping = json.loads(mapping_payload)
    return normalize_inventory_table(df, mapping)


@st.cache_data(show_spinner=False)
def read_receipt_table_cached(file_bytes: bytes, content_hash: str, sheet_name: str, header_rows: int) -> pd.DataFrame:
    return read_receipt_table(file_bytes, sheet_name, header_rows=header_rows)


@st.cache_data(show_spinner=False)
def normalize_receipt_table_cached(table_payload: str, mapping_payload: str) -> tuple[pd.DataFrame, dict, list[dict]]:
    df = pd.read_json(table_payload, orient="split")
    mapping = json.loads(mapping_payload)
    return normalize_receipt_table(df, mapping)


def model_to_dict(model: WarehouseModel) -> dict:
    payload = asdict(model)
    payload["model_version"] = MODEL_VERSION
    return payload


def cell_from_dict(data: dict) -> WarehouseCell:
    return WarehouseCell(**{k: v for k, v in data.items() if k in WarehouseCell.__dataclass_fields__})


def row_from_dict(data: dict) -> WarehouseRow:
    row_data = {k: v for k, v in data.items() if k in WarehouseRow.__dataclass_fields__ and k != "potential_cells"}
    row = WarehouseRow(**row_data)
    row.potential_cells = [cell_from_dict(item) for item in data.get("potential_cells", [])]
    return row


def sheet_from_dict(data: dict) -> WarehouseSheet:
    sheet_data = {k: v for k, v in data.items() if k in WarehouseSheet.__dataclass_fields__ and k != "rows"}
    sheet = WarehouseSheet(**sheet_data)
    sheet.rows = [row_from_dict(item) for item in data.get("rows", [])]
    return sheet


def model_from_dict(data: dict) -> WarehouseModel:
    version = data.get("model_version")
    if version != MODEL_VERSION:
        raise ValueError(f"Неподдерживаемая версия модели: {version}; ожидается {MODEL_VERSION}.")
    model = WarehouseModel(sheets=[sheet_from_dict(item) for item in data.get("sheets", [])])
    model.diagnostics = data.get("diagnostics", [])
    return model


def write_json_atomic(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp_path.replace(path)


def load_render_settings() -> dict:
    settings = dict(DEFAULT_RENDER_LABEL_SETTINGS)
    settings["colors"] = dict(DEFAULT_RENDER_COLOR_SETTINGS)
    if RENDER_SETTINGS_PATH.exists():
        try:
            payload = json.loads(RENDER_SETTINGS_PATH.read_text(encoding="utf-8-sig"))
            settings.update({key: payload.get(key, value) for key, value in DEFAULT_RENDER_LABEL_SETTINGS.items()})
            colors = dict(DEFAULT_RENDER_COLOR_SETTINGS)
            colors.update(payload.get("colors", {}))
            settings["colors"] = colors
        except json.JSONDecodeError:
            pass
    return settings


def save_render_settings(settings: dict) -> None:
    payload = {}
    if RENDER_SETTINGS_PATH.exists():
        try:
            payload = json.loads(RENDER_SETTINGS_PATH.read_text(encoding="utf-8-sig"))
        except json.JSONDecodeError:
            payload = {}
    payload.update({key: settings.get(key, value) for key, value in DEFAULT_RENDER_LABEL_SETTINGS.items()})
    existing_colors = payload.get("colors", {}) if isinstance(payload.get("colors"), dict) else {}
    colors = dict(existing_colors)
    colors.update({key: settings.get("colors", {}).get(key, value) for key, value in DEFAULT_RENDER_COLOR_SETTINGS.items()})
    payload["colors"] = colors
    write_json_atomic(RENDER_SETTINGS_PATH, payload)


def render_label_settings_editor(settings: dict) -> dict:
    with st.expander("Настройки подписей", expanded=False):
        st.caption("Выберите, какие подписи показывать на карте. После сохранения настройки применяются к текущему рендеру без перестроения склада.")
        c1, c2, c3, c4 = st.columns(4)
        settings["show_row_labels"] = c1.checkbox("Показывать номера рядов", value=bool(settings.get("show_row_labels", True)), key="show_row_labels")
        settings["show_cell_labels"] = c2.checkbox("Показывать номера ячеек", value=bool(settings.get("show_cell_labels", True)), key="show_cell_labels")
        settings["show_occupancy_labels"] = c3.checkbox("Показывать занятость ячеек", value=bool(settings.get("show_occupancy_labels", True)), key="show_occupancy_labels")
        settings["show_aisle_labels"] = c4.checkbox("Показывать подписи проездов", value=bool(settings.get("show_aisle_labels", True)), key="show_aisle_labels")
        c5, c6 = st.columns(2)
        label_modes = ["Авто", "Полные", "Короткие", "Только при наведении"]
        row_positions = ["авто", "сверху", "снизу", "сверху и снизу"]
        settings["label_mode"] = c5.selectbox("Режим подписей", label_modes, index=label_modes.index(settings.get("label_mode", "Авто")) if settings.get("label_mode") in label_modes else 0, key="label_mode")
        settings["row_label_position"] = c6.selectbox("Положение номера ряда", row_positions, index=row_positions.index(settings.get("row_label_position", "авто")) if settings.get("row_label_position") in row_positions else 0, key="row_label_position")
        st.caption("Если подпись не помещается, карта уменьшает шрифт до 4 px, затем скрывает текст на карте, но оставляет полную информацию в tooltip.")
    return settings


def render_color_settings_editor(settings: dict) -> dict:
    colors = dict(DEFAULT_RENDER_COLOR_SETTINGS)
    colors.update(settings.get("colors", {}))
    with st.expander("Настройки цветов карты", expanded=False):
        st.caption("Настройте цвета ячеек, проездов и состояний занятости. Сохранение обновляет только цвета и не меняет геометрию склада.")
        c1, c2, c3 = st.columns(3)
        colors["cell_color"] = c1.color_picker("Цвет обычных ячеек", colors["cell_color"], key="color_cell")
        colors["deep_lane_cell_color"] = c2.color_picker("Цвет набивных ячеек", colors["deep_lane_cell_color"], key="color_deep_lane")
        colors["aisle_color"] = c3.color_picker("Цвет проездов между рядами", colors["aisle_color"], key="color_aisle")
        colors["cross_aisle_color"] = c3.color_picker("Цвет поперечных проездов", colors["cross_aisle_color"], key="color_cross_aisle")
        c4, c5, c6 = st.columns(3)
        colors["top_road_color"] = c4.color_picker("Цвет верхнего проезда", colors["top_road_color"], key="color_top_road")
        colors["bottom_road_color"] = c5.color_picker("Цвет нижнего проезда", colors["bottom_road_color"], key="color_bottom_road")
        colors["exit_color"] = c6.color_picker("Цвет выходов", colors["exit_color"], key="color_exit")
        c7, c8, c9 = st.columns(3)
        colors["selected_cell_color"] = c7.color_picker("Цвет выбранной ячейки", colors["selected_cell_color"], key="color_selected")
        colors["hover_cell_color"] = c8.color_picker("Цвет ячейки при наведении", colors["hover_cell_color"], key="color_hover")
        colors["occupied_cell_color"] = c9.color_picker("Цвет занятой ячейки", colors["occupied_cell_color"], key="color_occupied")
        c10, c11 = st.columns(2)
        colors["deep_lane_partial_color"] = c10.color_picker("Цвет частично занятой набивной", colors["deep_lane_partial_color"], key="color_deep_partial")
        colors["deep_lane_full_color"] = c11.color_picker("Цвет полностью занятой набивной", colors["deep_lane_full_color"], key="color_deep_full")
        b1, b2 = st.columns(2)
        if b1.button("Сохранить цвета", key="save_render_colors"):
            settings["colors"] = colors
            st.success("Цвета карты сохранены.")
        if b2.button("Сбросить цвета по умолчанию", key="reset_render_colors"):
            colors = dict(DEFAULT_RENDER_COLOR_SETTINGS)
            settings["colors"] = colors
            st.success("Цвета сброшены по умолчанию.")
    settings["colors"] = colors
    return settings


def render_map_settings_editor() -> dict:
    settings = load_render_settings()
    saved_settings = copy.deepcopy(settings)
    settings = render_label_settings_editor(settings)
    settings = render_color_settings_editor(settings)
    if settings != saved_settings:
        save_render_settings(settings)
        update_data_revisions(st.session_state.get("geometry_model"), ["render_settings"], "save_render_settings")
    return settings


def save_uploaded_copy(uploaded_file, target_name: str) -> str:
    if uploaded_file is None:
        return ""
    LAST_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    target = LAST_IMPORT_DIR / target_name
    target.write_bytes(uploaded_file.getvalue())
    return str(target)


def save_last_import(model: WarehouseModel, render_cache: dict, meta: dict) -> float:
    started = perf_counter()
    write_json_atomic(MODEL_PATH, model_to_dict(model))
    write_json_atomic(RENDER_CACHE_PATH, render_cache)
    write_json_atomic(META_PATH, meta)
    return perf_counter() - started


def load_last_import() -> tuple[WarehouseModel | None, dict, dict, str | None]:
    if not MODEL_PATH.exists():
        return None, {}, {}, None
    try:
        model_payload = json.loads(MODEL_PATH.read_text(encoding="utf-8-sig"))
        model = model_from_dict(model_payload)
        render_cache = json.loads(RENDER_CACHE_PATH.read_text(encoding="utf-8-sig")) if RENDER_CACHE_PATH.exists() else prepare_render_cache(model)
        meta = json.loads(META_PATH.read_text(encoding="utf-8-sig")) if META_PATH.exists() else {}
        return model, render_cache, meta, None
    except Exception as exc:
        return None, {}, {}, str(exc)


def clear_last_import() -> None:
    if LAST_IMPORT_DIR.exists():
        shutil.rmtree(LAST_IMPORT_DIR)


def model_stats(model: WarehouseModel) -> dict:
    cells = model.cells
    rows = [row for sheet in model.sheets for row in sheet.rows]
    zones = {cell.fill_color or "без зоны" for cell in cells}
    tiers = {str(cell.tier_number) for cell in cells}
    disabled = [cell for cell in cells if "block" in (cell.source or "").lower() or "заблок" in " ".join(cell.warnings).lower()]
    return {
        "excel_colored_cells": sum(1 for cell in cells if cell.fill_color),
        "cells": len(cells),
        "active_cells": len(cells) - len(disabled),
        "disabled_cells": len(disabled),
        "rows": len({row.row_number for row in rows}),
        "zones": len(zones),
        "tiers": len(tiers),
    }


def ensure_loaded_model() -> None:
    if "virtual_warehouse_model" in st.session_state:
        return
    model, render_cache, meta, error = load_last_import()
    if error:
        st.session_state["last_import_error"] = error
    if model is not None:
        meta = dict(meta)
        meta["loaded_from_cache"] = True
        st.session_state["virtual_warehouse_model"] = model
        st.session_state["virtual_warehouse_render_cache"] = render_cache
        st.session_state["virtual_warehouse_meta"] = meta
        st.session_state["virtual_warehouse_source"] = "сохранённая модель"


def filter_sheet(sheet: WarehouseSheet, zone: str, row_number: str, tier: str, availability: str) -> WarehouseSheet:
    filtered = WarehouseSheet(sheet.name, sheet.max_row, sheet.max_column, sheet.values, sheet.merged_ranges, warnings=sheet.warnings)
    for row in sheet.rows:
        if row_number != "Все" and str(row.row_number) != row_number:
            continue
        new_row = WarehouseRow(row.sheet_name, row.row_number, row.min_row, row.min_col, row.max_row, row.max_col, row.direction, row.confidence, warnings=row.warnings)
        for cell in row.potential_cells:
            is_disabled = "block" in (cell.source or "").lower() or "заблок" in " ".join(cell.warnings).lower()
            if zone != "Все" and (cell.fill_color or "без зоны") != zone:
                continue
            if tier != "Все" and str(cell.tier_number) != tier:
                continue
            if availability == "Только активные" and is_disabled:
                continue
            if availability == "Только заблокированные" and not is_disabled:
                continue
            new_row.potential_cells.append(cell)
        if new_row.potential_cells:
            filtered.rows.append(new_row)
    return filtered


def render_diagnostics(model: WarehouseModel, meta: dict, diagnostics: list[dict]) -> None:
    stats = model_stats(model)
    st.subheader("Диагностика склада")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Источник", meta.get("source", st.session_state.get("virtual_warehouse_source", "новый Excel")))
    c2.metric("Excel-ячеек с цветом", stats["excel_colored_cells"])
    c3.metric("Виртуальных ячеек", stats["cells"])
    c4.metric("Заблокированных", stats["disabled_cells"])
    c5, c6, c7, c8 = st.columns(4)
    c5.metric("Активных", stats["active_cells"])
    c6.metric("Рядов", stats["rows"])
    c7.metric("Зон", stats["zones"])
    c8.metric("Ярусов", stats["tiers"])
    st.caption(f"Построено: {meta.get('imported_at', '—')} · Схема: {meta.get('scheme_file_name', '—')} · Файл ячеек: {meta.get('cells_file_name', '—')} · Модель: {MODEL_PATH}")
    warnings = [d for d in diagnostics if d.get("level") in {"warning", "error"}]
    if warnings:
        st.dataframe(pd.DataFrame(warnings), use_container_width=True)


def render_performance(meta: dict) -> None:
    perf = meta.get("performance", {})
    st.subheader("Производительность")
    p1, p2, p3, p4, p5 = st.columns(5)
    p1.metric("Чтение Excel", f"{perf.get('read_excel_seconds', 0):.2f} сек")
    p2.metric("Построение модели", f"{perf.get('build_model_seconds', 0):.2f} сек")
    p3.metric("Подготовка визуализации", f"{perf.get('prepare_render_seconds', 0):.2f} сек")
    p4.metric("Сохранение модели", f"{perf.get('save_model_seconds', 0):.2f} сек")
    p5.metric("Рендеринг", f"{perf.get('render_seconds', 0):.2f} сек")
    st.caption(f"Модель загружена из кэша: {'да' if meta.get('loaded_from_cache') else 'нет'}")



def _is_numeric_text(value: str) -> bool:
    try:
        float(str(value).strip())
        return str(value).strip() != ""
    except ValueError:
        return False


def _cell_label(cell: dict) -> str:
    code = cell.get("code") or "без кода"
    return f"ряд {cell.get('row_number')} · ячейка {cell.get('cell_number')} · ярус {cell.get('tier')} · {code}"


def _short_cell_value(cell: dict | None) -> str:
    if not cell:
        return "—"
    return f"ряд {cell.get('row_number')}, ячейка {cell.get('cell_number')}, ярус {cell.get('tier')}, код {cell.get('code') or '—'}"


def _source_label(value: str | None) -> str:
    return {
        "excel": "Excel",
        "manual_add": "добавлена вручную",
        "manual_update": "изменена вручную",
    }.get(str(value or "excel"), str(value or "Excel"))


def _validate_manual_cell(model: dict, new_cell: dict, original_key: str | None = None) -> list[str]:
    errors: list[str] = []
    if not str(new_cell.get("row_number", "")).strip():
        errors.append("Ряд не может быть пустым.")
    if not str(new_cell.get("cell_number", "")).strip():
        errors.append("Номер ячейки не может быть пустым.")
    if not str(new_cell.get("tier", "")).strip():
        errors.append("Ярус не может быть пустым.")
    if new_cell.get("row_number") and not _is_numeric_text(str(new_cell.get("row_number"))):
        errors.append("Ряд должен быть числом.")
    if new_cell.get("cell_number") and not _is_numeric_text(str(new_cell.get("cell_number"))):
        errors.append("Номер ячейки должен быть числом.")
    if new_cell.get("tier") and not _is_numeric_text(str(new_cell.get("tier"))):
        errors.append("Ярус должен быть числом.")
    new_key = cell_key(new_cell)
    for cell in model.get("cells", []):
        if cell_key(cell) == new_key and new_key != original_key:
            errors.append("Ячейка с такой комбинацией ряд + номер ячейки + ярус уже существует.")
            break
    return errors


def _save_model_after_manual_change(model: dict, overrides: dict) -> dict:
    updated = apply_manual_overrides(model, overrides)
    updated["manual_change_counts"] = manual_change_counts(overrides)
    save_geometry_model(updated)
    st.session_state["geometry_model"] = updated
    return updated


def _manual_changes_dataframe(overrides: dict | None) -> pd.DataFrame:
    rows = []
    labels = {"add": "Добавление", "update": "Изменение", "delete": "Удаление"}
    for change in (overrides or {}).get("changes", []):
        rows.append({
            "Дата/время": change.get("created_at", ""),
            "Тип изменения": labels.get(change.get("change_type"), change.get("change_type", "")),
            "Старое значение": _short_cell_value(change.get("old_value")),
            "Новое значение": _short_cell_value(change.get("new_value")),
            "Комментарий": change.get("comment", ""),
        })
    return pd.DataFrame(rows, columns=["Дата/время", "Тип изменения", "Старое значение", "Новое значение", "Комментарий"])


def render_manual_cell_editor(model: dict) -> None:
    st.subheader("Ручное редактирование ячеек")
    st.caption("Добавляйте, изменяйте или удаляйте ячейки без правки исходного Excel. Все ручные операции сохраняются в текущей модели и журнале изменений.")
    overrides = load_manual_overrides()
    if overrides and overrides.get("source_model_id") != model.get("model_id"):
        overrides = None
    counts = manual_change_counts(overrides)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Ручных изменений", counts["total"])
    c2.metric("Добавлено вручную", counts["add"])
    c3.metric("Изменено вручную", counts["update"])
    c4.metric("Удалено вручную", counts["delete"])
    if counts["total"] == 0:
        st.caption("Ручных изменений пока нет.")

    add_tab, edit_tab, delete_tab, log_tab = st.tabs(["Добавить ячейку", "Изменить ячейку", "Удалить ячейку", "Журнал изменений"])
    existing_rows = sorted({str(cell.get("row_number")) for cell in model.get("cells", [])}, key=lambda value: (not value.isdigit(), value))
    existing_codes = {str(cell.get("code")) for cell in model.get("cells", []) if str(cell.get("code", "")).strip()}

    with add_tab:
        st.caption("Добавление меняет только сохранённую модель проекта и не редактирует исходный Excel-файл.")
        a1, a2, a3, a4 = st.columns(4)
        code = a1.text_input("Код ячейки", key="manual_add_code")
        row_number = a2.text_input("Ряд", key="manual_add_row")
        cell_number = a3.text_input("Номер ячейки", key="manual_add_cell")
        tier = a4.text_input("Ярус", value="1", key="manual_add_tier")
        comment = st.text_input("Комментарий, необязательно", key="manual_add_comment")
        row_is_new = bool(row_number.strip() and row_number.strip() not in existing_rows)
        allow_new_row = True
        if row_is_new:
            st.warning("Такого ряда ещё нет в модели. Новый ряд будет создан только после подтверждения.")
            allow_new_row = st.checkbox("Создать новый ряд", key="manual_add_allow_new_row")
        if st.button("Добавить ячейку", key="manual_add_button", type="primary"):
            new_cell = {"code": code.strip(), "row_number": row_number.strip(), "cell_number": cell_number.strip(), "tier": tier.strip(), "source": "manual_add"}
            errors = _validate_manual_cell(model, new_cell)
            if row_is_new and not allow_new_row:
                errors.append("Подтвердите создание нового ряда.")
            if code.strip() and code.strip() in existing_codes:
                st.warning("Код ячейки уже встречается в модели. Проверьте, что это ожидаемо.")
            if errors:
                for error in errors:
                    st.error(error)
            else:
                overrides = append_manual_change(model, "add", None, new_cell, comment.strip())
                _save_model_after_manual_change(model, overrides)
                if not update_data_revisions(model, ["geometry"], "manual_geometry_add"):
                    return
                st.success("Ячейка добавлена, координаты пересчитаны, модель сохранена.")
                st.rerun()

    cell_options = {_cell_label(cell): cell for cell in model.get("cells", [])}
    labels = list(cell_options)

    with edit_tab:
        if not labels:
            st.info("В модели пока нет ячеек для изменения.")
        else:
            f1, f2, f3, f4 = st.columns(4)
            row_filter = f1.selectbox("Фильтр: ряд", ["Все"] + existing_rows, key="manual_edit_row_filter")
            cell_filter = f2.text_input("Фильтр: номер ячейки", key="manual_edit_cell_filter")
            code_filter = f3.text_input("Фильтр: код", key="manual_edit_code_filter")
            tier_filter = f4.text_input("Фильтр: ярус", key="manual_edit_tier_filter")
            filtered = []
            for label, cell in cell_options.items():
                if row_filter != "Все" and str(cell.get("row_number")) != row_filter:
                    continue
                if cell_filter and str(cell.get("cell_number")) != cell_filter.strip():
                    continue
                if code_filter and code_filter.strip().lower() not in str(cell.get("code", "")).lower():
                    continue
                if tier_filter and str(cell.get("tier")) != tier_filter.strip():
                    continue
                filtered.append(label)
            if not filtered:
                st.warning("По фильтрам ячейки не найдены.")
            else:
                selected_label = st.selectbox("Выберите ячейку", filtered, key="manual_edit_selected")
                selected = cell_options[selected_label]
                e1, e2, e3, e4 = st.columns(4)
                new_code = e1.text_input("Код ячейки", value=str(selected.get("code", "")), key="manual_edit_code")
                new_row = e2.text_input("Ряд", value=str(selected.get("row_number", "")), key="manual_edit_row")
                new_cell_number = e3.text_input("Номер ячейки", value=str(selected.get("cell_number", "")), key="manual_edit_cell")
                new_tier = e4.text_input("Ярус", value=str(selected.get("tier", "")), key="manual_edit_tier")
                edit_comment = st.text_input("Комментарий, необязательно", key="manual_edit_comment")
                b1, b2 = st.columns(2)
                if b1.button("Сохранить изменения", key="manual_edit_save", type="primary"):
                    new_value = {"code": new_code.strip(), "row_number": new_row.strip(), "cell_number": new_cell_number.strip(), "tier": new_tier.strip(), "source": "manual_update"}
                    original_key = cell_key(selected)
                    errors = _validate_manual_cell(model, new_value, original_key=original_key)
                    if errors:
                        for error in errors:
                            st.error(error)
                    else:
                        overrides = append_manual_change(model, "update", selected, new_value, edit_comment.strip())
                        _save_model_after_manual_change(model, overrides)
                        if not update_data_revisions(model, ["geometry"], "manual_geometry_update"):
                            return
                        st.success("Изменения сохранены, координаты пересчитаны.")
                        st.rerun()
                if b2.button("Отменить", key="manual_edit_cancel"):
                    st.info("Изменение отменено: модель не менялась.")

    with delete_tab:
        if not labels:
            st.info("В модели пока нет ячеек для удаления.")
        else:
            selected_label = st.selectbox("Выберите ячейку для удаления", labels, key="manual_delete_selected")
            selected = cell_options[selected_label]
            st.write({
                "Код": selected.get("code", ""),
                "Ряд": selected.get("row_number", ""),
                "Ячейка": selected.get("cell_number", ""),
                "Ярус": selected.get("tier", ""),
                "Координаты": f"x={selected.get('x_center', 0):.2f}, y={selected.get('y_center', 0):.2f}",
                "Источник": _source_label(selected.get("source", "excel")),
            })
            confirm = st.checkbox("Вы точно хотите удалить ячейку?", key="manual_delete_confirm")
            delete_comment = st.text_input("Комментарий, необязательно", key="manual_delete_comment")
            if st.button("Удалить ячейку", key="manual_delete_button", type="primary"):
                if not confirm:
                    st.error("Подтвердите удаление ячейки.")
                else:
                    overrides = append_manual_change(model, "delete", selected, None, delete_comment.strip())
                    _save_model_after_manual_change(model, overrides)
                    if not update_data_revisions(model, ["geometry"], "manual_geometry_delete"):
                        return
                    st.success("Ячейка удалена из модели, координаты пересчитаны.")
                    st.rerun()

    with log_tab:
        st.subheader("Журнал ручных изменений")
        log_df = _manual_changes_dataframe(overrides)
        if log_df.empty:
            st.info("Ручных изменений пока нет.")
        else:
            st.dataframe(log_df, use_container_width=True)
            st.download_button("Скачать журнал изменений", log_df.to_csv(index=False).encode("utf-8-sig"), file_name="manual_overrides_log.csv", mime="text/csv")
        st.download_button("Скачать текущую модель Excel", export_current_model_excel_bytes(model), file_name="current_warehouse_model.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        clear_confirm = st.checkbox("Подтверждаю очистку ручных изменений", key="manual_clear_confirm")
        if st.button("Очистить ручные изменения", key="manual_clear_button"):
            if not clear_confirm:
                st.error("Подтвердите очистку ручных изменений.")
            else:
                clear_manual_overrides()
                base_model = rebuild_geometry_from_cells(model, model.get("base_cells", model.get("cells", [])), keep_base_cells=False)
                base_model["manual_change_counts"] = manual_change_counts(None)
                save_geometry_model(base_model)
                st.session_state["geometry_model"] = base_model
                st.success("Ручные изменения очищены. Модель восстановлена к состоянию после последней загрузки Excel.")
                st.rerun()


def render_inventory_placement(model: dict) -> dict:
    st.subheader("Инвентаризация")
    st.caption("Загрузите инвент на конец дня, проверьте строки и зафиксируйте фактический переходящий остаток без перемещения сохранённых SKU.")
    with measure_step("load_inventory_state"):
        state, state_warning = load_placement_state(model)
    if state_warning:
        st.warning(state_warning)
    settings = state.setdefault("settings", {"allow_mixed_sku_in_deep_lane": False})
    allow_mixed = st.checkbox(
        "Разрешить несколько SKU в одной набивной ячейке",
        value=bool(settings.get("allow_mixed_sku_in_deep_lane", False)),
        key="placement_allow_mixed",
    )
    settings["allow_mixed_sku_in_deep_lane"] = allow_mixed

    upload_tab, unplaced_tab, manual_tab, edit_tab = st.tabs([
        "Загрузка и сверка инвента",
        "Товар без привязки к ячейкам",
        "Разместить вручную",
        "Редактировать размещение",
    ])

    with upload_tab:
        st.caption("Шаги: загрузите файл → проверьте сопоставление и предпросмотр → зафиксируйте переходящий остаток.")
        inventory_file = st.file_uploader("Загрузить Excel с остатками", type=["xlsx"], key="inventory_upload")
        if inventory_file is None:
            st.button("Зафиксировать переходящий остаток", type="primary", disabled=True, key="inventory_reconcile_disabled")
        else:
            inventory_bytes = inventory_file.getvalue()
            inventory_hash = file_hash(inventory_bytes)
            sheet_names = get_inventory_sheet_names(inventory_bytes)
            sheet_name = st.selectbox("Лист инвента", sheet_names, key="inventory_sheet")
            header_rows = st.radio("Строк заголовка инвента", [1, 2], index=0, horizontal=True, key="inventory_header_rows")
            inv_df = read_inventory_table_cached(inventory_bytes, inventory_hash, sheet_name, header_rows)
            st.caption(f"Прочитано строк инвента: {len(inv_df)}; колонок: {len(inv_df.columns)}")
            with st.expander("Предпросмотр инвента", expanded=False):
                st.dataframe(inv_df.head(30), use_container_width=True)
            detected = detect_inventory_columns(inv_df)
            columns = [None] + list(inv_df.columns)
            st.caption("Проверьте автоопределение колонок или выберите вручную.")
            c1, c2, c3, c4 = st.columns(4)
            c5, c6, c7, c8 = st.columns(4)
            mapping = {
                "sku_code": c1.selectbox("Код товара / SKU", columns, index=columns.index(detected["sku_code"]) if detected["sku_code"] in columns else 0, key="inv_map_sku"),
                "sku_name": c2.selectbox("Наименование товара", columns, index=columns.index(detected["sku_name"]) if detected["sku_name"] in columns else 0, key="inv_map_name"),
                "qty_pallets": c3.selectbox("Количество паллет", columns, index=columns.index(detected["qty_pallets"]) if detected["qty_pallets"] in columns else 0, key="inv_map_pallets"),
                "qty_boxes": c4.selectbox("Количество коробов", columns, index=columns.index(detected["qty_boxes"]) if detected["qty_boxes"] in columns else 0, key="inv_map_boxes"),
                "cell_address": c5.selectbox("Адрес ячейки", columns, index=columns.index(detected["cell_address"]) if detected["cell_address"] in columns else 0, key="inv_map_address"),
                "row_number": c6.selectbox("Ряд", columns, index=columns.index(detected["row_number"]) if detected["row_number"] in columns else 0, key="inv_map_row"),
                "cell_number": c7.selectbox("Ячейка", columns, index=columns.index(detected["cell_number"]) if detected["cell_number"] in columns else 0, key="inv_map_cell"),
                "tier": c8.selectbox("Ярус", columns, index=columns.index(detected["tier"]) if detected["tier"] in columns else 0, key="inv_map_tier"),
            }
            c9 = st.columns(1)[0]
            mapping["weight_class"] = c9.selectbox("Весовая категория", columns, index=columns.index(detected["weight_class"]) if detected.get("weight_class") in columns else 0, key="inv_map_weight_class")
            for key in ["expiry_date", "batch", "characteristic", "characteristic_code", "characteristic_name", "weight", "volume"]:
                mapping[key] = detected.get(key)
            normalized_inventory, inv_diagnostics = normalize_inventory_table_cached(inv_df.to_json(orient="split", force_ascii=False), json.dumps(mapping, ensure_ascii=False))
            if inv_diagnostics:
                st.dataframe(pd.DataFrame(inv_diagnostics), use_container_width=True)
            has_cell_columns = bool(mapping.get("cell_address") or (mapping.get("row_number") and mapping.get("cell_number")))
            if not has_cell_columns:
                st.warning("В инвенте нет адресов ячеек. Система не может восстановить фактическое расположение товара. Автоматическое размещение будет модельным и используется только для расчётов.")
            if st.button("Зафиксировать переходящий остаток", type="primary", key="inventory_reconcile_button"):
                if any(item.get("level") == "error" for item in inv_diagnostics):
                    st.error("Исправьте обязательные колонки перед сверкой.")
                else:
                    reconciled_state, report = reconcile_placements_with_inventory(model, state, normalized_inventory)
                    if not report.get("success"):
                        st.error("Сверка не выполнена: исправьте ошибки инвента. Текущее размещение не изменено.")
                        st.dataframe(pd.DataFrame(report.get("details", [])), use_container_width=True)
                    else:
                        save_placement_state(reconciled_state)
                        domains = ["inventory"] + (["placements"] if reconciled_state != state else [])
                        if not update_data_revisions(model, domains, "reconcile_inventory"):
                            return model
                        st.session_state["placement_state"] = reconciled_state
                        st.session_state["last_inventory_reconciliation_report"] = report
                        summary = report.get("summary", {})
                        st.success(
                            f"Переходящий остаток зафиксирован: SKU до — {summary.get('SKU до сверки', 0)}, "
                            f"по инвенту — {summary.get('SKU в инвенте', 0)}, после — {summary.get('SKU после сверки', 0)}."
                        )
                        r1, r2, r3 = st.columns(3)
                        r1.metric("Удалено SKU", summary.get("Полностью удалено SKU", 0))
                        r2.metric("Уменьшено SKU", summary.get("Уменьшено SKU", 0))
                        r3.metric("Освобождено ячеек", summary.get("Освобождено логических ячеек", 0))
                        with st.expander("Подробная диагностика"):
                            st.dataframe(pd.DataFrame(report.get("details", [])), use_container_width=True)
                        st.rerun()
            with st.expander("Импорт адресного инвента (служебный сценарий)"):
                st.caption("Сохраняет прежний сценарий импорта фактических адресов. Используйте его только для первичной загрузки склада.")
                if st.button("Импортировать адресный инвент", key="inventory_import_button"):
                    if any(item.get("level") == "error" for item in inv_diagnostics):
                        st.error("Исправьте обязательные колонки перед импортом.")
                    else:
                        state, placement_import_diag = import_inventory(model, normalized_inventory, allow_replace=True)
                        if not update_data_revisions(model, ["inventory", "placements"], "import_inventory"):
                            return model
                        st.session_state["placement_state"] = state
                        st.success("Адресный инвент импортирован.")
                        st.dataframe(pd.DataFrame(placement_import_diag), use_container_width=True)
                        st.rerun()

        last_report = st.session_state.get("last_inventory_reconciliation_report")
        if last_report:
            with st.expander("Результат последней сверки с инвентом"):
                _metric_grid(last_report.get("summary", {}), columns=4)
                st.dataframe(pd.DataFrame(last_report.get("details", [])), use_container_width=True)

    with unplaced_tab:
        unplaced = state.get("unplaced_inventory", [])
        total_unplaced_pallets = sum(float(item.get("qty_pallets", 0) or 0) for item in unplaced)
        total_unplaced_boxes = sum(float(item.get("qty_boxes", 0) or 0) for item in unplaced)
        u1, u2, u3 = st.columns(3)
        u1.metric("SKU без ячейки", len(unplaced))
        u2.metric("Паллет без ячейки", f"{total_unplaced_pallets:g}")
        u3.metric("Коробов без ячейки", f"{total_unplaced_boxes:g}")
        if unplaced:
            st.warning("Это модельное размещение, а не фактическое, потому что в инвенте нет адресов ячеек.")
            st.dataframe(pd.DataFrame(unplaced), use_container_width=True)
        else:
            st.info("Товаров без привязки к ячейкам сейчас нет.")
        st.caption("Новый приход размещается кнопкой «Добавить приход на текущий склад» на предыдущем шаге. Здесь остаются только ручные операции с неразмещённым товаром.")
        if st.button("Разложить автоматически по складу", disabled=not unplaced, key="auto_place_inventory"):
            state, auto_diag = auto_place_unplaced(model, state, allow_mixed_sku_in_deep_lane=allow_mixed)
            if not update_data_revisions(model, ["placements"], "auto_place_inventory"):
                return model
            st.session_state["placement_state"] = state
            st.success("Автоматическое модельное размещение выполнено.")
            st.dataframe(pd.DataFrame(auto_diag), use_container_width=True)
            st.rerun()

    with manual_tab:
        unplaced = state.get("unplaced_inventory", [])
        if not unplaced:
            st.info("Нет товара без ячейки для ручного размещения.")
        else:
            options = {f"{idx}: {item.get('sku_code')} · {item.get('sku_name')} · {item.get('qty_pallets')} паллет": idx for idx, item in enumerate(unplaced)}
            selected_label = st.selectbox("Товар из списка Без ячейки", list(options), key="manual_place_item")
            m1, m2, m3, m4 = st.columns(4)
            target_row = m1.text_input("Ряд", key="manual_place_row")
            target_cell = m2.text_input("Ячейка", key="manual_place_cell")
            target_tier = m3.text_input("Ярус", value="1", key="manual_place_tier")
            qty = m4.number_input("Паллет к размещению", min_value=0.0, value=1.0, step=1.0, key="manual_place_qty")
            if st.button("Разместить", key="manual_place_button", type="primary"):
                state, error = manual_place(model, state, options[selected_label], target_row, target_cell, target_tier, qty, allow_mixed_sku_in_deep_lane=allow_mixed)
                if error:
                    st.error(error)
                else:
                    if not update_data_revisions(model, ["placements"], "manual_place"):
                        return model
                    st.session_state["placement_state"] = state
                    st.success("Товар размещён вручную, placements.json обновлён.")
                    st.rerun()

    with edit_tab:
        placements = state.get("placements", [])
        if not placements:
            st.info("Размещений пока нет.")
        else:
            placement_options = {f"{p.get('sku_code')} · {p.get('cell_key')} · {p.get('qty_pallets')} паллет · {p.get('source')}": p for p in placements}
            selected = st.selectbox("Выберите размещение", list(placement_options), key="placement_edit_selected")
            selected_placement = placement_options[selected]
            e1, e2, e3, e4 = st.columns(4)
            new_qty = e1.number_input("Новое количество паллет", min_value=0.0, value=float(selected_placement.get("qty_pallets", 0) or 0), step=1.0, key="placement_edit_qty")
            move_row = e2.text_input("Новый ряд", value=str(selected_placement.get("row_number", "")), key="placement_move_row")
            move_cell = e3.text_input("Новая ячейка", value=str(selected_placement.get("cell_number", "")), key="placement_move_cell")
            move_tier = e4.text_input("Новый ярус", value=str(selected_placement.get("tier", "1")), key="placement_move_tier")
            a1, a2, a3 = st.columns(3)
            if a1.button("Изменить количество", key="placement_update_qty_button"):
                state, error = update_placement_qty(model, state, selected_placement["placement_id"], new_qty)
                if error:
                    st.error(error)
                else:
                    if not update_data_revisions(model, ["placements"], "update_placement_qty"):
                        return model
                    st.session_state["placement_state"] = state
                    st.success("Количество размещения изменено.")
                    st.rerun()
            if a2.button("Перенести в другую ячейку", key="placement_move_button"):
                state, error = move_placement(model, state, selected_placement["placement_id"], move_row, move_cell, move_tier, allow_mixed_sku_in_deep_lane=allow_mixed)
                if error:
                    st.error(error)
                else:
                    if not update_data_revisions(model, ["placements"], "move_placement"):
                        return model
                    st.session_state["placement_state"] = state
                    st.success("Размещение перенесено.")
                    st.rerun()
            if a3.button("Удалить размещение", key="placement_delete_button"):
                state, error = delete_placement(state, selected_placement["placement_id"])
                if error:
                    st.error(error)
                else:
                    if not update_data_revisions(model, ["placements"], "delete_placement"):
                        return model
                    st.session_state["placement_state"] = state
                    st.success("Размещение удалено, товар возвращён в Без ячейки.")
                    st.rerun()

    return attach_placements_to_model(model, state)


def _metric_grid(metrics: dict[str, object], columns: int = 5) -> None:
    items = list(metrics.items())
    for start in range(0, len(items), columns):
        cols = st.columns(min(columns, len(items) - start))
        for col, (label, value) in zip(cols, items[start:start + columns]):
            col.metric(label, value)


def render_placement_diagnostics_section(model: dict, state: dict) -> None:
    st.subheader("Диагностика размещения")
    st.caption("Раздел анализирует текущие warehouse_model.json, placements.json и receipts.json. Открытие диагностики не запускает размещение и не изменяет сохранённые файлы.")
    with measure_step("load_receipts"):
        receipts_state, receipts_warning = load_receipts_state(model)
    if receipts_warning:
        st.warning(receipts_warning)
        receipts_state = {"receipts": []}
    snapshot, snapshot_warning = load_pre_placement_snapshot(model)
    if snapshot_warning:
        st.info(snapshot_warning)
    diagnostics = build_placement_diagnostics(model, state, receipts_state, snapshot)
    if diagnostics.get("snapshot_warning"):
        st.caption(diagnostics["snapshot_warning"])
    _metric_grid(diagnostics["summary"], columns=4)

    st.markdown("**Аналитика по весовым зонам**")
    zone_df = pd.DataFrame(diagnostics["zone_rows"])
    if zone_df.empty:
        st.info("Нет данных по весовым зонам.")
    else:
        st.dataframe(zone_df, use_container_width=True, hide_index=True)

    st.markdown("**Изменение весовых зон**")
    zc1, zc2 = st.columns(2)
    zc1.metric("Рядов с изменённой зоной", diagnostics["changed_rows_count"])
    zc2.metric("Ячеек с изменённой зоной", diagnostics["changed_cells_count"])
    changes_df = pd.DataFrame(diagnostics["zone_changes"])
    if changes_df.empty:
        st.info("Нет данных о рядах.")
    else:
        st.dataframe(changes_df, use_container_width=True, hide_index=True)

    st.markdown("**Детальная таблица занятых ячеек**")
    detail_df = pd.DataFrame(diagnostics["occupied_rows"])
    if detail_df.empty:
        st.info("Занятых ячеек нет.")
    else:
        f1, f2, f3 = st.columns(3)
        row_filter = f1.multiselect("Ряд", sorted(detail_df["Ряд"].dropna().astype(str).unique()), key="diag_filter_row")
        zone_filter = f2.multiselect("Весовая зона", sorted(detail_df["Весовая зона ячейки"].dropna().astype(str).unique()), key="diag_filter_zone")
        category_filter = f3.multiselect("Категория SKU", sorted(detail_df["Категория SKU"].dropna().astype(str).unique()), key="diag_filter_category")
        f4, f5, f6 = st.columns(3)
        reason_filter = f4.multiselect("Причина", sorted(detail_df["Код причины размещения"].dropna().astype(str).unique()), key="diag_filter_reason")
        status_filter = f5.multiselect("Статус", sorted(detail_df["Источник"].dropna().astype(str).unique()), key="diag_filter_source")
        only_partial = f6.checkbox("Только частично заполненные", key="diag_filter_partial")
        only_full = f6.checkbox("Только полностью заполненные", key="diag_filter_full")
        filtered = detail_df.copy()
        if row_filter:
            filtered = filtered[filtered["Ряд"].astype(str).isin(row_filter)]
        if zone_filter:
            filtered = filtered[filtered["Весовая зона ячейки"].astype(str).isin(zone_filter)]
        if category_filter:
            filtered = filtered[filtered["Категория SKU"].astype(str).isin(category_filter)]
        if reason_filter:
            filtered = filtered[filtered["Код причины размещения"].astype(str).isin(reason_filter)]
        if status_filter:
            filtered = filtered[filtered["Источник"].astype(str).isin(status_filter)]
        if only_partial:
            filtered = filtered[(filtered["Стало после"] > 0) & (filtered["Стало после"] < filtered["Вместимость"])]
        if only_full:
            filtered = filtered[filtered["Стало после"] >= filtered["Вместимость"]]
        st.dataframe(filtered, use_container_width=True, hide_index=True)

    st.markdown("**Не размещено**")
    unplaced_df = pd.DataFrame(diagnostics["unplaced_rows"])
    if unplaced_df.empty:
        st.success("Неразмещённых позиций нет.")
    else:
        st.dataframe(unplaced_df, use_container_width=True, hide_index=True)

    st.download_button("Скачать размещение в Excel", export_placements_excel_bytes(model, state), file_name="placements.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    log_df = pd.DataFrame(state.get("journal", []))
    st.subheader("Журнал размещения")
    if log_df.empty:
        st.info("Журнал размещения пока пуст.")
    else:
        st.dataframe(log_df, use_container_width=True)
        st.download_button("Скачать журнал размещения", log_df.to_csv(index=False).encode("utf-8-sig"), file_name="placement_journal.csv", mime="text/csv")


RECEIPT_STATUS_LABELS = {
    "not_placed": "Не размещено",
    "partially_placed": "Частично размещено",
    "placed": "Размещено",
    "error": "Ошибка",
}

RECEIPT_WEIGHT_CLASS_LABELS = {
    "heavy": "Тяжёлое",
    "medium": "Среднее",
    "light": "Лёгкое",
    "fragile": "Хрупкое",
    "unclassified": "Не классифицировано",
}

RECEIPT_TABLE_COLUMNS = {
    "receipt_date": "Дата прихода",
    "receipt_number": "Номер приходного ордера",
    "receipt_line_id": "Строка прихода",
    "sku_key": "Ключ SKU",
    "receipt_document": "Документ прихода",
    "sku_code": "Код товара",
    "sku_name": "Наименование",
    "characteristic_name": "Характеристика",
    "qty_pallets": "Количество паллет",
    "qty_boxes": "Количество коробов",
    "expiry_date": "Срок годности",
    "source_weight": "Вес",
    "fragile_flag": "Признак хрупкости",
    "source_zone": "Исходная зона из 1С",
    "calculated_zone": "Рассчитанная зона",
    "zone_calculation_reason": "Причина расчёта",
    "source_weight_raw": "Исходный вес",
    "weight_parse_status": "Статус веса",
    "weight_parse_reason": "Причина ошибки веса",
    "zone_calculation_status": "Статус расчёта",
    "weight_class": "Зона размещения",
    "placement_status": "Статус размещения",
}


def _receipt_dataframe(receipts: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(receipts)
    if df.empty:
        return df
    columns = [column for column in RECEIPT_TABLE_COLUMNS if column in df.columns]
    result = df[columns].copy()
    for zone_column in ["weight_class", "calculated_zone"]:
        if zone_column in result.columns:
            result[zone_column] = result[zone_column].map(RECEIPT_WEIGHT_CLASS_LABELS).fillna(result[zone_column])
    if "placement_status" in result.columns:
        result["placement_status"] = result["placement_status"].map(RECEIPT_STATUS_LABELS).fillna(result["placement_status"])
    return result.rename(columns=RECEIPT_TABLE_COLUMNS)


def _receipt_zone_summary(receipts: list[dict]) -> dict[str, int]:
    summary = {"heavy": 0, "medium": 0, "light": 0, "fragile": 0, "unclassified": 0}
    for receipt in receipts:
        weight_class = str(receipt.get("calculated_zone") or receipt.get("weight_class") or "unclassified")
        if weight_class not in summary:
            weight_class = "unclassified"
        summary[weight_class] += 1
    return summary


def _zone_calculation_dataframe(receipts: list[dict]) -> pd.DataFrame:
    receipt_count_by_sku = {}
    for receipt in receipts:
        sku_key = receipt.get("sku_key", "")
        if sku_key:
            receipt_count_by_sku.setdefault(sku_key, set()).add(receipt.get("receipt_number", ""))
    rows = []
    for receipt in receipts:
        sku_key = receipt.get("sku_key", "")
        rows.append({
            "Номер приходного ордера": receipt.get("receipt_number", ""),
            "receipt_line_id": receipt.get("receipt_line_id", ""),
            "sku_key": sku_key,
            "SKU": receipt.get("sku_code", ""),
            "Номенклатура": receipt.get("sku_name", ""),
            "Характеристика": receipt.get("characteristic_name", ""),
            "Вес": receipt.get("source_weight", ""),
            "Исходное значение веса": receipt.get("source_weight_raw", ""),
            "Статус преобразования веса": receipt.get("weight_parse_status", ""),
            "Причина ошибки веса": receipt.get("weight_parse_reason", ""),
            "Признак хрупкости": "Да" if receipt.get("fragile_flag") else "Нет",
            "Исходная зона из 1С": receipt.get("source_zone", ""),
            "Рассчитанная зона": RECEIPT_WEIGHT_CLASS_LABELS.get(receipt.get("calculated_zone", "unclassified"), receipt.get("calculated_zone", "")),
            "Количество паллет": receipt.get("qty_pallets", ""),
            "Приходов с этим SKU": len(receipt_count_by_sku.get(sku_key, set())),
            "Причина расчёта": receipt.get("zone_calculation_reason", ""),
            "Статус": receipt.get("zone_calculation_status", ""),
        })
    return pd.DataFrame(rows)


def _render_zone_classification_result(state: dict) -> None:
    diag = state.get("zone_classification_diagnostics", {})
    if diag:
        c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
        c1.metric("Всего SKU", diag.get("Всего SKU", 0))
        c2.metric("Лёгких SKU", diag.get("Лёгких SKU", 0))
        c3.metric("Средних SKU", diag.get("Средних SKU", 0))
        c4.metric("Тяжёлых SKU", diag.get("Тяжёлых SKU", 0))
        c5.metric("Хрупких SKU", diag.get("Хрупких SKU", 0))
        c6.metric("Без категории", diag.get("SKU без рассчитанной категории", 0))
        c7.metric("Конфликтов", diag.get("Конфликтов данных", 0))
        st.metric("Несовпадений с исходной зоной 1С", diag.get("Несовпадений с исходной зоной 1С", 0))
    if state.get("receipts"):
        st.dataframe(_zone_calculation_dataframe(state.get("receipts", [])), use_container_width=True)
        bad_weight_rows = [
            {
                "Номер приходного ордера": receipt.get("receipt_number", ""),
                "receipt_line_id": receipt.get("receipt_line_id", ""),
                "sku_key": receipt.get("sku_key", ""),
                "Номенклатура": receipt.get("sku_name", ""),
                "Характеристика": receipt.get("characteristic_name", ""),
                "Исходное значение веса": receipt.get("source_weight_raw", ""),
                "Нормализованное значение": receipt.get("source_weight", ""),
                "Причина ошибки": receipt.get("weight_parse_reason", ""),
                "Номер строки Excel": receipt.get("source_row_number", ""),
            }
            for receipt in state.get("receipts", [])
            if receipt.get("weight_parse_status") != "ok"
        ]
        if bad_weight_rows:
            st.warning("Есть строки, где вес не удалось преобразовать.")
            st.dataframe(pd.DataFrame(bad_weight_rows), use_container_width=True)


def _render_receipt_placement_diagnostics(diag: dict | None) -> None:
    if not diag:
        return
    st.subheader("Результат расчёта размещения приходов")
    st.dataframe(pd.DataFrame([{"Показатель": key, "Значение": value} for key, value in diag.items() if key != "Неразмещённые позиции"]), use_container_width=True)
    unplaced = diag.get("Неразмещённые позиции") or []
    if unplaced:
        st.warning("Часть приходов не размещена. Смотрите причину в колонке unplaced_reason.")
        st.dataframe(pd.DataFrame(unplaced), use_container_width=True)


def render_receipts_section(model: dict) -> None:
    st.subheader("Приходы")
    state, state_warning = load_receipts_state(model)
    if state_warning:
        st.warning(state_warning)
    receipts = state.get("receipts", [])
    diagnostics = state.get("diagnostics", {})
    if receipts:
        c1, c2, c3, c4, c5 = st.columns(5)
        c1.metric("Файл", state.get("source_file_name", "—"))
        c2.metric("Дата загрузки", state.get("created_at", "—"))
        c3.metric("Строк", len(receipts))
        c4.metric("SKU", diagnostics.get("Всего SKU", 0))
        c5.metric("Паллет", f"{diagnostics.get('Всего паллет', 0):g}" if isinstance(diagnostics.get("Всего паллет", 0), (int, float)) else diagnostics.get("Всего паллет", 0))
        st.info("Приходы загружены. Алгоритм размещения по ячейкам пока не запущен. Все строки имеют статус ‘Не размещено’.")
    upload_tab, data_tab, diag_tab = st.tabs(["Загрузка приходов", "Приходы к размещению", "Диагностика приходов"])

    with upload_tab:
        st.caption("Загрузите Excel с приходами и проверьте соответствие колонок. После загрузки приходы сохраняются как отдельный слой данных.")
        receipt_file = st.file_uploader("Загрузить Excel с приходами", type=["xlsx"], key="receipt_upload")
        replace_current = st.checkbox("Заменить текущие загруженные приходы", value=True, key="receipt_replace_current")
        if receipt_file is not None:
            receipt_bytes = receipt_file.getvalue()
            receipt_hash = file_hash(receipt_bytes)
            sheet_names = get_receipt_sheet_names(receipt_bytes)
            sheet_name = st.selectbox("Выбрать лист", sheet_names, key="receipt_sheet")
            header_rows = st.radio("Строк заголовка приходов", [1, 2], index=0, horizontal=True, key="receipt_header_rows")
            receipt_df = read_receipt_table_cached(receipt_bytes, receipt_hash, sheet_name, header_rows)
            with st.expander("Предпросмотр", expanded=False):
                st.dataframe(receipt_df.head(30), use_container_width=True)
            detected = detect_receipt_columns(receipt_df)
            columns = [None] + list(receipt_df.columns)
            st.caption("Ручной выбор колонок: проверьте автоопределение или выберите колонки вручную.")
            c1, c2, c3, c4, c5 = st.columns(5)
            c6, c7, c8, c9, c10 = st.columns(5)
            c11, c12, c13, c14, c15 = st.columns(5)
            c16 = st.columns(1)[0]
            zone_detected = detect_zone_classification_columns(receipt_df)
            mapping = {
                "sku_code": c1.selectbox("Код товара", columns, index=columns.index(detected["sku_code"]) if detected["sku_code"] in columns else 0, key="receipt_map_sku"),
                "sku_name": c2.selectbox("Наименование", columns, index=columns.index(detected["sku_name"]) if detected["sku_name"] in columns else 0, key="receipt_map_name"),
                "qty_pallets": c3.selectbox("Количество паллет", columns, index=columns.index(detected["qty_pallets"]) if detected["qty_pallets"] in columns else 0, key="receipt_map_pallets"),
                "qty_boxes": c4.selectbox("Количество коробов", columns, index=columns.index(detected["qty_boxes"]) if detected["qty_boxes"] in columns else 0, key="receipt_map_boxes"),
                "qty_units": c5.selectbox("Базовое количество", columns, index=columns.index(detected["qty_units"]) if detected["qty_units"] in columns else 0, key="receipt_map_units"),
                "receipt_date": c6.selectbox("Дата прихода", columns, index=columns.index(detected["receipt_date"]) if detected["receipt_date"] in columns else 0, key="receipt_map_date"),
                "receipt_number": c7.selectbox("Номер документа", columns, index=columns.index(detected["receipt_number"]) if detected["receipt_number"] in columns else 0, key="receipt_map_number"),
                "receipt_document": c8.selectbox("Документ прихода", columns, index=columns.index(detected["receipt_document"]) if detected["receipt_document"] in columns else 0, key="receipt_map_document"),
                "warehouse": c9.selectbox("Склад", columns, index=columns.index(detected["warehouse"]) if detected["warehouse"] in columns else 0, key="receipt_map_warehouse"),
                "warehouse_zone": c10.selectbox("Складская зона", columns, index=columns.index(detected["warehouse_zone"]) if detected["warehouse_zone"] in columns else 0, key="receipt_map_zone"),
                "characteristic_code": c11.selectbox("Код характеристики", columns, index=columns.index(detected["characteristic_code"]) if detected["characteristic_code"] in columns else 0, key="receipt_map_char_code"),
                "characteristic_name": c12.selectbox("Характеристика", columns, index=columns.index(detected["characteristic_name"]) if detected["characteristic_name"] in columns else 0, key="receipt_map_char_name"),
                "batch": c13.selectbox("Партия", columns, index=columns.index(detected["batch"]) if detected["batch"] in columns else 0, key="receipt_map_batch"),
                "expiry_date": c14.selectbox("Срок годности", columns, index=columns.index(detected["expiry_date"]) if detected["expiry_date"] in columns else 0, key="receipt_map_expiry"),
                "comment": c15.selectbox("Комментарий", columns, index=columns.index(detected["comment"]) if detected["comment"] in columns else 0, key="receipt_map_comment"),
                "weight_class": None,
            }
            st.subheader("Правила определения зоны товара")
            st.caption("Настройте весовые границы и источник признака хрупкости. Система рассчитает категорию каждого SKU перед размещением.")
            zc1, zc2, zc3 = st.columns(3)
            mapping["source_weight"] = zc1.selectbox("Колонка с весом товара", columns, index=columns.index(zone_detected["weight_column"]) if zone_detected.get("weight_column") in columns else 0, key="receipt_map_source_weight")
            mapping["fragile_flag"] = zc2.selectbox("Колонка с признаком хрупкости", columns, index=columns.index(zone_detected["fragile_column"]) if zone_detected.get("fragile_column") in columns else 0, key="receipt_map_fragile_flag")
            mapping["source_zone"] = zc3.selectbox("Колонка с исходной зоной из 1С", columns, index=columns.index(zone_detected["source_zone_column"]) if zone_detected.get("source_zone_column") in columns else 0, key="receipt_map_source_zone")
            zw1, zw2 = st.columns(2)
            max_light_weight = zw1.number_input("Максимальный вес лёгкого товара, кг", min_value=0.0, value=5.0, step=0.1, key="receipt_max_light_weight")
            max_medium_weight = zw2.number_input("Максимальный вес среднего товара, кг", min_value=0.0, value=15.0, step=0.1, key="receipt_max_medium_weight")
            zone_settings = default_zone_classification_settings()
            zone_settings.update({"weight_column": mapping.get("source_weight"), "fragile_column": mapping.get("fragile_flag"), "source_zone_column": mapping.get("source_zone"), "max_light_weight_kg": max_light_weight, "max_medium_weight_kg": max_medium_weight})
            if max_medium_weight <= max_light_weight:
                st.error("Максимальный вес среднего товара должен быть строго больше максимального веса лёгкого товара.")
            normalized_receipts, receipt_diagnostics, receipt_messages = normalize_receipt_table_cached(receipt_df.to_json(orient="split", force_ascii=False), json.dumps(mapping, ensure_ascii=False))
            if receipt_messages:
                st.dataframe(pd.DataFrame(receipt_messages), use_container_width=True)
            if st.button("Загрузить приходы", type="primary", key="receipt_import_button"):
                if any(item.get("level") == "error" for item in receipt_messages):
                    st.error("Исправьте обязательные колонки или ошибки данных перед загрузкой приходов.")
                elif not replace_current and receipts:
                    st.error("Подтвердите замену текущих загруженных приходов или очистите их вручную.")
                else:
                    if max_medium_weight <= max_light_weight:
                        st.error("Исправьте границы веса перед загрузкой приходов.")
                        return
                    new_state = make_receipts_state(model, receipt_file.name, receipt_hash, normalized_receipts, receipt_diagnostics, mapping, zone_settings)
                    save_receipts_state(new_state)
                    if not update_data_revisions(model, ["receipts"], "save_receipts"):
                        return
                    st.success("Приходы загружены и сохранены. Все строки имеют статус ‘Не размещено’.")
                    st.rerun()

    with data_tab:
        if not receipts:
            st.info("Загруженных приходов пока нет. Сначала загрузите и проверьте файл прихода.")
            st.button("Добавить приход на текущий склад", type="primary", disabled=True, key="receipt_add_disabled")
        else:
            st.dataframe(_receipt_dataframe(receipts), use_container_width=True)
            st.subheader("Правила определения зоны товара")
            st.caption("Настройте весовые границы и источник признака хрупкости. Система рассчитает категорию каждого SKU перед размещением.")
            stored_zone_settings = {**default_zone_classification_settings(), **state.get("zone_classification_settings", {})}
            st.caption(
                "Колонка веса: "
                f"{stored_zone_settings.get('weight_column') or 'не выбрана'} · "
                "признак хрупкости: "
                f"{stored_zone_settings.get('fragile_column') or 'не выбран'} · "
                "исходная зона 1С: "
                f"{stored_zone_settings.get('source_zone_column') or 'не выбрана'}"
            )
            zw1, zw2 = st.columns(2)
            current_light_limit = zw1.number_input(
                "Максимальный вес лёгкого товара, кг",
                min_value=0.0,
                value=float(stored_zone_settings.get("max_light_weight_kg", 5.0) or 0.0),
                step=0.1,
                key="receipt_data_max_light_weight",
            )
            current_medium_limit = zw2.number_input(
                "Максимальный вес среднего товара, кг",
                min_value=0.0,
                value=float(stored_zone_settings.get("max_medium_weight_kg", 15.0) or 0.0),
                step=0.1,
                key="receipt_data_max_medium_weight",
            )
            current_zone_settings = {
                **stored_zone_settings,
                "max_light_weight_kg": current_light_limit,
                "max_medium_weight_kg": current_medium_limit,
            }
            current_settings_hash = zone_classification_settings_hash(current_zone_settings)
            saved_settings_hash = stored_zone_settings.get("settings_hash") or state.get("zone_classification_diagnostics", {}).get("settings_hash")
            if saved_settings_hash and saved_settings_hash != current_settings_hash:
                st.warning("Границы веса изменились. Старый расчёт зон товаров устарел — нажмите «Рассчитать зоны товаров» повторно.")
            if current_medium_limit <= current_light_limit:
                st.error("Максимальный вес среднего товара должен быть строго больше максимального веса лёгкого товара.")
            if st.button("Рассчитать зоны товаров", key="receipt_zone_calculate_button"):
                if current_medium_limit <= current_light_limit:
                    st.error("Исправьте границы веса перед расчётом зон товаров.")
                else:
                    updated_receipts, zone_diag = calculate_receipt_zones(receipts, current_zone_settings)
                    current_zone_settings["settings_hash"] = zone_diag.get("settings_hash", current_settings_hash)
                    current_zone_settings["calculated_at"] = pd.Timestamp.now(tz="UTC").isoformat()
                    state["receipts"] = updated_receipts
                    state["zone_classification_settings"] = current_zone_settings
                    state["zone_classification_diagnostics"] = zone_diag
                    state["diagnostics"] = build_receipt_diagnostics(updated_receipts, len(updated_receipts))
                    save_receipts_state(state)
                    if not update_data_revisions(model, ["receipts"], "calculate_receipt_zones"):
                        return
                    st.success("Зоны товаров рассчитаны. Размещение будет использовать только рассчитанную зону, а исходная зона 1С останется для сравнения.")
                    st.rerun()
            _render_zone_classification_result(state)
            receipts = state.get("receipts", [])
            zone_summary = _receipt_zone_summary(receipts)
            classified = len(receipts) - zone_summary.get("unclassified", 0)
            z1, z2, z3, z4, z5 = st.columns(5)
            z1.metric("С зоной", classified)
            z2.metric("Без зоны", zone_summary.get("unclassified", 0))
            z3.metric("Тяжёлое", zone_summary.get("heavy", 0))
            z4.metric("Среднее/лёгкое", zone_summary.get("medium", 0) + zone_summary.get("light", 0))
            z5.metric("Хрупкое", zone_summary.get("fragile", 0))
            if classified == 0:
                st.error("Зоны товаров ещё не рассчитаны или все SKU без категории. Настройте правила и нажмите «Рассчитать зоны товаров». Исходная зона 1С используется только для сравнения.")
            elif zone_summary.get("unclassified", 0):
                st.warning("У части строк прихода нет рассчитанной зоны. Эти строки не будут размещены автоматически и получат причину missing_calculated_zone.")
            st.caption("Чтобы добавить приход поверх текущего остатка, нажмите «Добавить приход на текущий склад». Операция не очищает фактический переходящий остаток.")
            b1, b3 = st.columns(2)
            b1.download_button("Скачать загруженные приходы", export_receipts_excel_bytes(state), file_name="receipts.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            if b3.button("Добавить приход на текущий склад", key="receipt_calculate_stub", type="primary"):
                if current_medium_limit <= current_light_limit:
                    st.error("Исправьте границы веса и пересчитайте зоны товаров перед размещением.")
                    return
                if classified == 0:
                    st.error("Сначала нажмите «Рассчитать зоны товаров». Размещение не использует исходную зону 1С как fallback.")
                    return
                if saved_settings_hash and saved_settings_hash != current_settings_hash:
                    st.error("Расчёт зон товаров устарел после изменения границ веса. Пересчитайте зоны товаров перед размещением.")
                    return
                placement_state, placement_warning = load_placement_state(model)
                if placement_warning:
                    st.warning(placement_warning)
                model = apply_active_boundaries_to_model(model)
                save_geometry_model(model)
                save_pre_placement_snapshot(model, placement_state, state, trigger="receipt_placement")
                placement_state, basic_diag = calculate_basic_weight_placement(model, placement_state, state)
                st.session_state["geometry_model"] = attach_placements_to_model(model, placement_state)
                st.session_state["placement_state"] = placement_state
                st.session_state["last_receipt_placement_diag"] = basic_diag
                summary = basic_diag or {}
                st.success(
                    f"Приход добавлен на текущий склад: размещено — {summary.get('Размещено паллет', summary.get('Размещено', 0))}, "
                    f"не размещено — {summary.get('Не размещено паллет', summary.get('Не размещено', 0))}."
                )
                with st.expander("Подробная диагностика", expanded=False):
                    _render_receipt_placement_diagnostics(basic_diag)
                st.rerun()
            elif st.session_state.get("last_receipt_placement_diag"):
                with st.expander("Подробная диагностика последнего добавления прихода", expanded=False):
                    _render_receipt_placement_diagnostics(st.session_state.get("last_receipt_placement_diag"))

    with diag_tab:
        st.subheader("Диагностика приходов")
        if diagnostics:
            st.dataframe(pd.DataFrame([{"Показатель": key, "Значение": value} for key, value in diagnostics.items() if key != "messages"]), use_container_width=True)
            messages = diagnostics.get("messages", [])
            if messages:
                st.dataframe(pd.DataFrame(messages), use_container_width=True)
        else:
            st.info("Диагностика появится после загрузки файла приходов.")


def _current_warehouse_state(model: dict) -> None:
    with measure_step("load_placements"):
        placement_state, _ = load_placement_state(model)
    with measure_step("load_receipts"):
        receipts_state, _ = load_receipts_state(model)
    placements = placement_state.get("placements", [])
    unplaced = placement_state.get("unplaced_inventory", [])
    carryover = [item for item in placements if item.get("source") == "inventory_carryover"]
    new_receipts = [item for item in placements if item.get("source") != "inventory_carryover"]
    sku_keys = {item.get("sku_key") for item in placements if item.get("sku_key")}
    occupied_cells = {item.get("cell_key") for item in placements if item.get("cell_key")}
    last_inventory = placement_state.get("last_inventory_reconciliation", {})
    operations = placement_state.get("journal", [])
    last_operation = operations[-1].get("created_at", "—") if operations else "—"
    cols = st.columns(7)
    metrics = [
        ("Последняя операция", last_operation),
        ("Последний инвент", last_inventory.get("inventory_date", "—")),
        ("SKU на складе", len(sku_keys)),
        ("Занято ячеек", len(occupied_cells)),
        ("Переходящий остаток", f"{sum(float(x.get('qty_pallets', 0) or 0) for x in carryover):g}"),
        ("Новый приход", f"{sum(float(x.get('qty_pallets', 0) or 0) for x in new_receipts):g}"),
        ("Не размещено", f"{sum(float(x.get('qty_pallets', 0) or 0) for x in unplaced):g}"),
    ]
    for col, (label, value) in zip(cols, metrics):
        col.metric(label, value)
    if receipts_state.get("receipts"):
        st.caption(f"Загружено строк прихода: {len(receipts_state['receipts'])}.")


def render_outbound_picking(model: dict) -> None:
    with measure_step("load_outbound_state"):
        orders_state = load_outbound_orders(model)
        execution_state = load_outbound_execution_state(model)
        execution_log = load_outbound_execution_log(model)
    rows = orders_state.get("rows", [])
    st.caption("Сборка выполняется последовательно в целых qty_units. Вес, масса и весовые коэффициенты не используются.")
    uploaded = st.file_uploader("Excel с расходными ордерами", type=["xlsx"], key="outbound_orders_upload")
    if uploaded is not None:
        file_bytes = uploaded.getvalue()
        sheet_names = get_outbound_sheet_names(file_bytes)
        sheet_name = st.selectbox("Лист с расходными ордерами", sheet_names, key="outbound_sheet")
        header_rows = st.radio("Строк заголовка РО", [1, 2], horizontal=True, key="outbound_header_rows")
        source_table = read_outbound_table(file_bytes, sheet_name, header_rows)
        detected = detect_outbound_columns(source_table)
        columns = [None, *source_table.columns.tolist()]
        mapping: dict[str, str | None] = {}
        labels = {
            "outbound_order_number": "Номер РО",
            "created_at": "Дата создания",
            "nomenclature": "Номенклатура",
            "characteristic": "Характеристика",
            "qty_units": "Количество юнитов",
            "unit_name": "Единица измерения",
            "warehouse": "Склад",
        }
        mapping_columns = st.columns(4)
        for index, (field, label) in enumerate(labels.items()):
            detected_column = detected.get(field)
            mapping[field] = mapping_columns[index % 4].selectbox(
                label,
                columns,
                index=columns.index(detected_column) if detected_column in columns else 0,
                key=f"outbound_map_{field}",
            )
        normalized_rows, diagnostics = normalize_outbound_table(source_table, mapping)
        with st.expander("Предпросмотр распознанных расходных ордеров"):
            st.dataframe(pd.DataFrame(normalized_rows).head(200), use_container_width=True)
            if diagnostics:
                st.dataframe(pd.DataFrame(diagnostics), use_container_width=True)
        if st.button("Загрузить расходные ордера", type="primary", key="outbound_save_upload"):
            if any(item.get("level") == "error" for item in diagnostics):
                st.error("Исправьте сопоставление обязательных колонок.")
            else:
                orders_state = {
                    "model_id": model.get("model_id"),
                    "loaded_at": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
                    "source_file_name": uploaded.name,
                    "source_file_hash": file_hash(file_bytes),
                    "rows": normalized_rows,
                }
                save_outbound_orders(orders_state)
                if not update_data_revisions(model, ["outbound"], "save_outbound_orders"):
                    return
                st.success(f"Загружено строк РО: {len(normalized_rows)}.")
                st.rerun()

    order_summary = summarize_outbound_orders(rows, execution_state)
    if not order_summary:
        st.info("Расходные ордера пока не загружены.")
        return
    st.markdown("**Расходные ордера**")
    st.dataframe(pd.DataFrame(order_summary), use_container_width=True, hide_index=True)
    available = [item for item in order_summary if not item["processed"]]
    option_to_key = {f"{item['created_at']} · {item['outbound_order_number']} · {item['requested_units']} юн.": item["order_key"] for item in available}
    selected_labels = st.multiselect("Выберите необработанные РО", list(option_to_key), key="outbound_selected_orders")
    selected_keys = [option_to_key[label] for label in selected_labels]

    def execute(keys: list[str]) -> None:
        placement_state, warning = load_placement_state(model)
        if warning:
            st.warning(warning)
        ensure_pre_outbound_snapshot(placement_state)
        updated_placements, updated_execution, updated_log, summary = execute_outbound_orders(
            model,
            placement_state,
            rows,
            execution_state,
            execution_log,
            keys,
        )
        save_placement_state(updated_placements)
        save_outbound_execution_state(updated_execution)
        save_outbound_execution_log(updated_log, model.get("model_id"))
        if not update_data_revisions(model, ["placements", "outbound"], "execute_outbound_orders"):
            return
        st.session_state["placement_state"] = updated_placements
        st.session_state["last_outbound_summary"] = summary
        st.success(f"Собрано юнитов: {summary.get('Собрано юнитов', 0)}; дефицит: {summary.get('Дефицит юнитов', 0)}.")
        st.rerun()

    c1, c2, c3 = st.columns(3)
    if c1.button("Собрать выбранные РО", disabled=not selected_keys, key="outbound_execute_selected"):
        execute(selected_keys)
    all_unprocessed = [item["order_key"] for item in available]
    if c2.button("Собрать все необработанные РО", disabled=not all_unprocessed, key="outbound_execute_all"):
        execute(all_unprocessed)
    if c3.button("Сбросить результаты сборки", disabled=not bool(execution_state.get("processed_orders")), key="outbound_reset"):
        restored, result = reset_outbound_execution(model)
        if result["success"]:
            if not update_data_revisions(model, ["placements", "outbound"], "reset_outbound_execution"):
                return
            st.session_state["placement_state"] = restored
            st.success(result["message"])
            st.rerun()
        else:
            st.error(result["message"])

    execution_state = load_outbound_execution_state(model)
    execution_log = load_outbound_execution_log(model)
    line_results = execution_state.get("line_results", [])
    summary = outbound_execution_summary(rows, execution_state, line_results)
    metric_columns = st.columns(5)
    for index, (label, value) in enumerate(summary.items()):
        metric_columns[index % 5].metric(label, value)
    if line_results:
        with st.expander("Результаты строк РО"):
            st.dataframe(pd.DataFrame(line_results), use_container_width=True)
        shortages = [item for item in line_results if int(item.get("shortage_units", 0) or 0) > 0 or item.get("line_status") not in {"completed"}]
        with st.expander("Дефицитные и отклонённые строки"):
            st.dataframe(pd.DataFrame(shortages), use_container_width=True)
    if execution_log:
        with st.expander("Журнал списаний по ячейкам"):
            st.dataframe(pd.DataFrame(execution_log), use_container_width=True)


@st.fragment
def render_map_geometry_fragment(model: dict) -> None:
    with measure_step("fragment_map_geometry"):
        render_geometry_map_view(model)


@st.fragment
def render_outbound_fragment(model: dict) -> None:
    with measure_step("fragment_map_outbound"):
        render_outbound_picking(model)


@st.fragment
def render_row_settings_fragment(model: dict) -> None:
    with measure_step("fragment_settings_rows"):
        render_unified_row_settings_editor(model)


@st.fragment
def render_cross_aisles_fragment(model: dict) -> None:
    with measure_step("fragment_settings_cross_aisles"):
        render_cross_aisle_settings_editor(model)


@st.fragment
def render_aisles_fragment(model: dict) -> None:
    with measure_step("fragment_settings_aisles"):
        render_active_model_aisle_editor(model)


@st.fragment
def render_zone_boundaries_fragment(model: dict) -> None:
    with measure_step("fragment_settings_zones"):
        render_zone_boundaries_editor(model)


@st.fragment
def render_receipts_fragment(model: dict) -> None:
    with measure_step("fragment_receipts"):
        render_receipts_section(model)


@st.fragment
def render_inventory_fragment(model: dict) -> None:
    with measure_step("fragment_inventory"):
        render_inventory_placement(model)


def render_operation_history(model: dict) -> None:
    placement_state, _ = load_placement_state(model)
    journal = placement_state.get("journal", [])
    if journal:
        st.dataframe(pd.DataFrame(journal), use_container_width=True)
    else:
        st.info("История операций пока пуста.")
    with st.expander("Запросы для выгрузки из 1С"):
        st.caption("Используйте действующие запросы проекта для подготовки файлов прихода и инвентаризации. Формат выгрузки в этой версии интерфейса не изменён.")


@st.fragment
def render_operation_history_fragment(model: dict) -> None:
    with measure_step("fragment_history"):
        render_operation_history(model)


def render_warehouse_map_tab(model: dict | None) -> None:
    if not model:
        st.info("Сначала загрузите схему склада на вкладке «Служебное».")
        return
    labels = {"map": "Карта склада", "outbound": "Моделирование сборки"}
    subsection = st.radio(
        "Подраздел карты склада",
        list(labels),
        format_func=labels.get,
        horizontal=True,
        key="warehouse_map_subsection",
        label_visibility="collapsed",
    )
    if subsection == "map":
        with measure_step("render_subsection_map_geometry"):
            render_map_geometry_fragment(model)
    elif subsection == "outbound":
        with measure_step("render_subsection_map_outbound"):
            render_outbound_fragment(model)


def render_warehouse_settings_tab(model: dict | None) -> None:
    st.subheader("Настройки склада")
    if not model:
        st.info("Сначала загрузите схему склада на вкладке «Служебное».")
        return
    labels = {
        "rows": "Настройки рядов",
        "cross_aisles": "Поперечные проезды",
        "aisles": "Межрядные проезды",
        "zones": "Весовые зоны",
    }
    subsection = st.radio(
        "Подраздел настроек склада",
        list(labels),
        format_func=labels.get,
        horizontal=True,
        key="warehouse_settings_subsection",
        label_visibility="collapsed",
    )
    if subsection == "rows":
        st.caption("Основной способ изменения склада — единый черновик настроек рядов. Карта остаётся режимом просмотра.")
        with measure_step("render_subsection_settings_rows"):
            render_row_settings_fragment(model)
    elif subsection == "cross_aisles":
        with measure_step("render_subsection_settings_cross_aisles"):
            render_cross_aisles_fragment(model)
    elif subsection == "aisles":
        with measure_step("render_subsection_settings_aisles"):
            render_aisles_fragment(st.session_state.get("geometry_model", model))
    elif subsection == "zones":
        with measure_step("render_subsection_settings_zones"):
            render_zone_boundaries_fragment(st.session_state.get("geometry_model", model))


def render_receipts_inventory_tab(model: dict | None) -> None:
    if not model:
        st.info("Для работы с приходами сначала загрузите схему склада на вкладке «Служебное».")
        return
    labels = {"receipts": "Приход", "inventory": "Инвентаризация", "history": "История операций"}
    subsection = st.radio(
        "Подраздел приходов и инвентаризации",
        list(labels),
        format_func=labels.get,
        horizontal=True,
        key="warehouse_receipts_subsection",
        label_visibility="collapsed",
    )
    if subsection == "receipts":
        with measure_step("render_subsection_receipts"):
            render_receipts_fragment(model)
    elif subsection == "inventory":
        with measure_step("render_subsection_inventory"):
            render_inventory_fragment(model)
    elif subsection == "history":
        with measure_step("render_subsection_history"):
            render_operation_history_fragment(model)


def render_analytics_tab(model: dict | None) -> None:
    if not model:
        st.info("Аналитика появится после загрузки модели склада.")
        return
    state, warning = load_placement_state(model)
    if warning:
        st.warning(warning)
    render_placement_diagnostics_section(model, state)
    unplaced = state.get("unplaced_inventory", [])
    st.subheader("Не размещено")
    if unplaced:
        st.dataframe(pd.DataFrame(unplaced), use_container_width=True)
    else:
        st.info("Неразмещённого товара нет.")
    last_report = st.session_state.get("last_inventory_reconciliation_report")
    if last_report:
        st.subheader("Результат последней сверки с инвентом")
        st.dataframe(pd.DataFrame(last_report.get("details", [])), use_container_width=True)


@st.fragment
def render_analytics_fragment(model: dict | None) -> None:
    with measure_step("fragment_analytics"):
        render_analytics_tab(model)


def render_service_tab(saved_model: dict | None, model: dict | None) -> None:
    render_data_revisions(model or saved_model)
    st.subheader("Производительность")
    enabled = st.toggle(
        "Включить измерение производительности",
        value=bool(st.session_state.get("performance_enabled", False)),
        key="performance_enabled",
    )
    if enabled:
        st.caption("Результаты текущего измерения появятся после следующего rerun.")
    else:
        st.caption("Профилирование выключено. Его также можно включить через WAREHOUSE_PERF=1.")
    if st.button("Очистить историю измерений", key="performance_clear_history"):
        clear_performance_history()
        st.success("История измерений очищена.")
    latest_run = load_latest_performance_run()
    if latest_run:
        steps = sorted(latest_run.get("steps", []), key=lambda item: item.get("duration_ms", 0), reverse=True)
        metrics = st.columns(5)
        metrics[0].metric("Тип запуска", "Холодный" if latest_run.get("is_cold_start") else "Повторный")
        metrics[1].metric("Общее время", f"{float(latest_run.get('total_duration_ms') or 0):.1f} мс")
        metrics[2].metric("Пиковая память", f"{float(latest_run.get('peak_traced_memory_mb') or 0):.1f} МБ")
        metrics[3].metric("Статус", latest_run.get("status", "—"))
        metrics[4].metric("Этапов", len(steps))
        st.caption("Пять самых медленных этапов: " + ", ".join(item.get("name", "—") for item in steps[:5]))
        total_ms = float(latest_run.get("total_duration_ms") or 0)
        rows = [{
            "Этап": item.get("name"),
            "Время, мс": round(float(item.get("duration_ms") or 0), 3),
            "Доля общего времени, %": round(float(item.get("duration_ms") or 0) * 100 / total_ms, 2) if total_ms else 0,
            "Количество вызовов": item.get("call_count", 1),
            "Метаданные": json.dumps(item.get("metadata") or {}, ensure_ascii=False),
        } for item in steps]
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
    else:
        st.info("Сохранённых измерений пока нет.")

    st.subheader("Загрузка и замена схемы склада")
    render_geometry_constructor_tab(saved_model, show_active_model=False)
    model = st.session_state.get("geometry_model", model)
    if model:
        with st.expander("Ручное редактирование и технические данные"):
            render_manual_cell_editor(model)
            render_geometry_data_tabs(st.session_state.get("geometry_model", model))
        with st.expander("Экспериментальное редактирование геометрии", expanded=False):
            st.warning("Ручной сдвиг может нарушить геометрию проездов и будущий расчёт маршрутов")
            row_choices = _row_options(model)
            selected_labels = st.multiselect("Ряды для сдвига", list(row_choices), key="service_geometry_shift_rows")
            selected_rows = [row_choices[label] for label in selected_labels]
            g1, g2, g3 = st.columns(3)
            dx = g1.number_input("Сдвиг X, м", value=0.0, step=0.1, key="service_geometry_shift_x")
            dy = g2.number_input("Сдвиг Y, м", value=0.0, step=0.1, key="service_geometry_shift_y")
            snap = g3.checkbox("Привязать к шагу 0,1 м", value=True, key="service_geometry_shift_snap")
            if st.button("Применить экспериментальный сдвиг", disabled=not selected_rows, key="service_geometry_shift_apply"):
                _save_map_edit_snapshot(model)
                ok, message = _shift_rows(model, selected_rows, dx, dy, snap, 0.1)
                if ok:
                    _persist_map_edit(model, message)
                    st.rerun()
                else:
                    st.error(message)
        st.download_button(
            "Экспортировать модель JSON",
            json.dumps(model, ensure_ascii=False, indent=2).encode("utf-8"),
            file_name="warehouse_model.json",
            mime="application/json",
            key="service_export_model_json",
        )
    with st.expander("Опасные действия", expanded=False):
        st.warning("Эти действия изменяют сохранённое состояние. Перед выполнением проверьте выбранное подтверждение.")
        if st.button("Сбросить рассчитанное размещение", key="placement_clear_all"):
            state, warning = load_placement_state(model or {})
            if warning:
                st.warning(warning)
            state = clear_calculated_placements(state)
            st.session_state["placement_state"] = state
            st.success("Рассчитанный приход очищен. Фактический переходящий остаток сохранён.")
        if st.button("Очистить журнал размещения", key="placement_clear_journal"):
            state, _ = load_placement_state(model or {})
            state["journal"] = []
            save_placement_state(state)
            st.success("Журнал размещения очищен.")
        if st.button("Очистить загруженный приход", key="receipt_clear_button"):
            receipts_state, _ = load_receipts_state(model or {})
            if receipts_state.get("receipts"):
                clear_receipts_state()
                if update_data_revisions(model, ["receipts"], "clear_receipts"):
                    st.success("Загруженный приход очищен.")
            else:
                st.info("Загруженных приходов нет — изменения не требуются.")
        reset_confirm = st.checkbox("Подтверждаю полный сброс проекта", key="service_full_reset_confirm")
        if st.button("Полный сброс проекта", disabled=not reset_confirm, key="geometry_clear_saved"):
            for path in [MODEL_PATH, META_PATH]:
                if path.exists():
                    path.unlink()
            clear_manual_overrides()
            clear_row_settings()
            clear_placement_state()
            clear_receipts_state()
            st.session_state.pop("geometry_model", None)
            st.session_state.pop("placement_state", None)
            st.success("Проект полностью сброшен.")
            st.rerun()


def render_excel_geometry_warehouse() -> None:
    st.title("Симулятор скорости сборки")
    st.caption("Рабочий процесс: настройте склад на карте, добавьте приход, зафиксируйте инвент и проверьте результат в аналитике.")
    with measure_step("load_geometry_model"):
        saved_model = load_geometry_model()
    if saved_model and "geometry_model" not in st.session_state:
        st.session_state["geometry_model"] = saved_model
    model = st.session_state.get("geometry_model")
    section_labels = {
        "map": "Карта склада",
        "settings": "Настройки склада",
        "receipts_inventory": "Приходы и инвент",
        "analytics": "Аналитика",
        "service": "Служебное",
    }
    active_section = st.radio(
        "Раздел склада",
        options=list(section_labels),
        format_func=section_labels.__getitem__,
        horizontal=True,
        key="warehouse_active_section",
    )

    if active_section == "map":
        with measure_step("render_section_map"):
            render_warehouse_map_tab(model)
    elif active_section == "settings":
        with measure_step("render_section_settings"):
            render_warehouse_settings_tab(model)
    elif active_section == "receipts_inventory":
        with measure_step("render_section_receipts_inventory"):
            render_receipts_inventory_tab(model)
    elif active_section == "analytics":
        with measure_step("render_section_analytics"):
            render_analytics_fragment(model)
    elif active_section == "service":
        with measure_step("render_section_service"):
            render_service_tab(saved_model, model)


def render_geometry_constructor_tab(saved_model: dict | None, *, show_active_model: bool = False) -> None:
    st.caption("Загрузите Excel со схемой склада и выберите лист с ячейками. После построения модель сохранится и будет доступна на вкладке «Карта склада».")
    uploaded = st.file_uploader("Excel со списком фактических ячеек", type=["xlsx"], key="geometry_cells_file")
    if uploaded is None:
        if saved_model:
            st.info("Загрузите новый Excel или используйте сохранённую модель из боковой панели.")
        else:
            st.info("Загрузите Excel со списком ячеек в формате: Код | Ряд | Ячейка | Ярус.")
        model = st.session_state.get("geometry_model")
        if model and show_active_model:
            render_geometry_constructor_view(model)
        return

    file_bytes = uploaded.getvalue()
    content_hash = file_hash(file_bytes)
    sheet_names = get_geometry_sheet_names(file_bytes)
    sheet_name = st.selectbox("Лист со списком ячеек", sheet_names, key="geometry_sheet")
    st.caption("Выберите лист, где находятся строки с кодом, рядом, ячейкой и ярусом. Предпросмотр ниже поможет проверить, что выбран правильный лист.")
    header_rows = st.radio("Строк заголовка", [1, 2], index=1, horizontal=True, help="Если в Excel сверху 'Ряд', а ниже 'Ссылка', выберите 2 строки заголовка.")

    timings: dict[str, float] = {}
    started = perf_counter()
    df = read_cell_table_cached(file_bytes, content_hash, sheet_name, header_rows)
    timings["read_excel_seconds"] = perf_counter() - started
    st.caption(f"Прочитано строк: {len(df)}; колонок: {len(df.columns)}")
    with st.expander("Предпросмотр первых строк", expanded=False):
        st.dataframe(df.head(30), use_container_width=True)

    detected = detect_column_mapping(df)
    st.subheader("Колонки")
    st.caption("Проверьте автоопределение колонок или выберите их вручную. Эти настройки используются только для чтения текущего Excel.")
    columns = [None] + list(df.columns)
    c1, c2, c3, c4 = st.columns(4)
    mapping = {
        "code": c1.selectbox("Код", columns, index=columns.index(detected["code"]) if detected["code"] in columns else 0),
        "row_number": c2.selectbox("Ряд", columns, index=columns.index(detected["row_number"]) if detected["row_number"] in columns else 0),
        "cell_number": c3.selectbox("Ячейка", columns, index=columns.index(detected["cell_number"]) if detected["cell_number"] in columns else 0),
        "tier": c4.selectbox("Ярус", columns, index=columns.index(detected["tier"]) if detected["tier"] in columns else 0),
    }

    norm_started = perf_counter()
    normalized_df, column_diagnostics = normalize_cell_table_cached(df.to_json(orient="split", force_ascii=False), json.dumps(mapping, ensure_ascii=False))
    timings["normalize_columns_seconds"] = perf_counter() - norm_started
    if any(item.get("level") == "error" for item in column_diagnostics):
        st.error("Не удалось определить обязательные колонки. Выберите 'Ряд' и 'Ячейка' вручную.")
        st.dataframe(pd.DataFrame(column_diagnostics), use_container_width=True)
        return

    st.subheader("Размеры и ярусы")
    st.caption("Укажите размеры ячеек, проездов и ярусы для построения склада. Нажатие «Построить склад» пересчитает геометрию по этим параметрам.")
    s1, s2, s3, s4, s5 = st.columns(5)
    cell_length_m = s1.number_input("Длина ячейки вдоль ряда, м", min_value=0.1, value=1.2, step=0.1)
    cell_width_m = s2.number_input("Ширина ряда, м", min_value=0.1, value=0.8, step=0.1)
    aisle_width_m = s3.number_input("Межрядный проезд, м", min_value=0.1, value=3.4, step=0.1)
    top_road_width_m = s4.number_input("Верхний проезд, м", min_value=0.1, value=3.4, step=0.1)
    bottom_road_width_m = s5.number_input("Нижний проезд, м", min_value=0.1, value=3.4, step=0.1)
    tier_values = sorted(normalized_df["tier"].dropna().astype(str).unique().tolist(), key=lambda value: (not value.isdigit(), value)) or ["1"]
    tier_mode = st.radio("Ярусы", ["selected", "all", "min_per_cell"], format_func={"selected": "только выбранный", "all": "все", "min_per_cell": "минимальный для ряд+ячейка"}.get, horizontal=True)
    selected_tier = st.selectbox("Выбранный ярус", tier_values, disabled=tier_mode != "selected")

    row_config_default = default_row_config(normalized_df)
    if st.session_state.get("geometry_row_config_hash") != content_hash:
        st.session_state["geometry_row_config_data"] = row_config_default
        st.session_state["geometry_row_config_hash"] = content_hash

    st.subheader("Настройки рядов")
    st.caption("Настройте порядок, направление и тип хранения каждого ряда. Эти значения попадут в модель после построения склада.")
    st.caption("Обычный ряд хранит одну паллету на системную ячейку. Набивной ряд хранит несколько физических паллетомест внутри одной системной ячейки.")
    row_config_source = st.session_state.get("geometry_row_config_data", row_config_default)
    row_config_display = row_config_source.copy()
    row_config_display["row_storage_type"] = row_config_display["row_storage_type"].map({"normal": "Обычный ряд", "deep_lane": "Набивной ряд"}).fillna(row_config_display["row_storage_type"])
    row_config_display["cell_direction"] = row_config_display["cell_direction"].map({"bottom_to_top": "Снизу вверх", "top_to_bottom": "Сверху вниз"}).fillna(row_config_display["cell_direction"])
    if "weight_zone" not in row_config_display.columns:
        row_config_display["weight_zone"] = "unassigned"
    row_config_display["weight_zone"] = row_config_display["weight_zone"].map({"heavy": "Тяжёлое", "medium": "Среднее", "light": "Лёгкое", "fragile": "Хрупкое", "unassigned": "Не назначено"}).fillna("Не назначено")
    row_config = st.data_editor(
        row_config_display,
        num_rows="dynamic",
        use_container_width=True,
        key="geometry_row_config",
        column_config={
            "row_number": "Ряд",
            "row_order": "Порядок ряда",
            "row_storage_type": st.column_config.SelectboxColumn("Тип ряда", options=["Обычный ряд", "Набивной ряд"]),
            "deep_lane_width": st.column_config.NumberColumn("Набивных паллетомест", min_value=1, max_value=7, step=1),
            "cell_direction": st.column_config.SelectboxColumn("Направление ячеек", options=["Снизу вверх", "Сверху вниз"]),
            "weight_zone": st.column_config.SelectboxColumn("Весовая зона", options=["Тяжёлое", "Среднее", "Лёгкое", "Хрупкое", "Не назначено"]),
            "top_offset_cells": st.column_config.NumberColumn("Отступ сверху, ячеек", min_value=0, step=1),
            "bottom_offset_cells": st.column_config.NumberColumn("Отступ снизу, ячеек", min_value=0, step=1),
            "row_group": "Группа рядов",
            "side": "Сторона/зона",
            "comment": "Комментарий",
        },
    )
    row_config["row_storage_type"] = row_config["row_storage_type"].map({"Обычный ряд": "normal", "Набивной ряд": "deep_lane"}).fillna(row_config["row_storage_type"])
    row_config["cell_direction"] = row_config["cell_direction"].map({"Снизу вверх": "bottom_to_top", "Сверху вниз": "top_to_bottom"}).fillna(row_config["cell_direction"])
    if "weight_zone" not in row_config.columns:
        row_config["weight_zone"] = "Не назначено"
    row_config["weight_zone"] = row_config["weight_zone"].map({"Тяжёлое": "heavy", "Среднее": "medium", "Лёгкое": "light", "Хрупкое": "fragile", "Не назначено": "unassigned"}).fillna("unassigned")
    st.session_state["geometry_row_config_data"] = row_config

    st.subheader("Проезды между рядами")
    st.caption("Если пары «ряд от → ряд до» нет в таблице, ряды стоят плотно. Если есть — между ними добавляется проезд.")
    aisle_config = st.data_editor(
        empty_aisle_config(),
        num_rows="dynamic",
        use_container_width=True,
        key="geometry_aisle_config",
        column_config={
            "row_from": "Ряд от",
            "row_to": "Ряд до",
            "aisle_width_m": "Ширина проезда, м",
            "aisle_type": "Тип проезда",
            "comment": "Комментарий",
        },
    )

    build_clicked = st.button("Построить склад", type="primary", key="geometry_build")
    if build_clicked:
        settings = GeometrySettings(
            cell_length_m=cell_length_m,
            cell_width_m=cell_width_m,
            aisle_width_m=aisle_width_m,
            top_road_width_m=top_road_width_m,
            bottom_road_width_m=bottom_road_width_m,
            selected_tier=str(selected_tier),
            tier_mode=tier_mode,
        )
        build_started = perf_counter()
        model, diagnostics = build_geometry_model(normalized_df, settings, row_config, aisle_config, uploaded.name, sheet_name, source_file_hash=content_hash)
        timings["build_geometry_seconds"] = perf_counter() - build_started
        model["performance"] = timings | model.get("performance", {})
        clear_manual_overrides()
        clear_row_settings()
        clear_placement_state()
        st.session_state.pop("placement_state", None)
        st.warning("Загружен новый Excel. Старые ручные изменения, настройки набивных рядов и размещение товара сброшены для новой модели.")
        save_started = perf_counter()
        with measure_step("save_geometry_model"):
            save_geometry_model(model)
        timings["save_model_seconds"] = perf_counter() - save_started
        model["performance"] = timings | model.get("performance", {})
        with measure_step("save_geometry_model"):
            save_geometry_model(model)
        st.session_state["geometry_model"] = model
        st.success(f"Геометрическая модель построена и сохранена: {len(model['rows'])} рядов, {len(model['cells'])} ячеек, {len(model['aisles'])} проездов.")

    model = st.session_state.get("geometry_model")
    if model and show_active_model:
        render_geometry_constructor_view(model)



RUSSIAN_COLUMN_LABELS = {
    "code": "Код",
    "row_number": "Ряд",
    "cell_number": "Ячейка",
    "tier": "Ярус",
    "source": "Источник",
    "source_line": "Строка источника",
    "storage_type": "Тип хранения",
    "row_storage_type": "Тип ряда",
    "deep_lane_width": "Набивных паллетомест",
    "capacity_pallets": "Вместимость, паллет",
    "volume_m3": "Объём, м³",
    "cell_direction": "Направление ячеек",
    "physical_slots": "Физические места",
    "row_order": "Порядок",
    "row_group": "Группа",
    "side": "Сторона/зона",
    "comment": "Комментарий",
    "cells_count": "Количество ячеек",
    "row_from": "Ряд от",
    "row_to": "Ряд до",
    "aisle_width_m": "Ширина проезда, м",
    "aisle_type": "Тип проезда",
    "road_type": "Тип дороги",
    "width_m": "Ширина, м",
    "length_m": "Длина, м",
    "node_id": "Узел",
    "node_type": "Тип узла",
    "from_node": "От узла",
    "to_node": "К узлу",
    "distance_m": "Расстояние, м",
    "edge_type": "Тип связи",
    "x_min": "X от",
    "x_max": "X до",
    "y_min": "Y от",
    "y_max": "Y до",
    "x_center": "X центр",
    "y_center": "Y центр",
    "top_offset_cells": "Отступ сверху, ячеек",
    "bottom_offset_cells": "Отступ снизу, ячеек",
    "top_offset_m": "Отступ сверху, м",
    "bottom_offset_m": "Отступ снизу, м",
}


def _localized_dataframe(rows: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    if "source" in df.columns:
        df["source"] = df["source"].map(_source_label)
    return df.rename(columns=RUSSIAN_COLUMN_LABELS)

def _model_aisle_config_dataframe(model: dict) -> pd.DataFrame:
    rows = []
    for aisle in model.get("aisles", []):
        rows.append({
            "row_from": aisle.get("row_from", ""),
            "row_to": aisle.get("row_to", ""),
            "aisle_width_m": aisle.get("aisle_width_m", model.get("settings", {}).get("aisle_width_m", 3.4)),
            "aisle_type": aisle.get("aisle_type", "межрядный проезд"),
            "comment": aisle.get("comment", ""),
        })
    return pd.DataFrame(rows, columns=["row_from", "row_to", "aisle_width_m", "aisle_type", "comment"])



def _boundary_rows(boundaries: dict, zone: str) -> str:
    boundary = (boundaries or {}).get(zone, {})
    start = boundary.get("start_row") or "—"
    end = boundary.get("end_row") or "—"
    return f"{start}–{end}" if start != "—" or end != "—" else "—"


def _boundary_table(model: dict, boundaries: dict) -> pd.DataFrame:
    rows = []
    for zone in ZONE_ORDER:
        boundary = (boundaries or {}).get(zone, {})
        rows.append({
            "Зона": ZONE_LABELS.get(zone, zone),
            "Начальный ряд": boundary.get("start_row", ""),
            "Конечный ряд": boundary.get("end_row", ""),
            "Количество рядов": boundary.get("row_count", 0),
            "Вместимость": boundary.get("capacity", 0),
        })
    return pd.DataFrame(rows)


def _calculated_boundary_table(model: dict, settings: dict) -> pd.DataFrame:
    base = settings.get("base_zone_boundaries", {})
    calculated = settings.get("calculated_zone_boundaries", {})
    details = settings.get("calculated_zone_diagnostics", {}).get("details", {})
    rows = []
    for zone in ZONE_ORDER:
        base_count = int((base.get(zone, {}) or {}).get("row_count", 0) or 0)
        calc_count = int((calculated.get(zone, {}) or {}).get("row_count", 0) or 0)
        detail = details.get(zone, {})
        rows.append({
            "Зона": ZONE_LABELS.get(zone, zone),
            "Базовые границы": _boundary_rows(base, zone),
            "Рассчитанные границы": _boundary_rows(calculated, zone),
            "Сдвиг в рядах": calc_count - base_count,
            "Потребность, паллет": detail.get("receipt_required_pallets", 0),
            "Фактически занято, паллет": detail.get("factual_occupied_pallets", 0),
            "Вместимость": detail.get("capacity", (calculated.get(zone, {}) or {}).get("capacity", 0)),
            "Резерв, %": detail.get("reserve_percent", settings.get("zone_reserve_percent", 0)),
            "Дефицит": detail.get("deficit", 0),
            "Статус": detail.get("status", "—"),
        })
    return pd.DataFrame(rows)


def render_zone_boundaries_editor(model: dict) -> dict:
    st.subheader("Границы зон размещения")
    st.caption("Базовые границы задаются вручную. Система может временно расширять и сужать зоны под состав прихода, сохраняя порядок зон и учитывая фактическую занятость склада.")
    settings = ensure_zone_boundary_settings(model)
    c1, c2, c3, c4, c5 = st.columns(5)
    settings["zone_reserve_percent"] = c1.number_input("Резерв зоны, %", min_value=0.0, max_value=100.0, value=float(settings.get("zone_reserve_percent", 0) or 0), step=1.0, key="zone_reserve_percent")
    minimum_rows = settings.setdefault("minimum_rows", {zone: 1 for zone in ZONE_ORDER})
    minimum_rows["heavy"] = int(c2.number_input("Мин. рядов: тяжёлое", min_value=0, value=int(minimum_rows.get("heavy", 1)), step=1, key="min_rows_heavy"))
    minimum_rows["medium"] = int(c3.number_input("Мин. рядов: среднее", min_value=0, value=int(minimum_rows.get("medium", 1)), step=1, key="min_rows_medium"))
    minimum_rows["light"] = int(c4.number_input("Мин. рядов: лёгкое", min_value=0, value=int(minimum_rows.get("light", 1)), step=1, key="min_rows_light"))
    minimum_rows["fragile"] = int(c5.number_input("Мин. рядов: хрупкое", min_value=0, value=int(minimum_rows.get("fragile", 1)), step=1, key="min_rows_fragile"))

    left, right = st.columns(2)
    with left:
        st.markdown("**Базовые границы**")
        st.dataframe(_boundary_table(model, settings.get("base_zone_boundaries", {})), use_container_width=True, hide_index=True)
    with right:
        st.markdown("**Расчёт под текущий приход**")
        st.dataframe(_calculated_boundary_table(model, settings), use_container_width=True, hide_index=True)

    b1, b2, b3 = st.columns(3)
    if b1.button("Рассчитать границы под приход", key="calculate_zone_boundaries"):
        receipts_state, receipts_warning = load_receipts_state(model)
        if receipts_warning:
            st.warning(receipts_warning)
        placement_state, placement_warning = load_placement_state(model)
        if placement_warning:
            st.warning(placement_warning)
        calculated, diagnostics = calculate_dynamic_zone_boundaries(model, receipts_state, placement_state)
        settings["calculated_zone_boundaries"] = calculated
        settings["calculated_zone_diagnostics"] = diagnostics
        save_geometry_model(model)
        st.session_state["geometry_model"] = model
        st.success("Расчётные границы построены. Нажмите «Применить рассчитанные границы», чтобы использовать их при размещении.")
        st.rerun()
    if b2.button("Применить рассчитанные границы", key="apply_calculated_zone_boundaries", disabled=not bool(settings.get("calculated_zone_boundaries"))):
        settings["active_zone_boundaries"] = settings.get("calculated_zone_boundaries", {})
        model = apply_active_boundaries_to_model(model, settings["active_zone_boundaries"])
        save_geometry_model(model)
        st.session_state["geometry_model"] = model
        st.success("Рассчитанные границы применены. Базовые границы не изменены.")
        st.rerun()
    if b3.button("Вернуть базовые границы", key="restore_base_zone_boundaries"):
        settings["active_zone_boundaries"] = settings.get("base_zone_boundaries", {})
        model = apply_active_boundaries_to_model(model, settings["active_zone_boundaries"])
        save_geometry_model(model)
        st.session_state["geometry_model"] = model
        st.success("Активные границы возвращены к базовым.")
        st.rerun()
    active = settings.get("active_zone_boundaries", {})
    if active:
        st.caption("Активные границы: " + "; ".join(f"{ZONE_LABELS.get(zone, zone)} { _boundary_rows(active, zone) }" for zone in ZONE_ORDER))
    return model


ROW_SETTINGS_COLUMNS = {
    "row_number": "Номер ряда",
    "row_order": "Порядок ряда",
    "cell_direction": "Направление сборки",
    "weight_zone": "Весовая зона",
    "row_storage_type": "Тип ряда",
    "cell_capacity_pallets": "Вместимость одной логической ячейки",
    "cells_count": "Количество логических ячеек",
    "row_capacity_pallets": "Общая вместимость ряда",
    "top_offset_cells": "Отступ сверху, ячеек",
    "bottom_offset_cells": "Отступ снизу, ячеек",
    "row_group": "Группа ряда",
    "side": "Сторона / зона",
    "comment": "Комментарий",
}
ROW_SETTINGS_REVERSE_COLUMNS = {label: key for key, label in ROW_SETTINGS_COLUMNS.items()}


def _row_settings_to_display(df: pd.DataFrame) -> pd.DataFrame:
    display_df = df.copy()
    display_df["cell_direction"] = display_df["cell_direction"].map(DIRECTION_LABELS).fillna(display_df["cell_direction"])
    display_df["weight_zone"] = display_df["weight_zone"].map(WEIGHT_ZONE_LABELS).fillna(display_df["weight_zone"])
    display_df["row_storage_type"] = display_df["row_storage_type"].map(ROW_STORAGE_TYPE_LABELS).fillna(display_df["row_storage_type"])
    return display_df.rename(columns=ROW_SETTINGS_COLUMNS)


def _row_settings_from_display(display_df: pd.DataFrame) -> pd.DataFrame:
    df = display_df.rename(columns=ROW_SETTINGS_REVERSE_COLUMNS).copy()
    df["cell_direction"] = df["cell_direction"].map(DIRECTION_LABEL_TO_VALUE).fillna(df["cell_direction"])
    df["weight_zone"] = df["weight_zone"].map(WEIGHT_ZONE_LABEL_TO_VALUE).fillna(df["weight_zone"])
    df["row_storage_type"] = df["row_storage_type"].map({label: value for value, label in ROW_STORAGE_TYPE_LABELS.items()}).fillna(df["row_storage_type"])
    for column in ["row_order", "cell_capacity_pallets", "cells_count", "row_capacity_pallets", "top_offset_cells", "bottom_offset_cells"]:
        if column in df.columns:
            df[column] = pd.to_numeric(df[column], errors="coerce").fillna(0)
    return df


def _merge_row_settings_edits(draft: pd.DataFrame, edited_display: pd.DataFrame) -> pd.DataFrame:
    edited = _row_settings_from_display(edited_display)
    result = draft.copy()
    editable = ["row_order", "cell_direction", "weight_zone", "row_storage_type", "cell_capacity_pallets", "top_offset_cells", "bottom_offset_cells", "row_group", "side", "comment"]
    for _, row in edited.iterrows():
        mask = result["row_number"].astype(str) == str(row.get("row_number"))
        for column in editable:
            if column in row:
                result.loc[mask, column] = row.get(column)
    result["cell_capacity_pallets"] = result.apply(lambda row: max(1, int(float(row["cell_capacity_pallets"] or 1))) if row["row_storage_type"] == "deep_lane" else 1, axis=1)
    result["row_capacity_pallets"] = result["cells_count"].astype(float) * result["cell_capacity_pallets"].astype(float)
    return result


def render_unified_row_settings_editor(model: dict) -> dict:
    st.subheader("Настройки рядов")
    st.caption("Измените несколько строк и примените их одной транзакцией. До нажатия кнопки модель, карта и файлы не изменяются.")
    model_id = str(model.get("model_id") or model.get("source_file_hash") or "active")
    state_key = "row_settings_state"
    state = st.session_state.get(state_key)
    if not isinstance(state, dict) or state.get("model_id") != model_id:
        state = create_row_settings_state(model)
        st.session_state[state_key] = state
    draft = pd.DataFrame(copy.deepcopy(state.get("draft", [])))
    changed_rows = changed_row_numbers(state)
    if changed_rows:
        st.warning(f"Есть несохранённые изменения. Изменено рядов: {len(changed_rows)}")
    else:
        st.caption("Несохранённых изменений нет. Правки внутри формы не запускают пересчёт карты.")
    apply_state = st.session_state.get("apply_state", {})
    if apply_state.get("status") == "cancelled":
        st.info(apply_state.get("message", "Черновик восстановлен."))
        st.session_state["apply_state"] = {}
    if st.session_state.get("row_settings_last_changes"):
        st.success("Последние применённые изменения рядов:")
        st.write(st.session_state.pop("row_settings_last_changes"))

    revision = int(state.get("editor_revision", 0))
    compact_columns = {
        "Номер ряда": "Ряд",
        "Порядок ряда": "Порядок",
        "Направление сборки": "Направление",
        "Весовая зона": "Зона",
        "Вместимость одной логической ячейки": "Вместимость",
    }
    with st.form(f"unified_row_settings_form_{model_id}_{revision}"):
        edited_display = st.data_editor(
            _row_settings_to_display(draft).rename(columns=compact_columns),
            use_container_width=True,
            hide_index=True,
            key=f"unified_row_settings_editor_{model_id}_{revision}",
            disabled=["Ряд"],
            column_config={
                "Ряд": st.column_config.TextColumn("Ряд", disabled=True),
                "Порядок": st.column_config.NumberColumn("Порядок", step=1),
                "Направление": st.column_config.SelectboxColumn("Направление", options=list(DIRECTION_LABELS.values())),
                "Зона": st.column_config.SelectboxColumn("Зона", options=list(WEIGHT_ZONE_LABELS.values())),
                "Тип ряда": st.column_config.SelectboxColumn("Тип ряда", options=list(ROW_STORAGE_TYPE_LABELS.values())),
                "Вместимость": st.column_config.NumberColumn("Вместимость", min_value=1, step=1),
                "Отступ сверху, ячеек": st.column_config.NumberColumn("Отступ сверху, ячеек", min_value=0, step=1),
                "Отступ снизу, ячеек": st.column_config.NumberColumn("Отступ снизу, ячеек", min_value=0, step=1),
            },
            column_order=["Ряд", "Порядок", "Направление", "Зона", "Тип ряда", "Вместимость", "Отступ сверху, ячеек", "Отступ снизу, ячеек"],
        )
        cancel_column, apply_column = st.columns(2)
        reset_submit = cancel_column.form_submit_button("Отменить изменения")
        apply_submit = apply_column.form_submit_button("Применить изменения", type="primary")

    updated_draft = _merge_row_settings_edits(
        draft,
        edited_display.rename(columns={value: key for key, value in compact_columns.items()}),
    )
    submitted_state = update_row_settings_state(state, updated_draft.to_dict(orient="records"))
    if reset_submit:
        st.session_state[state_key] = reset_row_settings_state(state)
        st.session_state["apply_state"] = {"status": "cancelled", "message": "Черновик восстановлен из текущей модели."}
        st.info("Черновик восстановлен из текущей модели.")
        return model
    if apply_submit:
        st.session_state[state_key] = submitted_state
        if not changed_row_numbers(submitted_state):
            st.info("Изменений нет")
            return model
        edited_rows = submitted_state["draft"]
        with measure_step("apply_row_settings"):
            updated_model, messages = apply_row_settings_transaction(model, edited_rows)
        if any(str(message).startswith("Ошибка:") for message in messages):
            st.error("Изменения рядов не применены. Модель полностью оставлена без изменений.")
            st.write(messages)
            st.session_state["apply_state"] = {"status": "error", "messages": messages}
            return model
        set_base_boundaries_from_current_rows(updated_model)
        with measure_step("save_geometry_model"):
            save_geometry_model(updated_model)
        if not update_data_revisions(model, ["geometry"], "apply_row_settings"):
            return model
        invalidate_geometry_render_cache()
        st.session_state["geometry_model"] = updated_model
        st.session_state[state_key] = create_row_settings_state(updated_model)
        st.session_state["apply_state"] = {"status": "applied", "messages": messages}
        st.session_state["row_settings_last_changes"] = messages
        st.rerun()
    return model

def render_cross_aisle_settings_editor(model: dict) -> dict:
    st.subheader("Поперечные проезды внутри рядов")
    st.caption("Таблица является отдельным черновиком. Карта и модель перестраиваются только после применения.")
    state_key = "cross_aisle_settings_state"
    model_id = str(model.get("model_id") or model.get("source_file_hash") or "active")
    state = st.session_state.get(state_key)
    if not isinstance(state, dict) or state.get("model_id") != model_id:
        state = create_cross_aisle_settings_state(model)
        st.session_state[state_key] = state
    changed = changed_cross_aisle_count(state)
    if changed:
        st.warning(f"Есть несохранённые изменения. Изменено записей: {changed}")
    else:
        st.caption("Несохранённых изменений нет. Редактирование строк не запускает пересчёт карты.")
    revision = int(state.get("editor_revision", 0))
    draft = pd.DataFrame(
        copy.deepcopy(state.get("draft", [])),
        columns=["row_number", "after_cell_number", "width_cells", "width_m", "comment"],
    )
    for column in ("row_number", "after_cell_number", "comment"):
        draft[column] = draft[column].fillna("").astype("string")
    draft["width_cells"] = pd.to_numeric(draft["width_cells"], errors="coerce").astype("Int64")
    cell_length = float((model.get("settings") or {}).get("cell_length_m", 1.0) or 1.0)
    draft["width_m"] = (draft["width_cells"].astype("Float64") * cell_length).astype("Float64")
    display = draft.rename(columns={"row_number": "Ряд", "after_cell_number": "После ячейки", "width_cells": "Ширина, ячеек", "width_m": "Ширина, м", "comment": "Комментарий"})
    with st.form(f"cross_aisle_form_{model_id}_{revision}"):
        edited = st.data_editor(
            display,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            key=f"cross_aisle_editor_{model_id}_{revision}",
            disabled=["Ширина, м"],
            column_config={
                "Ряд": st.column_config.TextColumn("Ряд"),
                "После ячейки": st.column_config.TextColumn("После ячейки"),
                "Ширина, ячеек": st.column_config.NumberColumn("Ширина, ячеек", min_value=1, step=1),
                "Ширина, м": st.column_config.NumberColumn("Ширина, м", disabled=True, format="%.2f"),
                "Комментарий": st.column_config.TextColumn("Комментарий"),
            },
        )
        cancel_col, apply_col = st.columns(2)
        cancel = cancel_col.form_submit_button("Отменить изменения проездов")
        apply = apply_col.form_submit_button("Применить поперечные проезды", type="primary")
    records = edited.rename(columns={"Ряд": "row_number", "После ячейки": "after_cell_number", "Ширина, ячеек": "width_cells", "Ширина, м": "width_m", "Комментарий": "comment"}).to_dict(orient="records")
    submitted = update_cross_aisle_settings_state(state, records)
    if cancel:
        st.session_state[state_key] = reset_cross_aisle_settings_state(state)
        st.info("Черновик восстановлен из текущей модели.")
        return model
    if apply:
        st.session_state[state_key] = submitted
        if not changed_cross_aisle_count(submitted):
            st.info("Изменений нет")
            return model
        with measure_step("apply_cross_aisles"):
            updated_model, errors = apply_cross_aisles_transaction(model, records)
        if errors:
            st.error("Поперечные проезды не применены. Исправьте все ошибки:")
            for error in errors:
                st.error(error)
            return model
        with measure_step("save_geometry_model"):
            save_geometry_model(updated_model)
        if not update_data_revisions(model, ["geometry"], "apply_cross_aisles"):
            return model
        invalidate_geometry_render_cache()
        st.session_state["geometry_model"] = updated_model
        st.session_state[state_key] = create_cross_aisle_settings_state(updated_model)
        st.rerun()
    return model


def render_active_model_aisle_editor(model: dict) -> dict:
    st.subheader("Настройки проездов между рядами")
    st.caption("Изменение проездов перестраивает только геометрию активной модели по текущим ячейкам и не очищает ручные изменения, размещение товара или приходы.")
    settings = model.get("settings", {})
    model_key = model.get("model_id", "model")
    c1, c2, c3 = st.columns(3)
    default_aisle_width = c1.number_input(
        "Межрядный проезд по умолчанию, м",
        min_value=0.1,
        value=float(settings.get("aisle_width_m", 3.4) or 3.4),
        step=0.1,
        key=f"active_aisle_default_width_{model_key}",
    )
    top_road_width = c2.number_input(
        "Верхний проезд, м",
        min_value=0.1,
        value=float(settings.get("top_road_width_m", 3.4) or 3.4),
        step=0.1,
        key=f"active_top_road_width_{model_key}",
    )
    bottom_road_width = c3.number_input(
        "Нижний проезд, м",
        min_value=0.1,
        value=float(settings.get("bottom_road_width_m", 3.4) or 3.4),
        step=0.1,
        key=f"active_bottom_road_width_{model_key}",
    )
    st.caption("Если пары «ряд от → ряд до» нет в таблице, ряды стоят плотно. Если есть — между ними добавляется проезд.")
    aisle_config = st.data_editor(
        _model_aisle_config_dataframe(model),
        num_rows="dynamic",
        use_container_width=True,
        key=f"active_aisle_config_{model_key}",
        column_config={
            "row_from": "Ряд от",
            "row_to": "Ряд до",
            "aisle_width_m": st.column_config.NumberColumn("Ширина проезда, м", min_value=0.1, step=0.1),
            "aisle_type": "Тип проезда",
            "comment": "Комментарий",
        },
    )
    b1, b2 = st.columns(2)
    if b1.button("Сохранить настройки проездов", key="active_aisle_save", type="primary"):
        geometry_settings = GeometrySettings(
            cell_length_m=float(settings.get("cell_length_m", 1.2) or 1.2),
            cell_width_m=float(settings.get("cell_width_m", 0.8) or 0.8),
            aisle_width_m=default_aisle_width,
            top_road_width_m=top_road_width,
            bottom_road_width_m=bottom_road_width,
            pallet_height_m=float(settings.get("pallet_height_m", 2.2) or 2.2),
            selected_tier=str(settings.get("selected_tier", "1") or "1"),
            tier_mode=str(settings.get("tier_mode", "selected") or "selected"),
            row_order_mode=str(settings.get("row_order_mode", "row_order_or_number") or "row_order_or_number"),
        )
        rebuilt = rebuild_geometry_from_cells(model, model.get("cells", []), keep_base_cells=True, settings=geometry_settings, aisle_config=aisle_config)
        rebuilt["manual_change_counts"] = model.get("manual_change_counts", rebuilt.get("manual_change_counts", {}))
        save_geometry_model(rebuilt)
        st.session_state["geometry_model"] = rebuilt
        st.success("Настройки проездов сохранены, геометрия активной модели перестроена.")
        st.rerun()
    if b2.button("Сбросить таблицу проездов", key="active_aisle_reset"):
        rebuilt = rebuild_geometry_from_cells(model, model.get("cells", []), keep_base_cells=True, aisle_config=empty_aisle_config())
        save_geometry_model(rebuilt)
        st.session_state["geometry_model"] = rebuilt
        st.success("Межрядные проезды удалены из активной модели.")
        st.rerun()
    return model


def _model_summary_metrics(model: dict) -> None:
    deep_lane_rows = [row for row in model.get("rows", []) if row.get("row_storage_type") == "deep_lane"]
    total_capacity = sum(float(cell.get("capacity_pallets", 1) or 1) for cell in model.get("cells", []))
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Рядов", len(model.get("rows", [])))
    c2.metric("Логических ячеек", len(model.get("cells", [])))
    c3.metric("Набивных рядов", len(deep_lane_rows))
    c4.metric("Вместимость, паллет", f"{total_capacity:g}")


def render_geometry_data_tabs(model: dict) -> None:
    tabs = st.tabs(["Ряды", "Ячейки", "Проезды", "Навигация", "JSON"])
    with tabs[0]:
        st.dataframe(_localized_dataframe(model.get("rows", [])), use_container_width=True)
    with tabs[1]:
        st.dataframe(_localized_dataframe(model.get("cells", [])).head(10000), use_container_width=True)
    with tabs[2]:
        st.dataframe(_localized_dataframe(model.get("aisles", [])), use_container_width=True)
        st.dataframe(_localized_dataframe(model.get("roads", [])), use_container_width=True)
        cross_aisles = model.get("cross_aisles", [])
        st.markdown("#### Поперечные проезды")
        m1, m2, m3 = st.columns(3)
        m1.metric("Количество", len(cross_aisles))
        m2.metric("Рядов с проездами", len({str(item.get("row_number", "")) for item in cross_aisles}))
        m3.metric("Суммарная ширина, м", f"{sum(float(item.get('width_m', 0) or 0) for item in cross_aisles):g}")
        columns = ["aisle_id", "row_number", "after_cell_number", "width_cells", "width_m", "x_min", "x_max", "y_min", "y_max", "comment"]
        st.dataframe(pd.DataFrame(cross_aisles).reindex(columns=columns), use_container_width=True, hide_index=True)
    with tabs[3]:
        st.dataframe(_localized_dataframe(model.get("navigation_nodes", [])), use_container_width=True)
        st.dataframe(_localized_dataframe(model.get("navigation_edges", [])), use_container_width=True)
    with tabs[4]:
        st.download_button("Скачать модель JSON", json.dumps(model, ensure_ascii=False, indent=2).encode("utf-8"), file_name="warehouse_model.json", mime="application/json")


def render_geometry_constructor_view(model: dict) -> None:
    st.subheader("Активная модель")
    st.caption("Служебный просмотр активной модели и ручных изменений.")
    overrides = load_manual_overrides()
    if overrides and overrides.get("source_model_id") != model.get("model_id"):
        overrides = None
    counts = manual_change_counts(overrides)
    st.caption(f"Последний склад загружен из Excel: {model.get('source_file_name', '—')} · Дата построения: {model.get('created_at', '—')}")
    st.caption(f"Ручных изменений: {counts['total']} · Добавлено вручную: {counts['add']} · Изменено вручную: {counts['update']} · Удалено вручную: {counts['delete']}")
    _model_summary_metrics(model)
    st.subheader("Диагностика импорта")
    st.caption("Проверьте предупреждения и статистику построения. Диагностика помогает найти проблемы в исходной схеме без изменения модели.")
    settings = model.get("settings", {})
    stats = [
        ("Проездов между рядами", len(model.get("aisles", []))),
        ("Верхний проезд", f"{settings.get('top_road_width_m', 0)} м"),
        ("Нижний проезд", f"{settings.get('bottom_road_width_m', 0)} м"),
    ]
    cols = st.columns(len(stats))
    for col, (label, value) in zip(cols, stats):
        col.metric(label, value)
    diagnostics = model.get("diagnostics", [])
    if diagnostics:
        st.dataframe(pd.DataFrame(diagnostics), use_container_width=True)
    render_manual_cell_editor(model)
    render_geometry_data_tabs(st.session_state.get("geometry_model", model))




def _map_cell_key(cell: dict) -> str:
    return f"{cell.get('row_number')}|{cell.get('cell_number')}|{cell.get('tier') or '1'}"


def _occupied_for_cell(model: dict, key: str) -> float:
    return sum(float(p.get("occupied_capacity_pallets", p.get("qty_pallets", 0)) or 0) for p in model.get("placements", []) if p.get("cell_key") == key)


def _save_map_edit_snapshot(model: dict) -> None:
    st.session_state["map_edit_undo_model"] = copy.deepcopy(model)


def _persist_map_edit(model: dict, message: str) -> None:
    save_geometry_model(model)
    if not update_data_revisions(model, ["geometry"], "manual_map_edit"):
        return
    invalidate_geometry_render_cache()
    st.session_state["geometry_model"] = model
    st.success(message)


def _physical_slots_for_cell(cell: dict, capacity: int) -> list[dict]:
    if capacity <= 1:
        return []
    x_min = float(cell.get("x_min", 0) or 0)
    x_max = float(cell.get("x_max", x_min) or x_min)
    y_min = float(cell.get("y_min", 0) or 0)
    y_max = float(cell.get("y_max", y_min) or y_min)
    slot_width = (x_max - x_min) / capacity if capacity else 0
    return [
        {
            "slot_index": slot_index,
            "x_min": x_min + (slot_index - 1) * slot_width,
            "x_max": x_min + slot_index * slot_width,
            "y_min": y_min,
            "y_max": y_max,
            "capacity_pallets": 1,
        }
        for slot_index in range(1, capacity + 1)
    ]


def _base_cell_width_m(model: dict, row: dict | None, row_cells: list[dict]) -> float:
    settings_width = float((model.get("settings") or {}).get("cell_width_m") or 0)
    if row and row.get("base_cell_width_m"):
        return float(row.get("base_cell_width_m") or settings_width or 1)
    for cell in row_cells:
        if cell.get("base_cell_width_m"):
            return float(cell.get("base_cell_width_m") or settings_width or 1)
    if settings_width > 0:
        return settings_width
    if row_cells:
        first = row_cells[0]
        current_width = float(first.get("x_max", 0) or 0) - float(first.get("x_min", 0) or 0)
        current_lane_width = int(float(first.get("deep_lane_width", 1) or 1)) if first.get("storage_type") == "deep_lane" else 1
        return current_width / max(current_lane_width, 1) if current_width > 0 else 1.0
    return 1.0


def _row_intersects_another(model: dict, row_number: str, x_min: float, x_max: float, y_min: float, y_max: float) -> bool:
    for other in model.get("rows", []):
        if str(other.get("row_number")) == str(row_number):
            continue
        overlap_x = x_min < float(other.get("x_max", 0) or 0) and x_max > float(other.get("x_min", 0) or 0)
        overlap_y = y_min < float(other.get("y_max", 0) or 0) and y_max > float(other.get("y_min", 0) or 0)
        if overlap_x and overlap_y:
            return True
    return False


def _refresh_linear_geometry_after_row_resize(model: dict) -> None:
    row_by_number = {str(row.get("row_number")): row for row in model.get("rows", [])}
    for aisle in model.get("aisles", []):
        row_from = row_by_number.get(str(aisle.get("row_from")))
        row_to = row_by_number.get(str(aisle.get("row_to")))
        if row_from and row_to:
            aisle["x_min"] = float(row_from.get("x_max", 0) or 0)
            aisle["x_max"] = float(row_to.get("x_min", aisle.get("x_min", 0)) or 0)
            aisle["aisle_width_m"] = max(float(aisle.get("x_max", 0) or 0) - float(aisle.get("x_min", 0) or 0), 0.0)
    total_width = max([float(row.get("x_max", 0) or 0) for row in model.get("rows", [])] + [0.0])
    for road in model.get("roads", []):
        road["x_max"] = total_width
    for node in model.get("navigation_nodes", []):
        row = row_by_number.get(str(node.get("row_number")))
        if row:
            node["x"] = float(row.get("x_center", 0) or 0)
        elif node.get("node_id") in {"road:bottom", "road:top"}:
            node["x"] = total_width / 2 if total_width else 0.0


def _apply_row_storage_geometry(model: dict, row_number: str, storage_type: str, capacity: float) -> tuple[bool, str]:
    logical_capacity = max(1, int(round(capacity))) if storage_type == "deep_lane" else 1
    row_cells = [cell for cell in model.get("cells", []) if str(cell.get("row_number")) == str(row_number)]
    row = next((item for item in model.get("rows", []) if str(item.get("row_number")) == str(row_number)), None)
    base_cell_width = _base_cell_width_m(model, row, row_cells)
    target_width = base_cell_width * logical_capacity
    row_x_min = float((row or row_cells[0]).get("x_min", 0) or 0) if (row or row_cells) else 0.0
    row_x_max = row_x_min + target_width
    row_y_min = float((row or {}).get("y_min", 0) or 0)
    row_y_max = float((row or {}).get("y_max", 0) or 0)
    if _row_intersects_another(model, row_number, row_x_min, row_x_max, row_y_min, row_y_max):
        return False, "Недостаточно места для расширения набивного ряда. Переместите соседние ряды или увеличьте расстояние между ними."
    for cell in row_cells:
        cell["base_cell_width_m"] = base_cell_width
        cell["x_min"] = row_x_min
        cell["x_max"] = row_x_max
        cell["x_center"] = (row_x_min + row_x_max) / 2
        cell["width_m"] = target_width
        cell["storage_type"] = storage_type
        cell["deep_lane_width"] = logical_capacity
        cell["capacity_pallets"] = logical_capacity
        cell["volume_m3"] = round(float(cell.get("length_m", 0) or 0) * base_cell_width * logical_capacity * float((model.get("settings") or {}).get("pallet_height_m", 1.7) or 1.7), 4)
        cell["physical_slots"] = _physical_slots_for_cell(cell, logical_capacity) if storage_type == "deep_lane" else []
    if row:
        row["base_cell_width_m"] = base_cell_width
        row["base_row_width_m"] = base_cell_width
        row["x_min"] = row_x_min
        row["x_max"] = row_x_max
        row["x_center"] = (row_x_min + row_x_max) / 2
        row["width_m"] = target_width
        row["row_storage_type"] = storage_type
        row["deep_lane_width"] = logical_capacity
        row["capacity_pallets"] = logical_capacity * len(row_cells)
        row["cells_count"] = len(row_cells)
    for setting in model.get("row_settings", []):
        if str(setting.get("row_number")) == str(row_number):
            setting["row_storage_type"] = storage_type
            setting["deep_lane_width"] = logical_capacity
            setting["base_cell_width_m"] = base_cell_width
            setting["base_row_width_m"] = base_cell_width
    _refresh_linear_geometry_after_row_resize(model)
    return True, "Геометрия ряда обновлена."


def _cell_options(model: dict) -> dict[str, str]:
    return {f"Ряд {c.get('row_number')} · ячейка {c.get('cell_number')} · ярус {c.get('tier') or '1'}": _map_cell_key(c) for c in model.get("cells", [])}


def _row_options(model: dict) -> dict[str, str]:
    return {f"Ряд {r.get('row_number')} · порядок {r.get('row_order')}": str(r.get("row_number")) for r in model.get("rows", [])}


def _find_map_cell(model: dict, key: str) -> dict | None:
    return next((cell for cell in model.get("cells", []) if _map_cell_key(cell) == key), None)


def _find_map_row(model: dict, row_number: str) -> dict | None:
    return next((row for row in model.get("rows", []) if str(row.get("row_number")) == str(row_number)), None)


def _cell_duplicate_exists(model: dict, row_number: str, cell_number: str, tier: str, original_key: str = "") -> bool:
    new_key = f"{row_number}|{cell_number}|{tier or '1'}"
    return any(_map_cell_key(cell) == new_key and _map_cell_key(cell) != original_key for cell in model.get("cells", []))


def _row_has_placements(model: dict, row_number: str) -> bool:
    row_cells = {_map_cell_key(cell) for cell in model.get("cells", []) if str(cell.get("row_number")) == str(row_number)}
    return any(p.get("cell_key") in row_cells for p in model.get("placements", []))


def _add_cell_near(model: dict, selected: dict, where: str, new_number: str) -> tuple[bool, str]:
    tier = str(selected.get("tier") or "1")
    row_number = str(selected.get("row_number"))
    if _cell_duplicate_exists(model, row_number, new_number, tier):
        return False, "Ячейка с таким адресом уже существует."
    _save_map_edit_snapshot(model)
    row_cells = [cell for cell in model.get("cells", []) if str(cell.get("row_number")) == row_number]
    selected_idx = sorted(row_cells, key=lambda c: float(c.get("y_min", 0))).index(selected) if selected in row_cells else len(row_cells) - 1
    insert_idx = selected_idx if where == "before" else selected_idx + 1
    length = float(selected.get("length_m", 1.2) or 1.2)
    y_min = insert_idx * length
    new_cell = dict(selected)
    new_cell.update({"code": "", "cell_number": str(new_number), "y_min": y_min, "y_max": y_min + length, "y_center": y_min + length / 2, "source": "manual_add"})
    if selected.get("storage_type") != "deep_lane":
        new_cell["capacity_pallets"] = 1
    model["cells"].append(new_cell)
    for idx, cell in enumerate(sorted([c for c in model["cells"] if str(c.get("row_number")) == row_number], key=lambda c: float(c.get("y_min", 0)))):
        cell["y_min"] = idx * length
        cell["y_max"] = cell["y_min"] + length
        cell["y_center"] = cell["y_min"] + length / 2
    row = _find_map_row(model, row_number)
    if row:
        row["cells_count"] = len([c for c in model["cells"] if str(c.get("row_number")) == row_number])
        row["capacity_pallets"] = sum(float(c.get("capacity_pallets", 1) or 1) for c in model["cells"] if str(c.get("row_number")) == row_number)
        row["y_max"] = row["cells_count"] * length
    return True, "Ячейка добавлена."




def _selection_stats(model: dict, row_numbers: list[str], cell_keys: list[str]) -> dict:
    selected_cells = [c for c in model.get("cells", []) if _map_cell_key(c) in set(cell_keys) or str(c.get("row_number")) in set(row_numbers)]
    unique = {_map_cell_key(c): c for c in selected_cells}.values()
    capacity = sum(float(c.get("capacity_pallets", 1) or 1) for c in unique)
    occupied = sum(_occupied_for_cell(model, _map_cell_key(c)) for c in unique)
    return {"rows": len(set(row_numbers)), "cells": len(list(unique)), "capacity": capacity, "occupied": occupied, "free": max(capacity - occupied, 0)}


def _shift_rows(model: dict, row_numbers: list[str], dx: float, dy: float, snap: bool, step: float) -> tuple[bool, str]:
    if snap and step > 0:
        dx = round(dx / step) * step
        dy = round(dy / step) * step
    moving = [row for row in model.get("rows", []) if str(row.get("row_number")) in set(row_numbers)]
    if not moving:
        return False, "Выберите ряды для сдвига."
    snapshots = {str(r.get("row_number")): dict(r) for r in moving}
    for row in moving:
        row["x_min"] = float(row.get("x_min", 0) or 0) + dx
        row["x_max"] = float(row.get("x_max", 0) or 0) + dx
        row["x_center"] = float(row.get("x_center", 0) or 0) + dx
        row["y_min"] = float(row.get("y_min", 0) or 0) + dy
        row["y_max"] = float(row.get("y_max", 0) or 0) + dy
    for cell in model.get("cells", []):
        if str(cell.get("row_number")) in set(row_numbers):
            for key in ["x_min", "x_max", "x_center"]:
                cell[key] = float(cell.get(key, 0) or 0) + dx
            for key in ["y_min", "y_max", "y_center"]:
                cell[key] = float(cell.get(key, 0) or 0) + dy
            for slot in cell.get("physical_slots", []):
                for key in ["x_min", "x_max"]:
                    slot[key] = float(slot.get(key, 0) or 0) + dx
                for key in ["y_min", "y_max"]:
                    slot[key] = float(slot.get(key, 0) or 0) + dy
    # block row intersections after move
    rows = model.get("rows", [])
    for idx, a in enumerate(rows):
        for b in rows[idx + 1:]:
            if str(a.get("row_number")) == str(b.get("row_number")):
                continue
            overlap_x = float(a.get("x_min", 0)) < float(b.get("x_max", 0)) and float(a.get("x_max", 0)) > float(b.get("x_min", 0))
            overlap_y = float(a.get("y_min", 0)) < float(b.get("y_max", 0)) and float(a.get("y_max", 0)) > float(b.get("y_min", 0))
            if overlap_x and overlap_y:
                for row in moving:
                    row.update(snapshots[str(row.get("row_number"))])
                return False, f"Нельзя сохранить: ряд {a.get('row_number')} пересекается с рядом {b.get('row_number')}."
    return True, "Ряды сдвинуты."


def render_map_edit_panel(model: dict) -> dict:
    with st.expander("Ручное редактирование ячеек", expanded=False):
        st.caption("Выберите ячейку из списка. Карта не сообщает клики обратно в Streamlit, поэтому выбранным считается только объект из этого списка.")
        options = _cell_options(model)
        if not options:
            st.info("В модели нет ячеек.")
            return model
        label = st.selectbox("Выбранная ячейка", list(options), key="map_cell_select")
        key = options[label]
        st.session_state["map_selected_cell_key"] = key
        st.session_state.pop("map_selected_row_number", None)
        cell = _find_map_cell(model, key)
        if not cell:
            return model
        occupied = _occupied_for_cell(model, key)
        capacity = float(cell.get("capacity_pallets", 1) or 1)
        is_active = "block" not in str(cell.get("source", "")).lower()
        st.markdown("#### Выбрана ячейка")
        info = st.columns(4)
        info[0].metric("Ряд / ячейка", f"{cell.get('row_number')} / {cell.get('cell_number')}")
        info[1].metric("Ярус", cell.get("tier") or "1")
        info[2].metric("Вместимость", f"{capacity:g}")
        info[3].metric("Свободно", f"{max(capacity - occupied, 0):g}")
        st.info(
            f"Адрес: {key} · Занято: {occupied:g} · "
            f"Тип: {display_label(STORAGE_TYPE_LABELS, cell.get('storage_type'))} · "
            f"Весовая зона: {display_label(WEIGHT_ZONE_LABELS, cell.get('weight_zone', 'unassigned'))} · "
            f"Состояние: {CELL_STATE_LABELS[is_active]}"
        )
        c1, c2, c3 = st.columns(3)
        new_number = c1.text_input("Новый номер ячейки", value=str(cell.get("cell_number", "")), key="map_cell_new_number")
        new_capacity = c2.number_input("Вместимость", min_value=0.0, value=capacity, step=1.0, key="map_cell_capacity")
        blocked = c3.checkbox("Заблокирована", value=not is_active, key="map_cell_blocked")
        if st.button("Применить изменения ячейки", key="map_cell_apply"):
            if _cell_duplicate_exists(model, str(cell.get("row_number")), new_number, str(cell.get("tier") or "1"), key):
                st.error("Ячейка с таким адресом уже существует.")
            elif new_capacity < occupied:
                st.error("Нельзя уменьшить вместимость ниже занятого количества паллет.")
            else:
                _save_map_edit_snapshot(model)
                old_key = _map_cell_key(cell)
                cell["cell_number"] = str(new_number)
                cell["capacity_pallets"] = new_capacity
                cell["source"] = "block_manual" if blocked else "manual_update"
                new_key = _map_cell_key(cell)
                for placement in model.get("placements", []):
                    if placement.get("cell_key") == old_key:
                        placement["cell_key"] = new_key
                        placement["cell_number"] = str(new_number)
                _persist_map_edit(model, "Ячейка обновлена.")
                st.rerun()
        add_number = st.text_input("Номер добавляемой ячейки", value=str(int(float(cell.get("cell_number", 0) or 0)) + 1) if str(cell.get("cell_number", "")).isdigit() else "", key="map_cell_add_number")
        a1, a2, a3 = st.columns(3)
        if a1.button("Добавить до", key="map_cell_add_before"):
            ok, msg = _add_cell_near(model, cell, "before", add_number)
            (st.success if ok else st.error)(msg)
            if ok:
                _persist_map_edit(model, msg); st.rerun()
        if a2.button("Добавить после", key="map_cell_add_after"):
            ok, msg = _add_cell_near(model, cell, "after", add_number)
            (st.success if ok else st.error)(msg)
            if ok:
                _persist_map_edit(model, msg); st.rerun()
        confirm = st.checkbox("Подтвердить удаление ячейки", key="map_cell_delete_confirm")
        if a3.button("Удалить ячейку", disabled=not confirm, key="map_cell_delete"):
            if occupied > 0:
                st.error("Ячейка занята. Сначала переместите или сбросьте размещение товара.")
            else:
                _save_map_edit_snapshot(model)
                model["cells"] = [c for c in model.get("cells", []) if _map_cell_key(c) != key]
                _persist_map_edit(model, "Ячейка удалена.")
                st.session_state.pop("map_selected_cell_key", None)
                st.rerun()
    return model

def render_geometry_map_view(model: dict) -> None:
    st.subheader("Карта склада")
    st.caption("Используйте кнопки масштаба, колесо мыши и перетаскивание для навигации по карте. Переключение вкладок не перестраивает склад и не сбрасывает ручные правки.")
    with measure_step("load_placements"):
        placement_state, placement_warning = load_placement_state(model)
    if placement_warning:
        st.warning(placement_warning)
    else:
        with measure_step("enrich_model_with_placements_and_diagnostics"):
            model = attach_placements_to_model(model, placement_state)
            if placement_state.get("placements"):
                snapshot, _ = load_pre_placement_snapshot(model)
                model = enrich_model_with_placement_diagnostics(model, placement_state, snapshot)
                st.caption("На карте показана занятость из сохранённого placements.json, включая рассчитанные приходы.")
            model = enrich_model_with_outbound_diagnostics(model, placement_state)
    _model_summary_metrics(model)
    st.markdown(
        " · ".join(
            f"<span style='display:inline-flex;align-items:center;gap:4px;margin-right:8px'><span style='display:inline-block;width:14px;height:14px;background:{color};border:1px solid #94A3B8'></span>{ZONE_LABELS_RU.get(zone, zone)}</span>"
            for zone, color in PLACEMENT_CATEGORY_COLORS.items()
        ) + "<span style='display:inline-flex;align-items:center;gap:4px;margin-right:8px'><span style='display:inline-block;width:14px;height:14px;background:#DCEBFF;border:1px solid #AAB4C3'></span>Свободно</span><span style='display:inline-flex;align-items:center;gap:4px'><span style='display:inline-block;width:14px;height:14px;background:#F3F4F6;border:2px dashed #6B7280'></span>Заблокировано</span>",
        unsafe_allow_html=True,
    )
    control_left, control_right = st.columns([1, 2])
    with control_left:
        detailed = st.toggle("Детальный режим", value=len(model.get("cells", [])) <= 1500, key="map_detailed_mode")
    with control_right:
        scale = st.slider("Масштаб, px/м", min_value=4.0, max_value=60.0, value=22.0, step=1.0, key="map_scale")
    label_settings = render_map_settings_editor()
    label_settings["edit_mode"] = False
    label_settings["selected_cell_key"] = st.session_state.get("map_selected_cell_key", "")
    label_settings["selected_row_number"] = ""
    render_started = perf_counter()
    with measure_step("build_map_html_svg", {"detailed": detailed}):
        html = build_geometry_html_cached(json.dumps(model, ensure_ascii=False), scale, detailed, json.dumps(label_settings, ensure_ascii=False, sort_keys=True))
    components.html(html, height=980, scrolling=True)
    st.caption(f"Рендер карты: {perf_counter() - render_started:.2f} сек. Модель: data/last_import/warehouse_model.json")


def render_virtual_warehouse_excel() -> None:
    st.sidebar.caption(f"Сборка приложения: {APP_BUILD_LABEL}")
    render_git_release_badge()
    mode = st.sidebar.radio(
        "Режим",
        ["Склад из Excel: ряды + ячейки + проезды", "Виртуальный склад по Excel-схеме"],
        index=0,
    )
    if mode == "Склад из Excel: ряды + ячейки + проезды":
        render_excel_geometry_warehouse()
        return

    st.title("Симулятор скорости сборки")
    st.header("Виртуальный склад по Excel-схеме")
    st.sidebar.info("Оставлен только режим виртуального склада Excel. Старые разделы скрыты из интерфейса.")

    ensure_loaded_model()
    if st.session_state.get("last_import_error"):
        st.warning(f"Сохранённая модель не загружена: {st.session_state['last_import_error']}. Загрузите Excel заново.")

    with st.sidebar:
        st.subheader("Последний склад")
        if MODEL_PATH.exists():
            if st.button("Использовать сохранённый склад"):
                for key in ["virtual_warehouse_model", "virtual_warehouse_render_cache", "virtual_warehouse_meta"]:
                    st.session_state.pop(key, None)
                ensure_loaded_model()
                st.rerun()
            if st.button("Очистить сохранённый склад"):
                clear_last_import()
                for key in list(st.session_state.keys()):
                    if key.startswith("virtual_warehouse") or key == "last_import_error":
                        st.session_state.pop(key, None)
                st.rerun()
        else:
            st.caption("Сохранённого склада пока нет.")

    with st.expander("Загрузка нового Excel", expanded="virtual_warehouse_model" not in st.session_state):
        schema_file = st.file_uploader("Excel-схема склада", type=["xlsx"], key="virtual_warehouse_schema_upload")
        selected_sheets: list[str] | None = None
        schema_hash = ""
        schema_bytes = b""
        if schema_file is not None:
            schema_bytes = schema_file.getvalue()
            schema_hash = file_hash(schema_bytes)
            sheet_names = get_excel_sheet_names(schema_bytes, schema_hash)
            selected_sheets = st.multiselect("Листы для обработки", sheet_names, default=sheet_names[:1])
        cell_file = st.file_uploader("Файл номеров ячеек (необязательно)", type=["xlsx", "csv"], key="virtual_warehouse_cells_upload")
        placement_file = st.file_uploader("Файл размещения товаров (необязательно)", type=["xlsx", "csv"], key="virtual_warehouse_placements_upload")
        build_clicked = st.button("Построить склад", disabled=schema_file is None or not selected_sheets)

    if build_clicked:
        diagnostics: list[dict] = []
        try:
            timings = {}
            started = perf_counter()
            model = parse_warehouse_excel_cached(schema_bytes, schema_hash, tuple(selected_sheets or []))
            timings["read_excel_seconds"] = perf_counter() - started
            build_started = perf_counter()
            if cell_file is not None:
                addresses_by_row, cell_diagnostics = import_cell_addresses(cell_file)
                diagnostics.extend(cell_diagnostics)
                diagnostics.extend(apply_cell_addresses(model, addresses_by_row))
            if placement_file is not None:
                placements, placement_diagnostics = import_placements(placement_file)
                diagnostics.extend(placement_diagnostics)
                diagnostics.extend(apply_placements(model, placements))
            timings["build_model_seconds"] = perf_counter() - build_started
            render_started = perf_counter()
            with measure_step("prepare_render_cache"):
                render_cache = prepare_render_cache_cached(model_to_dict(model))
            timings["prepare_render_seconds"] = perf_counter() - render_started
            meta = {
                "model_version": MODEL_VERSION,
                "source": "новый Excel",
                "imported_at": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
                "scheme_file_name": schema_file.name,
                "scheme_file_hash": schema_hash,
                "scheme_file_size": len(schema_bytes),
                "cells_file_name": getattr(cell_file, "name", ""),
                "cells_file_hash": file_hash(cell_file.getvalue()) if cell_file else "",
                "placement_file_name": getattr(placement_file, "name", ""),
                "placement_file_hash": file_hash(placement_file.getvalue()) if placement_file else "",
                "selected_sheets": selected_sheets,
                "model_path": str(MODEL_PATH),
                "diagnostics": diagnostics,
                "performance": timings,
                "loaded_from_cache": False,
            }
            save_uploaded_copy(schema_file, "source_scheme.xlsx")
            save_uploaded_copy(cell_file, "source_cells.xlsx")
            save_uploaded_copy(placement_file, "source_placement.xlsx")
            timings["save_model_seconds"] = save_last_import(model, render_cache, meta)
            meta["performance"] = timings
            write_json_atomic(META_PATH, meta)
            revision_domains = ["geometry"] + (["placements"] if placement_file is not None else [])
            if not update_data_revisions(model_to_dict(model), revision_domains, "load_new_model"):
                return
            st.session_state["virtual_warehouse_model"] = model
            st.session_state["virtual_warehouse_render_cache"] = render_cache
            st.session_state["virtual_warehouse_meta"] = meta
            st.session_state["virtual_warehouse_source"] = "новый Excel"
            st.success(f"Склад построен и сохранён: {len(model.cells)} ячеек. Старая модель перезаписана только после успешного построения.")
        except Exception as exc:
            st.error(f"Не удалось построить склад. Последняя успешная сохранённая модель не удалена: {exc}")

    model = st.session_state.get("virtual_warehouse_model")
    if model is None:
        st.info("Загрузите Excel-схему и нажмите «Построить склад» или используйте ранее сохранённый склад.")
        return

    meta = st.session_state.get("virtual_warehouse_meta", {})
    diagnostics = meta.get("diagnostics", st.session_state.get("virtual_warehouse_diagnostics", []))
    render_diagnostics(model, meta, diagnostics)
    render_performance(meta)

    sheet_names = [sheet.name for sheet in model.sheets]
    selected_sheet = next(sheet for sheet in model.sheets if sheet.name == st.selectbox("Лист склада", sheet_names))
    cells = selected_sheet.potential_cells if hasattr(selected_sheet, "potential_cells") else [cell for row in selected_sheet.rows for cell in row.potential_cells]
    zones = ["Все"] + sorted({cell.fill_color or "без зоны" for cell in cells})
    rows = ["Все"] + sorted({str(row.row_number) for row in selected_sheet.rows})
    tiers = ["Все"] + sorted({str(cell.tier_number) for cell in cells})
    f1, f2, f3, f4 = st.columns(4)
    zone_filter = f1.selectbox("Зона", zones)
    row_filter = f2.selectbox("Ряд", rows)
    tier_filter = f3.selectbox("Ярус", tiers)
    availability_filter = f4.selectbox("Показывать", ["Все", "Только активные", "Только заблокированные"])

    filtered_sheet = filter_sheet(selected_sheet, zone_filter, row_filter, tier_filter, availability_filter)
    filtered_count = sum(len(row.potential_cells) for row in filtered_sheet.rows)
    summary_mode = filtered_count > 2000 and row_filter == "Все"
    if summary_mode:
        st.warning("Включён общий вид: ячеек больше 2000. Выберите ряд/фильтр для детального интерактива по ячейкам.")
    scale = st.slider("Масштаб сетки", min_value=18, max_value=60, value=34, step=2)
    render_started = perf_counter()
    html = build_virtual_warehouse_html_cached(asdict(filtered_sheet), scale, summary_mode)
    components.html(html, height=760, scrolling=True)
    render_seconds = perf_counter() - render_started
    meta.setdefault("performance", {})["render_seconds"] = render_seconds
    st.caption(f"Последний рендер: {render_seconds:.2f} сек.; показано ячеек: {filtered_count}.")

    tab_rows, tab_cells, tab_diag = st.tabs(["Ряды", "Ячейки", "Диагностика"])
    with tab_rows:
        st.dataframe(pd.DataFrame([asdict(row) | {"cells": len(row.potential_cells)} for row in filtered_sheet.rows]).drop(columns=["potential_cells"], errors="ignore"), use_container_width=True)
    with tab_cells:
        st.dataframe(pd.DataFrame([asdict(cell) for row in filtered_sheet.rows for cell in row.potential_cells]).head(5000), use_container_width=True)
    with tab_diag:
        diag_df = pd.DataFrame(build_diagnostics(model, diagnostics))
        st.dataframe(diag_df, use_container_width=True)
        st.download_button("Скачать диагностику CSV", diag_df.to_csv(index=False).encode("utf-8-sig"), file_name="virtual_warehouse_diagnostics.csv", mime="text/csv")
        st.download_button("Скачать модель JSON", json.dumps(model_to_dict(model), ensure_ascii=False, indent=2).encode("utf-8"), file_name="warehouse_model.json", mime="application/json")


_VIRTUAL_WAREHOUSE_APP_RENDERED = False


def main() -> None:
    global _VIRTUAL_WAREHOUSE_APP_RENDERED
    if get_script_run_ctx(suppress_warning=True) is None and __name__ != "__main__":
        return
    if _VIRTUAL_WAREHOUSE_APP_RENDERED:
        return
    _VIRTUAL_WAREHOUSE_APP_RENDERED = True
    start_performance_run(enabled=is_performance_enabled(st.session_state))
    error = None
    try:
        with measure_step("streamlit_main"):
            render_virtual_warehouse_excel()
    except BaseException as exc:
        error = exc
        raise
    finally:
        model = st.session_state.get("geometry_model")
        finish_performance_run(
            exception=error,
            model_id=model.get("model_id") if isinstance(model, dict) else None,
        )


if __name__ == "__main__" or get_script_run_ctx(suppress_warning=True) is not None:
    main()
