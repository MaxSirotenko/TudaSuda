"""Versioned factual source registry for the authoritative Data workspace.

This module preserves source rows and their provenance.  It intentionally does
not convert source pallet references, inventory quantities, or historical pick
orders into simulator authority.  The existing one-day V1 adapters remain the
only benchmark input boundary.
"""
from __future__ import annotations

import gzip
import hashlib
import json
import math
import os
import sqlite3
import tempfile
import shutil
import time
from collections import Counter, defaultdict
from collections.abc import Iterable, Mapping
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from numbers import Real
from pathlib import Path
from typing import Any, Callable

import pandas as pd
from warehouse_perf_diagnostics import measure, profiled, record_file_read
from warehouse_factual_artifacts import (
    atomic_json as _artifact_atomic_json,
    iter_jsonl as _artifact_iter_jsonl,
    read_jsonl as _artifact_read_jsonl,
    write_jsonl as _artifact_write_jsonl,
)

from warehouse_business_identity import build_canonical_sku_identity, find_canonical_identity_collisions

PARSER_VERSION = "factual-july-v5"
# Level 6 retained ~97% of level-9 compression in the 10k-row benchmark while
# spending materially less time in the three gzip writers.
FACTUAL_GZIP_COMPRESSLEVEL = 6
IMPORT_PROGRESS_ROW_INTERVAL = 5_000
IMPORT_PROGRESS_TIME_INTERVAL_SECONDS = 0.75
ImportProgressCallback = Callable[[dict[str, Any]], None]
DATA_ROOT = Path("data/last_import/factual")
REGISTRY_PATH = DATA_ROOT / "registry.json"
UNKNOWN_SOURCE = "unknown"
SOURCE_LABELS = {
    "historical_placement": "Историческое размещение",
    "inventory": "Инвентаризации",
    "receipts": "Приходные ордера",
    "outbound": "Расходные ордера",
    "vgh": "ВГХ / паллетизация",
    UNKNOWN_SOURCE: "Неизвестный тип файла",
}

# Authoritative aliases are deliberately reviewable and immutable by accident:
# each spelling below is present in the audited contract document and a real
# project export or checked-in 1C query.  Compatibility aliases are kept in a
# separate table so a synthetic fixture cannot promote one to source authority.
AUTHORITATIVE_CONTRACTS: dict[str, dict[str, tuple[str, ...]]] = {
    "historical_placement": {
        "snapshot_at": ("ДатаСреза",), "source_pallet_ref": ("Паллета",),
        "nomenclature": ("Номенклатура",), "characteristic": ("Характеристика",),
        "source_stock_quantity": ("КоличествоОстатокТовара",), "cell": ("Ячейка",),
        "cell_picking_order": ("ПорядокСборки",),
        "source_position_balance": ("КоличествоОстатокПоложения",),
    },
    "inventory": {
        "inventory_ref": ("Ссылка",),
        "inventory_number": ("Номер",),
        "occurred_at": ("Дата",), "line_number": ("НомерСтроки",),
        "warehouse": ("Склад",), "nomenclature": ("Номенклатура",),
        "characteristic": ("Характеристика",), "actual_quantity": ("КоличествоФакт",),
        "accounting_quantity": ("КоличествоУчет",),
    },
    "receipts": {
        "document_ref": ("Ссылка",), "document_number": ("Номер",),
        "occurred_at": ("Дата",), "warehouse": ("Склад",),
        "line_number": ("НомерСтроки",), "nomenclature": ("Номенклатура",),
        "characteristic": ("Характеристика",), "box_quantity": ("КоличествоКоробок",),
        "reported_pallets": ("КоличествоПаллет",),
        "terminal_completed": ("ПриемкаТерминаломЗакончена",),
        "expected_receipt": ("ОжидаемыйПриход",),
    },
    "outbound": {
        "document_ref": ("СсылкаРО",),
        "document_number": ("НомерРО",),
        "occurred_at": ("ДатаРО",), "warehouse": ("Склад",),
        "line_number": ("НомерСтроки",), "nomenclature": ("Номенклатура",),
        "characteristic": ("Характеристика",),
        "quantity": ("РасчетноеОтгруженоКоробок",),
        "source_pick_order": ("ПорядокСборки",),
    },
    "vgh": {
        "nomenclature": ("Номенклатура",), "characteristic": ("Характеристика",),
        "weight": ("Вес",), "length": ("Длина",), "width": ("Ширина",), "height": ("Высота",),
        "boxes_per_layer": ("КоличествоКоробовВОдномСлоеНаПаллете",),
        "layers_per_pallet": ("КоличествоСлоевНаПаллете",),
        "quantity_per_box": ("КоличествоВКоробке",),
    },
}
LEGACY_CONTRACT_ALIASES: dict[str, dict[str, tuple[str, ...]]] = {
    "inventory": {
        "inventory_ref": ("Инвентаризация", "СсылкаИнвентаризации"),
        "inventory_number": ("НомерИнвентаризации",), "occurred_at": ("ДатаИнвентаризации",),
        "actual_quantity": ("ФактическоеКоличество",), "accounting_quantity": ("УчетноеКоличество",),
    },
    "receipts": {
        "document_ref": ("СсылкаПриходногоОрдера",), "document_number": ("НомерПриходногоОрдера",),
        "occurred_at": ("ДатаПриходногоОрдера",),
    },
}


def _executable_contracts() -> dict[str, dict[str, tuple[str, ...]]]:
    result = {source: dict(fields) for source, fields in AUTHORITATIVE_CONTRACTS.items()}
    for source, fields in LEGACY_CONTRACT_ALIASES.items():
        for field, aliases in fields.items():
            result[source][field] = result[source].get(field, ()) + aliases
    return result


CONTRACTS = _executable_contracts()
CONTRACT_ALIAS_STATUS = {
    source: {alias: "authoritative" for aliases in fields.values() for alias in aliases}
    for source, fields in AUTHORITATIVE_CONTRACTS.items()
}
for _source, _fields in LEGACY_CONTRACT_ALIASES.items():
    CONTRACT_ALIAS_STATUS.setdefault(_source, {}).update(
        {alias: "legacy_compatibility" for aliases in _fields.values() for alias in aliases})
# Each accepted spelling is traceable to the task's confirmed exports or to a
# checked-in 1C query.  Keeping the evidence beside the executable contract
# prevents a synthetic fixture from silently becoming source authority.
CONTRACT_EVIDENCE = {
    "historical_placement": "CONFIRMED_PROJECT_SOURCE:размещение июль.xlsx",
    "inventory": "CONFIRMED_PROJECT_SOURCE;LEGACY_COMPATIBILITY:queries_1c/inventory_results.query",
    "receipts": "CONFIRMED_PROJECT_SOURCE:ПО июль.xlsx;LEGACY_COMPATIBILITY:queries_1c/day_receipts.query",
    "outbound": "EXISTING_WORKING_QUERY:queries_1c/mass_outbound_orders.query",
    "vgh": "CONFIRMED_PROJECT_SOURCE",
}
REQUIRED = {
    "historical_placement": {"snapshot_at", "source_pallet_ref", "nomenclature", "characteristic", "source_stock_quantity", "cell", "cell_picking_order", "source_position_balance"},
    "inventory": {"inventory_ref", "inventory_number", "occurred_at", "line_number", "warehouse", "nomenclature", "characteristic", "actual_quantity", "accounting_quantity"},
    "receipts": {"document_ref", "document_number", "occurred_at", "warehouse", "line_number", "nomenclature", "characteristic", "box_quantity"},
    # Outbound document identity is an OR rule and is checked explicitly by
    # detect_source_type: document_ref, or document_number + occurred_at.
    "outbound": {"nomenclature", "characteristic", "quantity"},
    "vgh": {"nomenclature", "characteristic", "boxes_per_layer", "layers_per_pallet"},
}


def _norm(value: Any) -> str:
    return "" if value is None else "".join(str(value).split()).casefold().replace("ё", "е")


def _json_value(value: Any) -> Any:
    if value is None or (not isinstance(value, (list, dict, str)) and pd.isna(value)):
        return None
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat()
    if hasattr(value, "item"):
        value = value.item()
    if isinstance(value, float) and not math.isfinite(value):
        return None
    return value if isinstance(value, (str, int, float, bool)) else str(value)


def _number(value: Any) -> float | int | None:
    value = _json_value(value)
    if value is None or isinstance(value, bool):
        return None
    try:
        number = float(str(value).replace("\u00a0", "").replace(" ", "").replace(",", "."))
    except ValueError:
        return None
    if not math.isfinite(number):
        return None
    return int(number) if number.is_integer() else number


_SOURCE_DATETIME_FORMATS = (
    "%d.%m.%Y %H:%M:%S",
    "%d.%m.%Y %H:%M",
    "%d.%m.%Y",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d",
)


def normalize_source_datetime(value: Any) -> str | None:
    """Return a deterministic factual-source datetime in canonical ISO form.

    Native Excel dates retain their calendar value.  Text is accepted only in
    the confirmed Russian day-first and ISO source formats; no locale-dependent
    or fuzzy parser participates in factual partitioning.
    """
    if value is None or value is pd.NaT:
        return None
    if isinstance(value, pd.Timestamp):
        return None if pd.isna(value) else value.isoformat()
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text:
        return None
    # The common ISO representation is unambiguous and avoids repeated
    # strptime format scans in the hot loop.
    if len(text) >= 10 and text[4:5] == "-" and text[7:8] == "-":
        try:
            return datetime.fromisoformat(text.replace("Z", "+00:00")).isoformat()
        except ValueError:
            pass
    for source_format in _SOURCE_DATETIME_FORMATS:
        try:
            return datetime.strptime(text, source_format).isoformat()
        except ValueError:
            continue
    return None


def normalize_operational_day(value: Any) -> str | None:
    """Normalize an Excel placement snapshot value to one calendar day.

    Historical exports can mix genuinely typed Excel dates, unformatted Excel
    serial numbers and text in the same column.  The operational day is the
    calendar date written by the source; it must not be shifted by a timezone
    conversion and must never retain a time component.
    """
    if value is None or value is pd.NaT or isinstance(value, bool):
        return None
    if isinstance(value, (pd.Timestamp, datetime, date)):
        if pd.isna(value):
            return None
        return value.date().isoformat() if isinstance(value, (pd.Timestamp, datetime)) else value.isoformat()
    if isinstance(value, Real):
        if not math.isfinite(float(value)) or not 1 <= float(value) < 2_958_466:
            return None
        # Excel's default 1900 date system includes its historical leap-year
        # compatibility offset; 1899-12-30 is the established serial epoch.
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text:
        return None
    if len(text) >= 10 and text[4:5] == "-" and text[7:8] == "-":
        try:
            return date.fromisoformat(text[:10]).isoformat()
        except ValueError:
            pass
    for source_format in _SOURCE_DATETIME_FORMATS:
        try:
            return datetime.strptime(text, source_format).date().isoformat()
        except ValueError:
            continue
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError:
        return None
    return parsed.date().isoformat()


def content_hash(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def dataset_identity(data_hash: str, source_type: str, sheet: str, parser_version: str = PARSER_VERSION) -> str:
    payload = "\x1f".join((data_hash, source_type, sheet, parser_version))
    return "dataset:" + hashlib.sha256(payload.encode()).hexdigest()


def detect_source_type(columns: Iterable[Any]) -> dict[str, Any]:
    """Detect by exact observed columns; filename is deliberately not authority."""
    normalized = {_norm(column): str(column) for column in columns}
    matches: dict[str, dict[str, str]] = {}
    # Priority is documented for review, but never resolves a collision: all
    # matches are collected and an ambiguous schema is returned explicitly.
    detection_order = ("historical_placement", "inventory", "vgh", "outbound", "receipts")
    for source_type in detection_order:
        fields = CONTRACTS[source_type]
        mapping = {}
        for field, aliases in fields.items():
            found = next((normalized[_norm(alias)] for alias in aliases if _norm(alias) in normalized), None)
            if found:
                mapping[field] = found
        required = REQUIRED[source_type].issubset(mapping)
        if source_type == "outbound":
            required = required and ("document_ref" in mapping or {"document_number", "occurred_at"}.issubset(mapping))
        if required:
            matches[source_type] = mapping
    status = "detected" if len(matches) == 1 else "ambiguous_schema" if len(matches) > 1 else "unknown_schema"
    source_type = next(iter(matches)) if len(matches) == 1 else UNKNOWN_SOURCE
    missing: list[str] = []
    diagnostic_code: str | None = None
    diagnostic_found: list[str] = []
    diagnostic_missing: list[str] = []
    outbound_common = {"nomenclature", "characteristic", "quantity"}
    outbound_mapping = {field: next((normalized[_norm(alias)] for alias in aliases if _norm(alias) in normalized), None)
                        for field, aliases in CONTRACTS["outbound"].items()}
    outbound_mapping = {field: value for field, value in outbound_mapping.items() if value}
    if source_type == UNKNOWN_SOURCE and not matches and outbound_common.issubset(outbound_mapping):
        diagnostic_code = "outbound_document_identity_missing"
        diagnostic_found = [outbound_mapping[field] for field in ("nomenclature", "characteristic", "quantity", "occurred_at")
                            if field in outbound_mapping]
        diagnostic_missing = ["СсылкаРО или НомерРО + ДатаРО"]
    if source_type == UNKNOWN_SOURCE and not matches and diagnostic_code is None:
        # Family signatures are exact, confirmed source fields, not fuzzy
        # aliases.  They allow a useful mapping_required diagnostic without
        # authorising canonical rows.
        signatures = {"receipts": "box_quantity", "outbound": "quantity", "historical_placement": "source_pallet_ref",
                      "inventory": "actual_quantity", "vgh": "boxes_per_layer"}
        candidates = []
        partial = {}
        for family, signature in signatures.items():
            mapping = {field: next((normalized[_norm(alias)] for alias in aliases if _norm(alias) in normalized), None)
                       for field, aliases in CONTRACTS[family].items()}
            mapping = {field: value for field, value in mapping.items() if value}
            if signature in mapping and {"nomenclature", "characteristic"}.issubset(mapping):
                candidates.append(family); partial[family] = mapping
        if len(candidates) == 1:
            source_type = candidates[0]; status = "mapping_required"
            missing = sorted(REQUIRED[source_type] - partial[source_type].keys())
            matches = partial
    return {"source_type": source_type, "status": status, "mapping_status": "ready" if status == "detected" else status,
            "required_missing": missing, "detected_columns": list(map(str, columns)),
            "matches": sorted(matches), "mapping": matches.get(source_type, {}),
            "diagnostic_code": diagnostic_code, "diagnostic_found": diagnostic_found,
            "diagnostic_missing": diagnostic_missing}


def read_excel_source(data: bytes, sheet: str | None = None) -> tuple[str, pd.DataFrame]:
    with pd.ExcelFile(BytesIO(data)) as workbook:
        selected = sheet or workbook.sheet_names[0]
        if selected not in workbook.sheet_names:
            raise ValueError("sheet_not_found")
        table = pd.read_excel(workbook, sheet_name=selected)
    table.columns = [str(column).strip() for column in table.columns]
    return selected, table.dropna(how="all")


def _canonical_record(raw: Mapping[str, Any], mapping: Mapping[str, str], provenance: Mapping[str, Any]) -> dict[str, Any]:
    def get(field: str) -> Any:
        return raw.get(mapping[field]) if field in mapping else None
    identity = build_canonical_sku_identity({"nomenclature": get("nomenclature"), "characteristic": get("characteristic")})
    record = {**provenance, "sku_key": identity["sku_key"], "nomenclature": identity["nomenclature"],
              "characteristic": identity["characteristic"], "identity_diagnostics": identity["diagnostics"]}
    for field in mapping:
        if field in {"nomenclature", "characteristic"}:
            continue
        value = get(field)
        if field == "snapshot_at":
            record[field] = normalize_operational_day(value)
        elif field == "occurred_at":
            record[field] = normalize_source_datetime(value)
        elif field in {"source_stock_quantity", "source_position_balance", "cell_picking_order", "actual_quantity",
                       "accounting_quantity", "box_quantity", "reported_pallets", "quantity", "source_pick_order",
                       "weight", "length", "width", "height", "boxes_per_layer", "layers_per_pallet", "quantity_per_box"}:
            record[field] = _number(value)
            record[field + "_raw"] = _json_value(value)
        else:
            record[field] = _json_value(value)
    # The mass RO query intentionally exports both the source line quantity and
    # calculated shipped boxes.  Keep the former as control evidence, never as
    # canonical factual quantity.
    if provenance["source_type"] == "outbound" and "Количество" in raw:
        record["source_line_quantity_raw"] = _json_value(raw.get("Количество"))
    if provenance["source_type"] == "vgh":
        a, b = record.get("boxes_per_layer"), record.get("layers_per_pallet")
        record["source_boxes_per_pallet"] = a * b if isinstance(a, (int, float)) and isinstance(b, (int, float)) else None
        record["palletization_authority"] = "source_layer_norm" if record["source_boxes_per_pallet"] else "unresolved"
    return record


def normalize_table(table: pd.DataFrame, *, dataset_id: str, filename: str, data_hash: str,
                    source_type: str, sheet: str, imported_at: str, mapping: Mapping[str, str] | None = None,
                    parser_version: str = PARSER_VERSION,
                    ) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    mapping = dict(mapping or detect_source_type(table.columns)["mapping"])
    raw_records, canonical = [], []
    for ordinal, (source_index, series) in enumerate(table.iterrows(), 2):
        provenance = {"dataset_id": dataset_id, "source_file_name": filename, "content_hash": data_hash,
                      "source_type": source_type, "parser_version": parser_version, "sheet": sheet,
                      "source_row": ordinal, "source_index": _json_value(source_index), "imported_at": imported_at}
        source_values = {str(key): value for key, value in series.items()}
        raw = {key: _json_value(value) for key, value in source_values.items()}
        raw_records.append({**provenance, "raw": raw})
        canonical.append(_canonical_record(source_values, mapping, provenance))
    return raw_records, canonical


def _day(record: Mapping[str, Any]) -> str | None:
    value = record.get("snapshot_at") or record.get("occurred_at")
    return normalize_operational_day(value)


def _dataset_operational_dates(dataset: Mapping[str, Any]) -> set[str]:
    """Return canonical days advertised by an imported dataset.

    Registry entries created by earlier parser versions can contain the same
    Excel day in several representations.  Treat the index as an import
    boundary too, rather than comparing its raw values with ISO day strings.
    """
    values = dataset.get("index", {}).get("dates", dataset.get("partitions", []))
    return {day for value in values if (day := normalize_operational_day(value))}


def _duplicates(raw_records: list[dict[str, Any]]) -> int:
    values = [json.dumps(row["raw"], ensure_ascii=False, sort_keys=True, separators=(",", ":")) for row in raw_records]
    return len(values) - len(set(values))


def validate_records(source_type: str, raw: list[dict[str, Any]], rows: list[dict[str, Any]],
                     geometry_cells: set[str] | None = None) -> dict[str, Any]:
    geometry_cells = geometry_cells or set()
    days = sorted({day for row in rows if (day := _day(row))})
    sku = {row["sku_key"] for row in rows if row.get("sku_key")}
    diagnostics: dict[str, Any] = {"rows": len(raw), "unique_sku": len(sku), "missing_sku": sum(not r.get("sku_key") for r in rows),
        "duplicate_raw_rows": _duplicates(raw), "period_from": days[0] if days else None, "period_to": days[-1] if days else None,
        "snapshot_count": len(days) if source_type == "historical_placement" else None,
        "zero_quantities": 0, "negative_quantities": 0, "missing_quantities": 0,
        "identity_collisions": find_canonical_identity_collisions(rows), "warnings": [], "errors": []}
    quantity_field = {"receipts": "box_quantity", "outbound": "quantity", "inventory": "actual_quantity"}.get(source_type)
    if quantity_field:
        values = [row.get(quantity_field) for row in rows]
        diagnostics.update(zero_quantities=sum(v == 0 for v in values), negative_quantities=sum(isinstance(v, (int, float)) and v < 0 for v in values), missing_quantities=sum(v is None for v in values))
    if source_type == "historical_placement":
        def grouped(field: str, value: str, *, across_snapshots: bool = False) -> int:
            groups = defaultdict(set)
            for row in rows:
                if row.get(field) not in (None, "") and row.get(value) not in (None, ""):
                    key = str(row[field]) if across_snapshots else (row.get("snapshot_at"), str(row[field]))
                    groups[key].add(str(row[value]))
            return sum(len(items) > 1 for items in groups.values())
        cells = {str(r.get("cell")) for r in rows if r.get("cell") not in (None, "")}
        orders = defaultdict(set)
        for row in rows:
            if row.get("cell") and row.get("cell_picking_order") is not None:
                orders[str(row["cell"])].add(row["cell_picking_order"])
        diagnostics.update(unique_cells=len(cells), sku_in_multiple_cells=grouped("sku_key", "cell"),
            multiple_sku_per_cell=grouped("cell", "sku_key"), multiple_sku_per_pallet=grouped("source_pallet_ref", "sku_key"),
            pallet_in_multiple_cells=grouped("source_pallet_ref", "cell", across_snapshots=True),
            # No authoritative resolver from the historical 1C address string
            # to geometry cell_key exists in the project.  Do not compare it
            # with ad-hoc display/code fields.
            unknown_geometry_cells=None,
            historical_cell_resolution="historical_cell_not_resolved_to_geometry",
            cell_picking_order_conflicts=sum(len(values) > 1 for values in orders.values()),
            cell_picking_order_evidence=[{"cell": cell, "picking_orders": sorted(values), "conflict": len(values) > 1} for cell, values in sorted(orders.items())])
        diagnostics["warnings"].append("historical_cell_not_resolved_to_geometry")
    for key, label in (("missing_sku", "missing_sku"), ("duplicate_raw_rows", "duplicate_raw_rows"),
                       ("identity_collisions", "identity_collisions"), ("negative_quantities", "negative_quantities"),
                       ("missing_quantities", "missing_quantities")):
        if diagnostics.get(key): diagnostics["warnings"].append(label)
    return diagnostics


def _known_july_placement_check(filename: str, rows: list[dict[str, Any]]) -> dict[str, Any] | None:
    """Apply the supplied fingerprint only to the explicitly known export."""
    if _norm(Path(filename).name) != _norm("размещение июль.xlsx"):
        return None
    expected = [(date(2026, 7, 1) + timedelta(days=offset)).isoformat() for offset in range(32)]
    counts = Counter(_day(row) for row in rows)
    missing = [day for day in expected if not counts.get(day)]
    return {"expected_snapshot_count": 32, "actual_snapshot_count": len({day for day in counts if day}),
            "period_expected": [expected[0], expected[-1]], "missing_dates": missing,
            "rows_2026_07_15": counts.get("2026-07-15", 0), "rows_2026_07_15_expected": 587,
            "snapshot_count_matches": len({day for day in counts if day}) == 32,
            "july_15_fingerprint_matches": counts.get("2026-07-15", 0) == 587}


def _atomic_json(path: Path, value: Any) -> None:
    _artifact_atomic_json(path, value)


def _write_jsonl(path: Path, rows: Iterable[Mapping[str, Any]]) -> None:
    _artifact_write_jsonl(path, rows, compresslevel=FACTUAL_GZIP_COMPRESSLEVEL)


def _iter_jsonl(path: Path) -> Iterable[dict[str, Any]]:
    yield from _artifact_iter_jsonl(path)


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    """Retain the legacy materialized-reader contract for its existing callers."""
    return _artifact_read_jsonl(path)


def load_registry(root: Path = DATA_ROOT) -> dict[str, Any]:
    path = root / "registry.json"
    if not path.exists(): return {"registry_version": 2, "datasets": [], "diagnostics": []}
    try:
        with measure("factual.load_registry"):
            state = json.loads(path.read_text(encoding="utf-8"))
            record_file_read("factual_registry", path.stat().st_size)
    except (OSError, json.JSONDecodeError):
        return {"registry_version": 1, "datasets": [], "warning": "registry_unreadable"}
    if not isinstance(state.get("datasets"), list):
        return {"registry_version": 2, "datasets": [], "warning": "registry_invalid", "diagnostics": []}
    # PR #161 registries did not distinguish immutable versions from active
    # source slots.  Migrate in memory deterministically and flag ambiguous
    # slots rather than allowing every version into business queries.
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in state["datasets"]:
        logical = item.get("logical_source_id") or logical_source_identity(
            item.get("source_type", UNKNOWN_SOURCE), item.get("source_file_name", ""), item.get("sheet", "Sheet1"))
        item.setdefault("logical_source_id", logical)
        item.setdefault("version", item.get("content_hash"))
        item.setdefault("superseded_by", None)
        item.setdefault("supersedes", None)
        groups[logical].append(item)
    diagnostics = list(state.get("diagnostics", []))
    for logical, versions in groups.items():
        if all("active" not in item for item in versions):
            ordered = sorted(versions, key=lambda item: (item.get("imported_at", ""), item.get("dataset_id", "")))
            for item in ordered: item["active"] = False
            ordered[-1]["active"] = True
            if len(ordered) > 1:
                diagnostics.append({"code": "registry_activation_review_required", "logical_source_id": logical})
    state.update(registry_version=2, diagnostics=diagnostics)
    return state


def logical_source_identity(source_type: str, filename: str, sheet: str) -> str:
    normalized_name = _norm(Path(filename).name)
    payload = "\x1f".join((source_type, normalized_name, _norm(sheet)))
    return "source:" + hashlib.sha256(payload.encode()).hexdigest()


def active_datasets(registry: Mapping[str, Any], source_type: str | None = None) -> list[dict[str, Any]]:
    """The single business-query boundary for lifecycle filtering."""
    return [item for item in registry.get("datasets", [])
            if item.get("active", True)
            and (not item.get("parser_version") or item.get("parser_version") == PARSER_VERSION)
            and (source_type is None or item.get("source_type") == source_type)]


def ensure_compact_scope_indexes(registry: dict[str, Any], *, source_types: Iterable[str] | None = None,
                                 root: Path = DATA_ROOT) -> dict[str, Any]:
    """Persist a one-time compact warehouse index for pre-index datasets.

    Old partitions are streamed rather than materialized. Once upgraded,
    normal UI reruns read registry metadata only.
    """
    changed = False; requested = set(source_types or {"outbound", "receipts", "inventory"})
    for dataset in registry.get("datasets", []):
        if dataset.get("source_type") not in requested: continue
        index = dataset.setdefault("index", {})
        if "warehouses" in index and "warehouses_by_date" in index: continue
        by_date: dict[str, list[str]] = {}; all_warehouses: set[str] = set()
        artifact = Path(str(dataset.get("artifact") or root / str(dataset.get("dataset_id", "")).removeprefix("dataset:")))
        for raw_day in index.get("dates", dataset.get("partitions", [])):
            day = normalize_operational_day(raw_day) or str(raw_day)
            path = artifact / "canonical" / f"date={raw_day}.jsonl.gz"
            warehouses = sorted({str(row.get("warehouse")) for row in _iter_jsonl(path) if row.get("warehouse")}) if path.exists() else []
            by_date[day] = warehouses; all_warehouses.update(warehouses)
            index.setdefault("daily", {}).setdefault(day, {})["warehouses"] = warehouses
        index["warehouses"] = sorted(all_warehouses); index["warehouses_by_date"] = by_date; changed = True
    if changed: _atomic_json(root / "registry.json", registry)
    return registry


def _import_excel_dataset_buffered(data: bytes, filename: str, *, sheet: str | None = None, root: Path = DATA_ROOT,
                         geometry_cells: Iterable[str] | None = None, reimport: bool = False,
                         parser_version: str = PARSER_VERSION) -> dict[str, Any]:
    # Hash and consult content provenance before opening the workbook.  With no
    # explicit sheet, a prior default-sheet import is safe to reuse; an
    # explicitly selected sheet must match exactly.
    digest = content_hash(data)
    root.mkdir(parents=True, exist_ok=True)
    registry = load_registry(root)
    obsolete_content_match = any(item.get("content_hash") == digest
        and item.get("parser_version") != parser_version
        and (sheet is None or item.get("sheet") == sheet) for item in registry["datasets"])
    content_match = next((item for item in registry["datasets"]
        if item.get("content_hash") == digest and item.get("parser_version") == parser_version
        and (sheet is None or item.get("sheet") == sheet)), None)
    if content_match and not reimport:
        return {**content_match, "reused": True, "duplicate_filename": filename if filename != content_match.get("source_file_name") else None}
    selected, table = read_excel_source(data, sheet)
    detection = detect_source_type(table.columns)
    if detection["source_type"] == UNKNOWN_SOURCE or detection["status"] != "detected":
        return {"status": detection["status"], "source_type": UNKNOWN_SOURCE, "source_label": SOURCE_LABELS[UNKNOWN_SOURCE],
                "detected_source_family": detection["source_type"] if detection["source_type"] != UNKNOWN_SOURCE else None,
                "required_missing": detection["required_missing"], "detected_columns": detection["detected_columns"],
                "matches": detection["matches"], "diagnostic_code": detection["diagnostic_code"],
                "diagnostic_found": detection["diagnostic_found"], "diagnostic_missing": detection["diagnostic_missing"],
                "errors": [detection["status"]]}
    source_type = detection["source_type"]
    dataset_id = dataset_identity(digest, source_type, selected, parser_version)
    existing = next((item for item in registry["datasets"] if item.get("dataset_id") == dataset_id), None)
    if existing and not reimport:
        return {**existing, "reused": True}
    imported_at = datetime.now(timezone.utc).isoformat(timespec="microseconds")
    logical_source_id = logical_source_identity(source_type, filename, selected)
    raw, canonical = normalize_table(table, dataset_id=dataset_id, filename=filename, data_hash=digest,
        source_type=source_type, sheet=selected, imported_at=imported_at, mapping=detection["mapping"], parser_version=parser_version)
    diagnostics = validate_records(source_type, raw, canonical, set(map(str, geometry_cells or [])))
    known_check = _known_july_placement_check(filename, canonical) if source_type == "historical_placement" else None
    if known_check:
        diagnostics["known_july_validation"] = known_check
        if known_check["missing_dates"]: diagnostics["warnings"].append("known_july_missing_dates")
        if not known_check["july_15_fingerprint_matches"]: diagnostics["warnings"].append("known_july_15_fingerprint_mismatch")
    artifact = root / dataset_id.removeprefix("dataset:")
    staging = Path(tempfile.mkdtemp(dir=root, prefix=".staging-"))
    _write_jsonl(staging / "raw.jsonl.gz", raw)
    partitions = defaultdict(list)
    for row in canonical: partitions[_day(row) or "undated"].append(row)
    daily: dict[str, Any] = {}
    sku_index = sorted({row.get("sku_key") for row in canonical if row.get("sku_key")})
    document_keys: set[str] = set()
    for day, rows in partitions.items():
        _write_jsonl(staging / "canonical" / f"date={day}.jsonl.gz", rows)
        docs = {(r.get("document_ref") or r.get("inventory_ref"), r.get("document_number") or r.get("inventory_number"), r.get("occurred_at")) for r in rows}
        document_keys.update(json.dumps(key, ensure_ascii=False, default=str) for key in docs)
        positive = [r.get("quantity") for r in rows if isinstance(r.get("quantity"), (int, float)) and r.get("quantity") > 0]
        daily[day] = {"rows": len(rows), "sku": len({r.get("sku_key") for r in rows if r.get("sku_key")}),
            "cells": len({r.get("cell") for r in rows if r.get("cell")}), "documents": len(docs),
            "positive_quantity": sum(positive), "positive_sku_keys": sorted({r.get("sku_key") for r in rows if r.get("sku_key") and isinstance(r.get("quantity"), (int, float)) and r.get("quantity") > 0})}
    index = {"sku_keys": sku_index, "dates": sorted(partitions), "daily": daily,
             "warehouses": sorted({str(row.get("warehouse")) for row in canonical if row.get("warehouse")}),
             "warehouses_by_date": {day: sorted({str(row.get("warehouse")) for row in rows if row.get("warehouse")})
                                    for day, rows in partitions.items()},
             "document_keys": sorted(document_keys)}
    metadata = {"dataset_id": dataset_id, "source_file_name": filename, "content_hash": digest, "source_type": source_type,
        "source_label": SOURCE_LABELS[source_type], "parser_version": parser_version, "sheet": selected, "rows": len(raw),
        "period_from": diagnostics["period_from"], "period_to": diagnostics["period_to"], "unique_sku": diagnostics["unique_sku"],
        "imported_at": imported_at, "status": "ready_with_warnings" if diagnostics["warnings"] else "ready", "errors": diagnostics["errors"],
        "warnings": diagnostics["warnings"], "detected_columns": detection["detected_columns"],
        "mapping": detection["mapping"], "mapping_status": "ready", "artifact": str(artifact), "partitions": sorted(partitions),
        "index": index, "diagnostics": diagnostics, "logical_source_id": logical_source_id, "version": digest,
        "active": True, "superseded_by": None, "supersedes": None}
    previous = [item for item in registry["datasets"] if item.get("active", True) and item.get("logical_source_id") == logical_source_id and item.get("dataset_id") != dataset_id]
    metadata["supersedes"] = previous[-1]["dataset_id"] if previous else None
    _atomic_json(staging / "metadata.json", metadata)
    if artifact.exists():
        # A reparse of the same immutable identity already has a complete artifact.
        import shutil; shutil.rmtree(staging)
    else:
        os.replace(staging, artifact)
    for item in previous:
        item["active"] = False
        item["superseded_by"] = dataset_id
    registry["datasets"] = [item for item in registry["datasets"] if item.get("dataset_id") != dataset_id] + [metadata]
    # Different logical files may coexist, but overlapping dates/business keys
    # are surfaced rather than silently summed without a warning.
    overlaps = []
    for item in active_datasets(registry, source_type):
        if item["dataset_id"] == dataset_id: continue
        dates = set(item.get("index", {}).get("dates", item.get("partitions", []))) & set(index["dates"])
        keys = set(item.get("index", {}).get("document_keys", [])) & set(index["document_keys"])
        if dates or keys: overlaps.append({"dataset_id": item["dataset_id"], "dates": sorted(dates), "duplicate_document_keys": len(keys)})
    if overlaps:
        metadata["diagnostics"]["overlapping_active_sources"] = overlaps
        metadata["warnings"].append("overlapping_active_sources")
    registry["datasets"].sort(key=lambda item: (item.get("imported_at", ""), item["dataset_id"]))
    registry["registry_version"] = 2
    _atomic_json(root / "registry.json", registry)
    return {**metadata, "reused": False, "reparsed_for_parser_upgrade": obsolete_content_match}


BUSINESS_KEYS = {
    "outbound": ("document_ref", "line_number"),
    "receipts": ("document_ref", "line_number"),
    "inventory": ("inventory_ref", "line_number"),
    "vgh": ("sku_key",),
}
MATERIAL_FIELDS = {
    "outbound": ("occurred_at", "warehouse", "sku_key", "quantity", "source_pick_order"),
    "receipts": ("occurred_at", "warehouse", "sku_key", "box_quantity", "reported_pallets",
                 "terminal_completed", "expected_receipt"),
    "inventory": ("occurred_at", "warehouse", "sku_key", "actual_quantity", "accounting_quantity"),
    "vgh": ("boxes_per_layer", "layers_per_pallet", "quantity_per_box", "source_boxes_per_pallet"),
}


def _fingerprint(values: Any) -> str:
    encoded = json.dumps(values, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str)
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _business_evidence(row: Mapping[str, Any], source_type: str) -> dict[str, Any] | None:
    fields = BUSINESS_KEYS.get(source_type)
    if not fields:
        return None
    key_values = [row.get(field) for field in fields]
    if any(value in (None, "") for value in key_values):
        return None
    payload = {field: row.get(field) for field in MATERIAL_FIELDS[source_type]}
    return {"business_key": json.dumps(key_values, ensure_ascii=False, separators=(",", ":"), default=str),
            "payload_fingerprint": _fingerprint(payload), "payload": payload,
            "day": _day(row), "dataset_id": row.get("dataset_id"),
            "source_file_name": row.get("source_file_name"), "source_row": row.get("source_row")}


def _stream_xlsx_dataset(data: bytes, filename: str, *, sheet: str | None, root: Path,
                         geometry_cells: Iterable[str] | None, parser_version: str,
                         fail_after_rows: int | None = None,
                         progress_callback: ImportProgressCallback | None = None) -> dict[str, Any]:
    """Import XLSX with one canonical and one RAW record resident at a time.

    Memory is bounded by the uploaded XLSX bytes, openpyxl's read-only parser,
    exact SHA-256 row fingerprints, SKU/business-key indexes, and per-day
    aggregate sets.  It never creates a pandas table or complete row lists.
    """
    from openpyxl import load_workbook

    import_started = time.perf_counter()
    stage_seconds: dict[str, float] = defaultdict(float)
    digest = content_hash(data)
    root.mkdir(parents=True, exist_ok=True)
    registry = load_registry(root)
    obsolete_content_match = any(item.get("content_hash") == digest
        and item.get("parser_version") != parser_version
        and (sheet is None or item.get("sheet") == sheet) for item in registry["datasets"])
    content_match = next((item for item in registry["datasets"]
        if item.get("content_hash") == digest and item.get("parser_version") == parser_version
        and (sheet is None or item.get("sheet") == sheet)), None)
    if content_match:
        state = "existing_active_version" if content_match.get("active", True) else "existing_superseded_version"
        return {**content_match, "reused": True, "reuse_state": state, "status": state,
                "duplicate_filename": filename if filename != content_match.get("source_file_name") else None}

    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        selected = sheet or workbook.sheetnames[0]
        if selected not in workbook.sheetnames:
            raise ValueError("sheet_not_found")
        worksheet = workbook[selected]
        iterator = worksheet.iter_rows(values_only=True)
        # ``max_row`` comes from the worksheet dimensions already read by
        # openpyxl.  It is only a progress hint (some producers overstate the
        # used range), but unlike counting rows it does not require a second
        # pass through a large XLSX.
        dimension_rows = worksheet.max_row
        total_rows = max(0, dimension_rows - 1) if isinstance(dimension_rows, int) else None
        read_started = time.perf_counter()
        headers = [str(value).strip() if value is not None else "" for value in next(iterator, ())]
        stage_seconds["xlsx_read"] += time.perf_counter() - read_started
        detection = detect_source_type(headers)
        if detection["source_type"] == UNKNOWN_SOURCE or detection["status"] != "detected":
            return {"status": detection["status"], "source_type": UNKNOWN_SOURCE,
                    "source_label": SOURCE_LABELS[UNKNOWN_SOURCE],
                    "detected_source_family": detection["source_type"] if detection["source_type"] != UNKNOWN_SOURCE else None,
                    "required_missing": detection["required_missing"], "detected_columns": headers,
                    "matches": detection["matches"], "diagnostic_code": detection["diagnostic_code"],
                    "diagnostic_found": detection["diagnostic_found"],
                    "diagnostic_missing": detection["diagnostic_missing"], "errors": [detection["status"]]}
        source_type = detection["source_type"]
        header_positions = tuple((index, header) for index, header in enumerate(headers) if header)
        dataset_id = dataset_identity(digest, source_type, selected, parser_version)
        imported_at = datetime.now(timezone.utc).isoformat(timespec="microseconds")
        logical_source_id = logical_source_identity(source_type, filename, selected)
        staging = Path(tempfile.mkdtemp(dir=root, prefix=".staging-"))
        writers: dict[str, Any] = {}
        raw_writer = index_writer = None
        row_count = missing_sku = duplicate_raw = zero = negative = missing_qty = 0
        raw_hashes: set[str] = set(); sku_keys: set[str] = set(); days: set[str] = set(); warehouses: set[str] = set()
        daily: dict[str, dict[str, Any]] = defaultdict(lambda: {"rows": 0, "sku_keys": set(), "cells": set(),
            "documents": set(), "positive_quantity": 0, "positive_sku_keys": set()})
        placement_sku_cells: dict[tuple[Any, Any], set[str]] = defaultdict(set)
        placement_cell_sku: dict[tuple[Any, Any], set[str]] = defaultdict(set)
        placement_pallet_sku: dict[tuple[Any, Any], set[str]] = defaultdict(set)
        placement_pallet_cells: dict[str, set[str]] = defaultdict(set)
        cell_orders: dict[str, set[Any]] = defaultdict(set)
        business_key_count = 0
        worksheet_row = 1
        last_progress_at = import_started
        last_progress_row = 0

        def report_progress(stage: str, *, force: bool = False) -> None:
            nonlocal last_progress_at, last_progress_row
            if progress_callback is None:
                return
            now = time.perf_counter()
            if not force and row_count - last_progress_row < IMPORT_PROGRESS_ROW_INTERVAL \
                    and now - last_progress_at < IMPORT_PROGRESS_TIME_INTERVAL_SECONDS:
                return
            elapsed = now - import_started
            event = {"processed_rows": row_count, "elapsed_seconds": elapsed,
                "rows_per_second": row_count / elapsed if elapsed else 0.0, "stage": stage,
                "total_rows": total_rows, "filename": filename}
            try:
                progress_callback(event)
            except Exception:
                # Presentation failures must not corrupt or roll back a valid
                # data import, especially after atomic publication.
                pass
            last_progress_at, last_progress_row = now, row_count

        report_progress("starting", force=True)
        try:
            raw_writer = gzip.open(staging / "raw.jsonl.gz", "wt", encoding="utf-8",
                                   compresslevel=FACTUAL_GZIP_COMPRESSLEVEL)
            index_writer = gzip.open(staging / "business_index.jsonl.gz", "wt", encoding="utf-8",
                                     compresslevel=FACTUAL_GZIP_COMPRESSLEVEL)
            while True:
                read_started = time.perf_counter()
                try:
                    values = next(iterator)
                except StopIteration:
                    stage_seconds["xlsx_read"] += time.perf_counter() - read_started
                    break
                worksheet_row += 1
                stage_seconds["xlsx_read"] += time.perf_counter() - read_started
                source_row = worksheet_row
                if not any(value is not None for value in values):
                    continue
                if fail_after_rows is not None and row_count >= fail_after_rows:
                    raise RuntimeError("injected_streaming_import_failure")
                started = time.perf_counter()
                source_values = {header: values[index] for index, header in header_positions}
                raw_values = {header: _json_value(value) for header, value in source_values.items()}
                provenance = {"dataset_id": dataset_id, "source_file_name": filename, "content_hash": digest,
                    "source_type": source_type, "parser_version": parser_version, "sheet": selected,
                    "source_row": source_row, "source_index": source_row - 2, "imported_at": imported_at}
                raw_record = {**provenance, "raw": raw_values}
                stage_seconds["raw_prepare"] += time.perf_counter() - started
                started = time.perf_counter()
                raw_json = json.dumps(raw_values, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
                stage_seconds["raw_serialization"] += time.perf_counter() - started
                started = time.perf_counter()
                raw_fp = hashlib.sha256(raw_json.encode()).hexdigest()
                duplicate_raw += raw_fp in raw_hashes; raw_hashes.add(raw_fp)
                stage_seconds["raw_hash"] += time.perf_counter() - started
                started = time.perf_counter()
                raw_record_json = json.dumps(raw_record, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
                stage_seconds["raw_serialization"] += time.perf_counter() - started
                started = time.perf_counter(); raw_writer.write(raw_record_json + "\n")
                stage_seconds["raw_write"] += time.perf_counter() - started
                started = time.perf_counter()
                record = _canonical_record(source_values, detection["mapping"], provenance)
                stage_seconds["canonical"] += time.perf_counter() - started
                day = _day(record) or "undated"; days.add(day)
                if day not in writers:
                    target = staging / "canonical" / f"date={day}.jsonl.gz"; target.parent.mkdir(parents=True, exist_ok=True)
                    writers[day] = gzip.open(target, "wt", encoding="utf-8",
                                             compresslevel=FACTUAL_GZIP_COMPRESSLEVEL)
                started = time.perf_counter()
                canonical_json = json.dumps(record, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
                stage_seconds["canonical_serialization"] += time.perf_counter() - started
                started = time.perf_counter(); writers[day].write(canonical_json + "\n")
                stage_seconds["canonical_write"] += time.perf_counter() - started
                started = time.perf_counter()
                sku = record.get("sku_key"); missing_sku += not bool(sku)
                if sku: sku_keys.add(sku); daily[day]["sku_keys"].add(sku)
                if record.get("warehouse") not in (None, ""):
                    warehouse = str(record["warehouse"]); warehouses.add(warehouse)
                    daily[day].setdefault("warehouses", set()).add(warehouse)
                daily[day]["rows"] += 1
                if record.get("cell") not in (None, ""): daily[day]["cells"].add(str(record["cell"]))
                doc = (record.get("document_ref") or record.get("inventory_ref"),
                       record.get("document_number") or record.get("inventory_number"), record.get("occurred_at"))
                if any(doc): daily[day]["documents"].add(json.dumps(doc, ensure_ascii=False, default=str))
                qty_field = {"receipts": "box_quantity", "outbound": "quantity", "inventory": "actual_quantity"}.get(source_type)
                if qty_field:
                    qty = record.get(qty_field); zero += qty == 0; negative += isinstance(qty, (int, float)) and qty < 0; missing_qty += qty is None
                    if qty_field == "quantity" and isinstance(qty, (int, float)) and qty > 0:
                        daily[day]["positive_quantity"] += qty
                        if sku: daily[day]["positive_sku_keys"].add(sku)
                stage_seconds["index_maintenance"] += time.perf_counter() - started
                started = time.perf_counter(); evidence = _business_evidence(record, source_type)
                stage_seconds["business_evidence"] += time.perf_counter() - started
                if evidence:
                    started = time.perf_counter()
                    evidence_json = json.dumps(evidence, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
                    stage_seconds["business_serialization"] += time.perf_counter() - started
                    started = time.perf_counter(); index_writer.write(evidence_json + "\n")
                    stage_seconds["business_write"] += time.perf_counter() - started
                    business_key_count += 1
                started = time.perf_counter()
                if source_type == "historical_placement":
                    snap, cell, pallet = record.get("snapshot_at"), str(record.get("cell") or ""), str(record.get("source_pallet_ref") or "")
                    if sku and cell: placement_sku_cells[(snap, sku)].add(cell); placement_cell_sku[(snap, cell)].add(sku)
                    if pallet and sku: placement_pallet_sku[(snap, pallet)].add(sku)
                    if pallet and cell: placement_pallet_cells[pallet].add(cell)
                    if cell and record.get("cell_picking_order") is not None: cell_orders[cell].add(record["cell_picking_order"])
                row_count += 1
                stage_seconds["index_maintenance"] += time.perf_counter() - started
                report_progress("reading_and_normalizing")
        except Exception:
            for stream in (raw_writer, index_writer, *writers.values()):
                if stream: stream.close()
            shutil.rmtree(staging, ignore_errors=True)
            raise
        finally:
            for stream in (raw_writer, index_writer, *writers.values()):
                if stream and not stream.closed: stream.close()
    finally:
        workbook.close()

    report_progress("finalizing", force=True)
    finalization_started = time.perf_counter()

    real_days = sorted(day for day in days if day != "undated")
    warnings = [name for value, name in ((missing_sku, "missing_sku"), (duplicate_raw, "duplicate_raw_rows"),
        (negative, "negative_quantities"), (missing_qty, "missing_quantities")) if value]
    diagnostics: dict[str, Any] = {"rows": row_count, "unique_sku": len(sku_keys), "missing_sku": missing_sku,
        "duplicate_raw_rows": duplicate_raw, "period_from": real_days[0] if real_days else None,
        "period_to": real_days[-1] if real_days else None, "snapshot_count": len(real_days) if source_type == "historical_placement" else None,
        "zero_quantities": zero, "negative_quantities": negative, "missing_quantities": missing_qty,
        "identity_collisions": [], "warnings": warnings, "errors": [], "streaming_import": True,
        "memory_bound": "xlsx bytes + openpyxl read-only window + exact row hashes + compact unique-key indexes"}
    if source_type == "historical_placement":
        diagnostics.update(unique_cells=len({cell for _, cell in placement_cell_sku}),
            sku_in_multiple_cells=sum(len(x) > 1 for x in placement_sku_cells.values()),
            multiple_sku_per_cell=sum(len(x) > 1 for x in placement_cell_sku.values()),
            multiple_sku_per_pallet=sum(len(x) > 1 for x in placement_pallet_sku.values()),
            pallet_in_multiple_cells=sum(len(x) > 1 for x in placement_pallet_cells.values()), unknown_geometry_cells=None,
            historical_cell_resolution="historical_cell_not_resolved_to_geometry",
            cell_picking_order_conflicts=sum(len(x) > 1 for x in cell_orders.values()),
            cell_picking_order_evidence=[{"cell": c, "picking_orders": sorted(v), "conflict": len(v) > 1} for c, v in sorted(cell_orders.items())])
        diagnostics["warnings"].append("historical_cell_not_resolved_to_geometry")
    index_daily = {day: {"rows": values["rows"], "sku": len(values["sku_keys"]), "cells": len(values["cells"]),
        "documents": len(values["documents"]), "positive_quantity": values["positive_quantity"],
        "positive_sku_keys": sorted(values["positive_sku_keys"]),
        "warehouses": sorted(values.get("warehouses", set()))} for day, values in daily.items()}
    index = {"sku_keys": sorted(sku_keys), "dates": sorted(days), "daily": index_daily,
             "warehouses": sorted(warehouses),
             "warehouses_by_date": {day: values["warehouses"] for day, values in index_daily.items()},
             "business_key_count": business_key_count, "business_index_artifact": "business_index.jsonl.gz"}
    stage_seconds["final_metadata_publication"] += time.perf_counter() - finalization_started
    elapsed = time.perf_counter() - import_started
    performance = {"rows": row_count, "elapsed_seconds": elapsed,
        "rows_per_second": row_count / elapsed if elapsed else 0.0,
        "stage_seconds": {name: round(seconds, 6) for name, seconds in sorted(stage_seconds.items())},
        "gzip_compresslevel": FACTUAL_GZIP_COMPRESSLEVEL}
    diagnostics["import_performance"] = performance
    artifact = root / dataset_id.removeprefix("dataset:")
    previous = [item for item in registry["datasets"] if item.get("active", True) and item.get("logical_source_id") == logical_source_id]
    metadata = {"dataset_id": dataset_id, "source_file_name": filename, "content_hash": digest,
        "source_type": source_type, "source_label": SOURCE_LABELS[source_type], "parser_version": parser_version,
        "sheet": selected, "rows": row_count, "period_from": diagnostics["period_from"], "period_to": diagnostics["period_to"],
        "unique_sku": len(sku_keys), "imported_at": imported_at, "status": "ready_with_warnings" if diagnostics["warnings"] else "ready",
        "errors": [], "warnings": diagnostics["warnings"], "detected_columns": headers, "mapping": detection["mapping"],
        "mapping_status": "ready", "artifact": str(artifact), "partitions": sorted(days), "index": index,
        "diagnostics": diagnostics, "logical_source_id": logical_source_id, "version": digest, "active": True,
        "superseded_by": None, "supersedes": previous[-1]["dataset_id"] if previous else None}
    _atomic_json(staging / "metadata.json", metadata)
    if artifact.exists(): shutil.rmtree(staging)
    else: os.replace(staging, artifact)
    for item in previous: item["active"] = False; item["superseded_by"] = dataset_id
    registry["datasets"] = [item for item in registry["datasets"] if item.get("dataset_id") != dataset_id] + [metadata]
    registry["datasets"].sort(key=lambda item: (item.get("imported_at", ""), item["dataset_id"]))
    registry["registry_version"] = 3
    _atomic_json(root / "registry.json", registry)
    elapsed = time.perf_counter() - import_started
    performance["elapsed_seconds"] = elapsed
    performance["rows_per_second"] = row_count / elapsed if elapsed else 0.0
    report_progress("completed", force=True)
    return {**metadata, "diagnostics": {**diagnostics, "import_performance": performance},
            "reused": False, "reparsed_for_parser_upgrade": obsolete_content_match}


def import_excel_dataset(data: bytes, filename: str, *, sheet: str | None = None, root: Path = DATA_ROOT,
                         geometry_cells: Iterable[str] | None = None, reimport: bool = False,
                         parser_version: str = PARSER_VERSION, _fail_after_rows: int | None = None,
                         progress_callback: ImportProgressCallback | None = None) -> dict[str, Any]:
    """Hash first, then use bounded-memory XLSX import; retain legacy XLS compatibility."""
    if Path(filename).suffix.casefold() == ".xlsx" and not reimport:
        return _stream_xlsx_dataset(data, filename, sheet=sheet, root=root, geometry_cells=geometry_cells,
                                    parser_version=parser_version, fail_after_rows=_fail_after_rows,
                                    progress_callback=progress_callback)
    result = _import_excel_dataset_buffered(data, filename, sheet=sheet, root=root, geometry_cells=geometry_cells,
                                            reimport=reimport, parser_version=parser_version)
    if Path(filename).suffix.casefold() == ".xls" and len(data) > 20_000_000:
        result.setdefault("warnings", []).append("large_xls_uses_buffered_import")
    return result


def load_dataset_rows(dataset: Mapping[str, Any], day: str | None = None, *, raw: bool = False) -> list[dict[str, Any]]:
    artifact = Path(str(dataset["artifact"]))
    if raw: return _read_jsonl(artifact / "raw.jsonl.gz")
    if day is not None: return _read_jsonl(artifact / "canonical" / f"date={day}.jsonl.gz")
    rows = []
    for partition in dataset.get("partitions", []): rows.extend(_read_jsonl(artifact / "canonical" / f"date={partition}.jsonl.gz"))
    return rows


def positive_outbound(rows: Iterable[Mapping[str, Any]]) -> list[dict[str, Any]]:
    return [dict(row) for row in rows if isinstance(row.get("quantity"), (int, float)) and row["quantity"] > 0]


@profiled("factual.date_summary")
def date_summary(registry: Mapping[str, Any], day: str) -> dict[str, Any]:
    result = {"operational_day": day, "placement": {"snapshot_exists": False, "rows": 0, "sku": 0, "cells": 0},
              "inventory": {"documents": 0, "rows": 0}, "receipts": {"documents": 0, "rows": 0, "accepted_boxes": None},
              "outbound": {"documents": 0, "lines": 0, "positive_demand": 0}, "vgh": {"relevant_sku": 0, "covered_sku": 0},
              "next_day_placement_available": False}
    indexes: dict[str, list[Mapping[str, Any]]] = defaultdict(list)
    for dataset in active_datasets(registry):
        daily = dataset.get("index", {}).get("daily", {}).get(day)
        if daily: indexes[dataset["source_type"]].append(daily)
    def total(source: str, field: str) -> float:
        return sum(float(item.get(field) or 0) for item in indexes[source])
    result["placement"] = {"snapshot_exists": bool(indexes["historical_placement"]),
        "rows": int(total("historical_placement", "rows")), "sku": int(total("historical_placement", "sku")),
        "cells": int(total("historical_placement", "cells"))}
    for source, target in (("inventory", "inventory"), ("receipts", "receipts")):
        result[target].update(documents=int(total(source, "documents")), rows=int(total(source, "rows")))
    demanded = {sku for item in indexes["outbound"] for sku in item.get("positive_sku_keys", [])}
    result["outbound"] = {"documents": int(total("outbound", "documents")), "lines": int(total("outbound", "rows")),
                          "positive_demand": total("outbound", "positive_quantity")}
    vgh = set()
    for dataset in active_datasets(registry):
        if dataset.get("source_type") == "vgh": vgh.update(dataset.get("index", {}).get("sku_keys", []))
        if dataset.get("source_type") == "historical_placement":
            tomorrow = (pd.Timestamp(day).date() + timedelta(days=1)).isoformat()
            result["next_day_placement_available"] |= tomorrow in dataset.get("index", {}).get("dates", dataset.get("partitions", []))
    result["vgh"] = {"relevant_sku": len(demanded), "covered_sku": len(demanded & vgh)}
    return result


@profiled("factual.cross_source_coverage")
def cross_source_coverage(registry: Mapping[str, Any]) -> list[dict[str, Any]]:
    indexes: dict[str, set[str]] = defaultdict(set)
    for dataset in active_datasets(registry):
        indexes[dataset["source_type"]].update(dataset.get("index", {}).get("sku_keys", []))
    scope = indexes["outbound"] | indexes["historical_placement"]
    return [{"sku_key": sku, "outbound": sku in indexes["outbound"], "placement": sku in indexes["historical_placement"],
             "vgh": sku in indexes["vgh"], "inventory": sku in indexes["inventory"], "receipts": sku in indexes["receipts"]} for sku in sorted(scope)]


def activate_dataset_version(dataset_id: str, *, root: Path = DATA_ROOT) -> dict[str, Any]:
    """Explicitly activate an immutable version without reparsing its workbook."""
    registry = load_registry(root)
    selected = next((item for item in registry["datasets"] if item.get("dataset_id") == dataset_id), None)
    if selected is None:
        raise KeyError("dataset_not_found")
    logical = selected.get("logical_source_id")
    current = [item for item in registry["datasets"] if item.get("active", True)
               and item.get("logical_source_id") == logical and item is not selected]
    predecessor = current[-1] if current else None
    for item in current:
        item["active"] = False
        item["superseded_by"] = dataset_id
    selected["active"] = True
    selected["superseded_by"] = None
    selected["supersedes"] = predecessor.get("dataset_id") if predecessor else selected.get("supersedes")
    registry["diagnostics"] = [d for d in registry.get("diagnostics", [])
        if not (d.get("code") == "registry_activation_review_required" and d.get("logical_source_id") == logical)]
    registry["readiness_invalidated_at"] = datetime.now(timezone.utc).isoformat(timespec="microseconds")
    _atomic_json(root / "registry.json", registry)
    return dict(selected)


def _source_conflicts(registry: Mapping[str, Any], source_type: str, day: str | None = None) -> dict[str, Any]:
    groups: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    for dataset in active_datasets(registry, source_type):
        path = Path(str(dataset["artifact"])) / "business_index.jsonl.gz"
        if path.exists():
            evidence_rows = _read_jsonl(path)
        else:
            evidence_rows = [e for row in load_dataset_rows(dataset, day) if (e := _business_evidence(row, source_type))]
        for evidence in evidence_rows:
            if day is None or evidence.get("day") == day or source_type == "vgh":
                groups[evidence["business_key"]][evidence["payload_fingerprint"]].append(evidence)
    duplicates, conflicts = [], []
    for key, payloads in sorted(groups.items()):
        occurrences = [item for values in payloads.values() for item in values]
        if len(payloads) > 1:
            conflicts.append({"code": "conflicting_factual_business_key", "source_type": source_type,
                "business_key": json.loads(key), "occurrences": occurrences})
        elif len(occurrences) > 1:
            duplicates.append({"code": "duplicate_factual_evidence_collapsed", "source_type": source_type,
                "business_key": json.loads(key), "occurrences": occurrences})
    return {"duplicates": duplicates, "conflicts": conflicts, "groups": groups}


@profiled("factual.readiness_compact_conflicts")
def _readiness_source_conflicts(registry: Mapping[str, Any], source_type: str,
                                *, preview_limit: int = 10, occurrence_limit: int = 10) -> dict[str, Any]:
    """Find exact source conflicts using disk storage and bounded Python memory."""
    fd, database_name = tempfile.mkstemp(prefix=f"factual-{source_type}-", suffix=".sqlite3")
    os.close(fd)
    connection: sqlite3.Connection | None = None
    try:
        connection = sqlite3.connect(database_name)
        connection.execute("PRAGMA journal_mode=OFF")
        connection.execute("PRAGMA synchronous=OFF")
        connection.execute("CREATE TABLE evidence (business_key TEXT, fingerprint TEXT, evidence TEXT)")
        batch: list[tuple[str, str, str]] = []

        def add(evidence: Mapping[str, Any]) -> None:
            batch.append((str(evidence["business_key"]), str(evidence["payload_fingerprint"]),
                          json.dumps(evidence, ensure_ascii=False, separators=(",", ":"), default=str)))
            if len(batch) >= 1000:
                connection.executemany("INSERT INTO evidence VALUES (?, ?, ?)", batch)
                batch.clear()

        for dataset in active_datasets(registry, source_type):
            artifact = Path(str(dataset["artifact"]))
            business_index = artifact / "business_index.jsonl.gz"
            if business_index.exists():
                with measure(f"factual.readiness_business_index_scan.{source_type}"):
                    for evidence in _iter_jsonl(business_index):
                        add(evidence)
            else:
                # Legacy artifacts are deliberately visited one partition at a time.
                for day in dataset.get("partitions", dataset.get("index", {}).get("dates", [])):
                    partition = artifact / "canonical" / f"date={day}.jsonl.gz"
                    for row in _iter_jsonl(partition):
                        evidence = _business_evidence(row, source_type)
                        if evidence is not None:
                            add(evidence)
        if batch:
            connection.executemany("INSERT INTO evidence VALUES (?, ?, ?)", batch)
        connection.execute("CREATE INDEX evidence_key_fp ON evidence (business_key, fingerprint)")
        connection.commit()
        conflict_count = connection.execute(
            "SELECT COUNT(*) FROM (SELECT business_key FROM evidence GROUP BY business_key HAVING COUNT(DISTINCT fingerprint) > 1)"
        ).fetchone()[0]
        duplicate_count = connection.execute(
            "SELECT COUNT(*) FROM (SELECT business_key FROM evidence GROUP BY business_key "
            "HAVING COUNT(DISTINCT fingerprint) = 1 AND COUNT(*) > 1)"
        ).fetchone()[0]
        keys = [row[0] for row in connection.execute(
            "SELECT business_key FROM evidence GROUP BY business_key "
            "HAVING COUNT(DISTINCT fingerprint) > 1 ORDER BY business_key LIMIT ?", (preview_limit,))]
        preview = []
        for key in keys:
            occurrences = [json.loads(row[0]) for row in connection.execute(
                "SELECT evidence FROM evidence WHERE business_key = ? ORDER BY fingerprint, rowid LIMIT ?",
                (key, occurrence_limit))]
            preview.append({"code": "conflicting_factual_business_key", "source_type": source_type,
                            "business_key": json.loads(key), "occurrences": occurrences})
        return {"source_type": source_type, "conflict_count": conflict_count,
                "duplicate_count": duplicate_count, "preview": preview}
    finally:
        if connection is not None:
            connection.close()
        try:
            os.unlink(database_name)
        except FileNotFoundError:
            pass


def build_data_contract_diagnostics(registry: Mapping[str, Any], *, root: Path = DATA_ROOT,
                                    preview_limit: int = 10) -> dict[str, Any]:
    """Build read-only diagnostics from registry metadata and compact indexes."""
    datasets = [dict(item) for item in registry.get("datasets", [])]
    active = [item for item in datasets if item.get("active", True)]
    analyses = {source: _readiness_source_conflicts(registry, source, preview_limit=preview_limit)
                for source in BUSINESS_KEYS}
    overlaps: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for index, left in enumerate(active):
        for right in active[index + 1:]:
            if left.get("source_type") != right.get("source_type"):
                continue
            dates = sorted(_dataset_operational_dates(left) & _dataset_operational_dates(right))
            keys = set(left.get("index", {}).get("document_keys", [])) & set(right.get("index", {}).get("document_keys", []))
            if not dates and not keys:
                continue
            for current, other in ((left, right), (right, left)):
                overlaps[str(current.get("dataset_id"))].append({
                    "dataset_id": other.get("dataset_id"), "source_filename": other.get("source_file_name"),
                    "period_from": other.get("period_from"), "period_to": other.get("period_to"),
                    "overlapping_dates": dates, "overlapping_document_keys": len(keys)})
    rows = []
    for item in datasets:
        source = str(item.get("source_type", UNKNOWN_SOURCE))
        analysis = analyses.get(source, {"duplicate_count": 0, "conflict_count": 0, "preview": []})
        compatible = not item.get("parser_version") or item.get("parser_version") == PARSER_VERSION
        warnings = list(item.get("warnings", [])); errors = list(item.get("errors", []))
        if item.get("active", True) and not compatible: errors.append("parser_reimport_required")
        if overlaps.get(str(item.get("dataset_id"))): warnings.append("overlapping_active_sources")
        rows.append({"source_type": source, "source_filename": item.get("source_file_name"),
            "sheet": item.get("sheet"), "parser_version": item.get("parser_version"),
            "current_parser_version": PARSER_VERSION, "parser_compatible": compatible,
            "dataset_id": item.get("dataset_id"), "logical_source_id": item.get("logical_source_id"),
            "active": bool(item.get("active", True)), "superseded_by": item.get("superseded_by"),
            "detected_source_columns": item.get("detected_columns", []),
            "canonical_mapping": item.get("mapping", {}), "rows": int(item.get("rows", 0) or 0),
            "period_from": item.get("period_from"), "period_to": item.get("period_to"),
            "unique_sku": int(item.get("unique_sku", 0) or 0),
            "documents": len(item.get("index", {}).get("document_keys", [])),
            "duplicate_business_keys": analysis["duplicate_count"] if item.get("active", True) else 0,
            "conflicting_business_keys": analysis["conflict_count"] if item.get("active", True) else 0,
            "conflict_preview": analysis["preview"] if item.get("active", True) else [],
            "overlapping_active_sources": overlaps.get(str(item.get("dataset_id")), []),
            "mapping_status": item.get("mapping_status", "unknown"),
            "warnings": sorted(set(warnings)), "errors": sorted(set(errors))})
    return {"parser_version": PARSER_VERSION, "datasets": rows,
            "requires_reimport": [row["dataset_id"] for row in rows if row["active"] and not row["parser_compatible"]],
            "read_only": True}


@profiled("factual.load_effective_rows")
def load_effective_rows(source_type: str, day: str | None = None, *, registry: Mapping[str, Any] | None = None,
                        root: Path = DATA_ROOT, strict: bool = True) -> dict[str, Any]:
    """Authoritative replay boundary: active evidence in, deduplicated factual rows out."""
    registry = registry or load_registry(root)
    if source_type == "historical_placement":
        candidates = [dataset for dataset in active_datasets(registry, source_type)
                      if day is None or normalize_operational_day(day) in _dataset_operational_dates(dataset)]
        if day is not None and len(candidates) > 1:
            conflict = {"code": "multiple_active_placement_sources_for_snapshot", "day": day,
                        "dataset_ids": [item["dataset_id"] for item in candidates]}
            if strict: raise ValueError("multiple_active_placement_sources_for_snapshot")
            return {"rows": [], "duplicates": [], "conflicts": [conflict], "authoritative": False}
        rows = [row for dataset in candidates for row in load_dataset_rows(dataset, day)]
        return {"rows": rows, "duplicates": [], "conflicts": [], "authoritative": len(candidates) <= 1}
    analysis = _source_conflicts(registry, source_type, day)
    if analysis["conflicts"] and strict:
        raise ValueError("conflicting_factual_business_key")
    representatives: dict[str, dict[str, Any]] = {}
    for dataset in sorted(active_datasets(registry, source_type), key=lambda x: x["dataset_id"]):
        for row in load_dataset_rows(dataset, day):
            evidence = _business_evidence(row, source_type)
            if evidence is None:
                representatives[f"unkeyed:{dataset['dataset_id']}:{row.get('source_row')}"] = dict(row)
                continue
            key = evidence["business_key"]
            if key in representatives: continue
            occurrences = [x for values in analysis["groups"].get(key, {}).values() for x in values]
            enriched = dict(row)
            enriched["duplicate_source_count"] = len(occurrences)
            enriched["duplicate_source_ids"] = sorted({str(x.get("dataset_id")) for x in occurrences})
            enriched["duplicate_source_filenames"] = sorted({str(x.get("source_file_name")) for x in occurrences})
            enriched["duplicate_row_evidence"] = [{"dataset_id": x.get("dataset_id"), "source_file_name": x.get("source_file_name"),
                                                     "source_row": x.get("source_row")} for x in occurrences]
            representatives[key] = enriched
    return {"rows": list(representatives.values()), "duplicates": analysis["duplicates"],
            "conflicts": analysis["conflicts"], "authoritative": not analysis["conflicts"]}


def geometry_model_signature(model: Mapping[str, Any]) -> str:
    cells = sorted(str(cell.get("cell_key") or "") for cell in model.get("cells", []) if isinstance(cell, Mapping))
    return _fingerprint({"model_id": model.get("model_id"), "cells": cells})


def load_cell_mappings(root: Path = DATA_ROOT) -> dict[str, Any]:
    path = root / "historical_cell_mappings.json"
    if not path.exists(): return {"version": 1, "mappings": []}
    try: return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError): return {"version": 1, "mappings": [], "warning": "mapping_registry_unreadable"}


def save_historical_cell_mapping(source_cell: Any, geometry_cell_key: str, model: Mapping[str, Any], *,
                                 root: Path = DATA_ROOT, source_evidence: Any = None) -> dict[str, Any]:
    """Persist only an explicit user choice, scoped to exact geometry identity."""
    source = str(source_cell).replace("\u00a0", " ").strip()
    targets = [cell for cell in model.get("cells", []) if str(cell.get("cell_key") or "") == str(geometry_cell_key)]
    if len(targets) != 1: raise ValueError("unknown_or_ambiguous_geometry_target")
    state = load_cell_mappings(root); signature = geometry_model_signature(model)
    record = {"model_id": model.get("model_id"), "model_signature": signature, "source_cell": source,
        "geometry_cell_key": geometry_cell_key, "mapping_method": "user_confirmed",
        "created_at": datetime.now(timezone.utc).isoformat(timespec="microseconds"),
        "source_evidence": source_evidence, "validation_status": "valid"}
    state["mappings"] = [item for item in state["mappings"]
        if not (item.get("model_signature") == signature and item.get("source_cell") == source)] + [record]
    _atomic_json(root / "historical_cell_mappings.json", state)
    return record


def resolve_historical_cell(source_cell: Any, model: Mapping[str, Any], *, root: Path = DATA_ROOT,
                            mappings: Mapping[str, Any] | None = None) -> dict[str, Any]:
    """Resolve through proven source_cell identity or explicit confirmation; never guess."""
    source = str(source_cell).replace("\u00a0", " ").strip()
    signature = geometry_model_signature(model)
    cells = [cell for cell in model.get("cells", []) if isinstance(cell, Mapping)]
    valid_keys = {str(cell.get("cell_key")) for cell in cells if cell.get("cell_key") not in (None, "")}
    state = mappings or load_cell_mappings(root)
    persisted = next((item for item in state.get("mappings", []) if item.get("source_cell") == source
                      and item.get("model_signature") == signature), None)
    if persisted:
        target = str(persisted.get("geometry_cell_key") or "")
        if target in valid_keys: return {**persisted, "resolution_status": "resolved"}
        return {**persisted, "geometry_cell_key": None, "resolution_status": "stale", "diagnostic": "historical_cell_unresolved"}
    # source_cell is the only model field accepted as shared authoritative identity.
    exact = [str(cell.get("cell_key")) for cell in cells
             if str(cell.get("source_cell") or "").replace("\u00a0", " ").strip() == source]
    if len(exact) == 1:
        return {"source_cell": source, "geometry_cell_key": exact[0], "mapping_method": "exact_authoritative",
                "model_signature": signature, "resolution_status": "resolved"}
    if len(exact) > 1:
        return {"source_cell": source, "geometry_cell_key": None, "candidates": sorted(exact),
                "resolution_status": "ambiguous", "diagnostic": "historical_cell_ambiguous"}
    return {"source_cell": source, "geometry_cell_key": None, "resolution_status": "unresolved",
            "diagnostic": "historical_cell_unresolved"}


@profiled("factual.load_effective_placement")
def load_effective_placement(day: str, model: Mapping[str, Any], *, registry: Mapping[str, Any] | None = None,
                             root: Path = DATA_ROOT, strict: bool = True) -> dict[str, Any]:
    view = load_effective_rows("historical_placement", day, registry=registry, root=root, strict=strict)
    resolved = []
    for row in view["rows"]:
        resolution = resolve_historical_cell(row.get("cell"), model, root=root)
        resolved.append({**row, "source_cell": row.get("cell"), "resolved_geometry_cell_key": resolution.get("geometry_cell_key"),
                         "cell_resolution_status": resolution["resolution_status"], "cell_resolution": resolution})
    return {**view, "rows": resolved}


def build_fact_route_readiness(placement_rows: Iterable[Mapping[str, Any]], demanded_sku_keys: Iterable[str],
                               usable_cell_keys: Iterable[str]) -> dict[str, Any]:
    rows = list(placement_rows); demanded = set(demanded_sku_keys); usable = set(usable_cell_keys)
    cells = {str(r.get("source_cell") or r.get("cell")) for r in rows if r.get("source_cell") or r.get("cell")}
    demand_rows = [r for r in rows if r.get("sku_key") in demanded]
    demand_cells = {str(r.get("source_cell") or r.get("cell")) for r in demand_rows}
    resolved = {str(r.get("source_cell") or r.get("cell")) for r in rows if r.get("resolved_geometry_cell_key")}
    demand_resolved = {str(r.get("source_cell") or r.get("cell")) for r in demand_rows
                       if r.get("resolved_geometry_cell_key") in usable}
    ambiguous = {str(r.get("source_cell") or r.get("cell")) for r in rows if r.get("cell_resolution_status") == "ambiguous"}
    demand_unresolved = demand_cells - demand_resolved
    return {"placement_rows": len(rows), "unique_factual_cells": len(cells), "resolved_cells": len(resolved),
        "unresolved_cells": len(cells - resolved), "ambiguous_cells": len(ambiguous),
        "overall_cell_coverage": {"resolved": len(resolved), "total": len(cells)},
        "demand_relevant_cell_coverage": {"resolved": len(demand_resolved), "total": len(demand_cells)},
        "demand_relevant_unresolved_cells": sorted(demand_unresolved),
        "resolved_row_sku_coverage": {"resolved": sum(bool(r.get("resolved_geometry_cell_key")) for r in rows), "total": len(rows)},
        "cells_without_usable_access_node": sorted({str(r.get("resolved_geometry_cell_key")) for r in demand_rows
            if r.get("resolved_geometry_cell_key") and r.get("resolved_geometry_cell_key") not in usable}),
        "fact_route_ready": demand_cells <= demand_resolved}


@profiled("factual.build_monthly_data_readiness")
def build_monthly_data_readiness(registry: Mapping[str, Any], model: Mapping[str, Any], period_from: str,
                                 period_to: str, *, root: Path = DATA_ROOT,
                                 usable_cell_keys: Iterable[str] | None = None,
                                 receipts_required: bool = True) -> dict[str, Any]:
    """Pure, deterministic factual authority contract; it performs no simulation."""
    start, end = date.fromisoformat(period_from), date.fromisoformat(period_to)
    required_days = [(start + timedelta(days=i)).isoformat() for i in range((end - start).days + 2)]
    blockers: list[dict[str, Any]] = []; warnings: list[dict[str, Any]] = []
    registry_blockers = [d for d in registry.get("diagnostics", []) if d.get("code") == "registry_activation_review_required"]
    if registry_blockers: blockers.append({"code": "registry_activation_review_required", "message": "Требуется подтвердить активную версию источника."})
    obsolete_active = [item for item in registry.get("datasets", []) if item.get("active", True)
                       and item.get("parser_version") and item.get("parser_version") != PARSER_VERSION]
    if obsolete_active:
        blockers.append({"code": "parser_reimport_required",
            "message": "Сохранённые factual artifacts созданы несовместимой версией parser; повторно импортируйте исходные Excel.",
            "datasets": [{"dataset_id": item.get("dataset_id"), "source_filename": item.get("source_file_name"),
                          "parser_version": item.get("parser_version")} for item in obsolete_active]})
    placement_datasets = active_datasets(registry, "historical_placement")
    placement_dates_by_dataset = {str(dataset.get("dataset_id")): _dataset_operational_dates(dataset)
                                  for dataset in placement_datasets}
    placement_sources = {day: [dataset for dataset in placement_datasets
        if day in placement_dates_by_dataset[str(dataset.get("dataset_id"))]] for day in required_days}
    missing = [day for day, sources in placement_sources.items() if not sources]
    overlaps = [day for day, sources in placement_sources.items() if len(sources) > 1]
    imported_placement_dates = sorted({day for dates in placement_dates_by_dataset.values() for day in dates})
    extra_placement_dates = sorted(set(imported_placement_dates) - set(required_days))
    if missing: blockers.append({"code": "missing_placement_snapshot", "dates": missing})
    if overlaps: blockers.append({"code": "multiple_active_placement_sources_for_snapshot", "dates": overlaps})
    analyses = {source: _readiness_source_conflicts(registry, source)
                for source in ("outbound", "receipts", "inventory", "vgh")}
    for source, analysis in analyses.items():
        if analysis["conflict_count"]:
            issue = {"code": "conflicting_factual_business_key", "source_type": source,
                     "count": analysis["conflict_count"], "preview": analysis["preview"]}
            (warnings if source == "vgh" else blockers).append(issue)
    demanded_sku_by_day: dict[str, set[str]] = {day: set() for day in required_days[:-1]}
    positive_demand = False
    with measure("factual.readiness_demand_index"):
        for dataset in active_datasets(registry, "outbound"):
            daily = dataset.get("index", {}).get("daily", {})
            has_compact_demand_index = any("positive_sku_keys" in counts for counts in daily.values())
            for day in required_days[:-1]:
                counts = daily.get(day, {})
                sku_keys = counts.get("positive_sku_keys", []) if has_compact_demand_index else None
                if sku_keys is None:
                    # Parser-vintage fallback: stream only this day's partition.
                    sku_keys = set()
                    partition = Path(str(dataset["artifact"])) / "canonical" / f"date={day}.jsonl.gz"
                    for row in _iter_jsonl(partition):
                        if (isinstance(row.get("quantity"), (int, float)) and row["quantity"] > 0
                                and row.get("sku_key")):
                            sku_keys.add(row["sku_key"])
                    positive_demand |= bool(sku_keys)
                else:
                    positive_demand |= bool(sku_keys) or float(counts.get("positive_quantity", 0) or 0) > 0
                demanded_sku_by_day[day].update(str(sku) for sku in sku_keys if sku)
    demanded = set().union(*demanded_sku_by_day.values()) if demanded_sku_by_day else set()
    vgh_keys = {sku for dataset in active_datasets(registry, "vgh") for sku in dataset.get("index", {}).get("sku_keys", [])}
    missing_vgh = sorted(demanded - vgh_keys)
    if missing_vgh:
        warnings.append({"code": "missing_vgh_for_demanded_sku", "count": len(missing_vgh),
                         "preview": missing_vgh[:10]})
    if not positive_demand: blockers.append({"code": "outbound_not_available"})
    receipt_dates = {day for dataset in active_datasets(registry, "receipts") for day in dataset.get("index", {}).get("dates", [])}
    if receipts_required and not (set(required_days[:-1]) & receipt_dates): blockers.append({"code": "receipts_not_available"})
    route_checks = []
    usable = set(usable_cell_keys or [str(c.get("cell_key")) for c in model.get("cells", []) if c.get("cell_key")])
    with measure("factual.readiness_daily_cell_resolution"):
        if not overlaps:
            for day in required_days[:-1]:
                placement = load_effective_placement(day, model, registry=registry, root=root, strict=False)
                route_checks.append(build_fact_route_readiness(placement["rows"], demanded_sku_by_day[day], usable))
    unresolved_demand = sum(x["demand_relevant_cell_coverage"]["total"] - x["demand_relevant_cell_coverage"]["resolved"] for x in route_checks)
    unresolved_demand_cells = sorted({cell for check in route_checks
                                      for cell in check.get("demand_relevant_unresolved_cells", [])})
    if unresolved_demand:
        blockers.append({"code": "historical_cell_unresolved",
                         "demand_relevant_cells": unresolved_demand,
                         "unique_source_cells": len(unresolved_demand_cells),
                         "source_cell_preview": unresolved_demand_cells[:20]})
    active_signature = _fingerprint(sorted(d["dataset_id"] for d in active_datasets(registry)))
    mapping_signature = _fingerprint(load_cell_mappings(root))
    source_conflict_count = sum(x["conflict_count"] for x in analyses.values()) + len(overlaps)
    placement_ready = not missing and not overlaps
    outbound_ready = positive_demand and not analyses["outbound"]["conflict_count"]
    receipts_ready = (not receipts_required or bool(set(required_days[:-1]) & receipt_dates)) and not analyses["receipts"]["conflict_count"]
    vgh_ready = not missing_vgh and not analyses["vgh"]["conflict_count"]
    cell_ready = not unresolved_demand and bool(route_checks)
    registry_ready = not registry_blockers and not obsolete_active
    period_days = set(required_days[:-1])

    def source_counts(source_type: str) -> tuple[int, int]:
        datasets = active_datasets(registry, source_type)
        rows = documents = 0
        for dataset in datasets:
            daily = dataset.get("index", {}).get("daily", {})
            for day, counts in daily.items():
                if day in period_days:
                    rows += int(counts.get("rows", 0) or 0)
                    documents += int(counts.get("documents", 0) or 0)
        return rows, documents

    outbound_count, outbound_documents = source_counts("outbound")
    receipt_rows, receipt_documents = source_counts("receipts")
    inventory_rows, inventory_documents = source_counts("inventory")
    required_sku = len(demanded)
    covered_sku = len(demanded & vgh_keys)
    checks = [
        {"name": "placement_snapshot", "status": "pass" if placement_ready else "fail",
         "title": "Срезы размещения", "expected_days": len(required_days),
         "available_days": len(required_days) - len(missing),
         "imported_days": len(imported_placement_dates), "missing_dates": missing,
         "extra_dates": extra_placement_dates,
         "expected_dates": required_days, "detected_dates": imported_placement_dates,
         "details": f"{len(required_days) - len(missing)} / {len(required_days)} дней"},
        {"name": "outbound", "status": "pass" if outbound_ready else "fail", "title": "Расходные ордера",
         "available": positive_demand, "rows": outbound_count, "documents": outbound_documents,
         "details": f"{outbound_count} строк · {outbound_documents} документов" if positive_demand else "отсутствуют"},
        {"name": "receipts", "status": "pass" if receipts_ready else "fail", "title": "Приходы",
         "available": bool(period_days & receipt_dates), "rows": receipt_rows, "documents": receipt_documents,
         "details": f"{receipt_rows} строк · {receipt_documents} документов" if period_days & receipt_dates else "отсутствуют"},
        {"name": "vgh_coverage", "status": "pass" if vgh_ready else "warning", "title": "ВГХ",
         "covered_sku": covered_sku, "total_required_sku": required_sku,
         "percentage": round(100 * covered_sku / required_sku, 1) if required_sku else 100.0,
         "missing_sku_count": len(missing_vgh), "details": f"{covered_sku} / {required_sku} SKU"},
        {"name": "inventory", "status": "pass" if inventory_rows else "info", "title": "Инвентаризация",
         "available": bool(inventory_rows), "rows": inventory_rows, "documents": inventory_documents,
         "details": f"{inventory_rows} строк · {inventory_documents} документов" if inventory_rows else "отсутствует (не блокирует replay)"},
        {"name": "parser_compatibility", "status": "fail" if obsolete_active else "pass",
         "title": "Версия parser", "datasets": len(obsolete_active),
         "details": "требуется reimport/reparse исходных Excel" if obsolete_active else PARSER_VERSION},
    ]
    return {"monthly_replay_ready": not blockers, "period_from": period_from, "period_to": period_to,
        "control_endpoint": required_days[-1], "placement_ready": placement_ready, "outbound_ready": outbound_ready,
        "receipts_ready": receipts_ready, "vgh_ready": vgh_ready, "cell_resolution_ready": cell_ready,
        "registry_ready": registry_ready, "hard_blockers": blockers, "warnings": warnings,
        "coverage": {"placement_checkpoints": len(required_days) - len(missing), "placement_checkpoints_required": len(required_days),
                     "demanded_sku": len(demanded), "vgh_covered_sku": len(demanded & vgh_keys),
                     "demand_relevant_unresolved_cells": unresolved_demand, "source_conflicts": source_conflict_count},
        "diagnostics": {"ready": not blockers, "checks": checks},
        "active_dataset_signature": active_signature, "cell_mapping_signature": mapping_signature}
